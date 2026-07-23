"""
routers/torneos.py
Router FastAPI para el módulo completo de Gestión de Torneos.
Cubre: equipos, jugadores, planilla, goles, tarjetas, posiciones, sanciones, W.O.
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from typing import Optional, List
from pydantic import BaseModel, Field
from datetime import datetime, date
import uuid
import json
import os
import shutil

from database import get_session
from services.facial_recognition import FacialRecognitionService
import time
from security import get_password_hash
from email_service import email_service
import secrets
import string

# Import schemas from schemas_regiones
from schemas_regiones import RegionCreate, CiudadCreate, PlayoffRegionalCreate

router = APIRouter(prefix="/cancha/torneos", tags=["Torneos"])

def generate_random_password(length: int = 10) -> str:
    characters = string.ascii_letters + string.digits
    return ''.join(secrets.choice(characters) for _ in range(length))


# ============================================================
# SCHEMAS PYDANTIC
# ============================================================

class CategoriaCreate(BaseModel):
    nombre: str
    formato: str = "liga"  # liga | eliminacion_simple | mixto | suizo
    max_equipos: int = 16
    costo_inscripcion: float = 0.0
    pts_victoria: int = 3
    pts_empate: int = 1
    pts_derrota: int = 0
    configuracion: Optional[dict] = {}

class EventoCreate(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    deporte: str = "Fútbol 5"
    fecha_inicio: str
    fecha_fin: Optional[str] = None
    complejo_id: Optional[str] = None
    organizador_id: Optional[int] = None
    reglas: Optional[list[str]] = []
    premios: Optional[list[dict]] = []
    categorias: List[CategoriaCreate] = []

class OrganizadorCreate(BaseModel):
    usuario_id: int
    nombre: str
    plan: Optional[str] = "basico"
    max_torneos: Optional[int] = 3

class EquipoCreate(BaseModel):
    nombre: str
    capitan_nombre: Optional[str] = None
    capitan_telefono: Optional[str] = None
    capitan_email: Optional[str] = None
    logo_url: Optional[str] = None
    color_principal: Optional[str] = None
    color_secundario: Optional[str] = None
    promocion: int = 0

class EquipoUpdate(BaseModel):
    nombre: Optional[str] = None
    capitan_nombre: Optional[str] = None
    capitan_telefono: Optional[str] = None
    capitan_email: Optional[str] = None
    inscripcion_confirmada: Optional[bool] = None

class JugadorCreate(BaseModel):
    nombre: str
    dni: str
    email: Optional[str] = None
    telefono: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    numero_camiseta: Optional[int] = None
    posicion: Optional[str] = None
    foto_url: Optional[str] = None
    egreso_ano: Optional[int] = None
    es_exalumno: bool = True
    documento_firmado_url: Optional[str] = None
    cedula_anverso_url: Optional[str] = None
    cedula_reverso_url: Optional[str] = None
    modalidad: Optional[str] = None
    genero: Optional[str] = None
    peso: Optional[float] = None
    estatura: Optional[float] = None

class JugadorUpdate(BaseModel):
    nombre: Optional[str] = None
    dni: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    numero_camiseta: Optional[int] = None
    posicion: Optional[str] = None
    foto_url: Optional[str] = None
    estado: Optional[str] = None
    egreso_ano: Optional[int] = None
    es_exalumno: Optional[bool] = None
    modalidad: Optional[str] = None
    genero: Optional[str] = None
    peso: Optional[float] = None
    estatura: Optional[float] = None
    documento_firmado_url: Optional[str] = None
    cedula_anverso_url: Optional[str] = None
    cedula_reverso_url: Optional[str] = None

class EquipoTecnicoCreate(BaseModel):
    nombre: str
    dni: Optional[str] = None
    rol: Optional[str] = "Entrenador"
    foto_url: Optional[str] = None

class EquipoTecnicoUpdate(BaseModel):
    nombre: Optional[str] = None
    dni: Optional[str] = None
    rol: Optional[str] = None
    foto_url: Optional[str] = None

class PartidoUpdate(BaseModel):
    goles_local: Optional[int] = None
    goles_visitante: Optional[int] = None
    estado: Optional[str] = None
    estadisticas: Optional[dict] = None
    observaciones: Optional[str] = None
    fecha_hora: Optional[str] = None
    cancha: Optional[str] = None
    equipo_local_id: Optional[str] = None
    equipo_visitante_id: Optional[str] = None

class GolCreate(BaseModel):
    player_id: Optional[str] = None
    equipo_id: str
    minuto: Optional[int] = None
    tipo: str = "normal"    # normal | penal | autogol
    tipo_evento_id: Optional[int] = None  # FK a cancha.tipos_evento (opcional, retrocompat.)

class TarjetaCreate(BaseModel):
    player_id: str
    equipo_id: str
    minuto: Optional[int] = None
    tipo: str               # amarilla | roja_directa | roja_segunda
    tipo_evento_id: Optional[int] = None  # FK a cancha.tipos_evento (opcional, retrocompat.)

class SustitucionCreate(BaseModel):
    """Registro de una sustitución: jugador que entra (player_id) y el que sale (player_out_id)."""
    player_id: str          # Jugador que ENTRA
    player_out_id: str      # Jugador que SALE
    equipo_id: str
    minuto: Optional[int] = None
    observaciones: Optional[str] = None

class WORequest(BaseModel):
    equipo_infractor_id: str
    marcador_ganador: int = 2
    marcador_perdedor: int = 0


# ============================================================
# HELPERS
# ============================================================

def _row_to_dict(keys: list, row) -> dict:
    return {k: (str(v) if isinstance(v, uuid.UUID) else
                v.isoformat() if isinstance(v, (datetime, date)) else v)
            for k, v in zip(keys, row)}


async def _recalcular_posiciones(torneo_id: str, session: AsyncSession):
    """Recalcula la tabla de posiciones completa para un torneo."""
    # Obtener configuración de puntos del torneo
    cfg = await session.execute(
        text("SELECT puntos_victoria, puntos_empate, puntos_derrota FROM torneos.torneos WHERE id = :id"),
        {"id": torneo_id}
    )
    cfg_row = cfg.fetchone()
    if not cfg_row:
        return
    pts_v, pts_e, pts_d = cfg_row[0] or 3, cfg_row[1] or 1, cfg_row[2] or 0

    # Obtener todos los equipos del torneo
    eq_res = await session.execute(
        text("SELECT id FROM torneos.equipos WHERE torneo_id = :tid AND estado_inscripcion = 'confirmado'"),
        {"tid": torneo_id}
    )
    equipos = [str(r[0]) for r in eq_res.fetchall()]

    # Obtener partidos finalizados
    p_res = await session.execute(
        text("""
            SELECT equipo_local_id, equipo_visitante_id, goles_local, goles_visitante
            FROM torneos.partidos
            WHERE torneo_id = :tid AND estado IN ('finalizado', 'wo')
        """),
        {"tid": torneo_id}
    )
    partidos = p_res.fetchall()

    # Calcular stats por equipo
    stats: dict = {eid: {"pj": 0, "pg": 0, "pe": 0, "pp": 0, "gf": 0, "gc": 0, "pts": 0, "opponents": []} for eid in equipos}

    for local_id, visitante_id, gl, gv in partidos:
        lid = str(local_id)
        vid = str(visitante_id)
        gl = gl or 0
        gv = gv or 0
        if lid in stats:
            stats[lid]["pj"] += 1
            stats[lid]["gf"] += gl
            stats[lid]["gc"] += gv
        if vid in stats:
            stats[vid]["pj"] += 1
            stats[vid]["gf"] += gv
            stats[vid]["gc"] += gl

        if gl > gv:
            if lid in stats: 
                stats[lid]["pg"] += 1; stats[lid]["pts"] += pts_v
                stats[lid]["opponents"].append((vid, 'W'))
            if vid in stats: 
                stats[vid]["pp"] += 1; stats[vid]["pts"] += pts_d
                stats[vid]["opponents"].append((lid, 'L'))
        elif gv > gl:
            if vid in stats: 
                stats[vid]["pg"] += 1; stats[vid]["pts"] += pts_v
                stats[vid]["opponents"].append((lid, 'W'))
            if lid in stats: 
                stats[lid]["pp"] += 1; stats[lid]["pts"] += pts_d
                stats[lid]["opponents"].append((vid, 'L'))
        else:
            if lid in stats: 
                stats[lid]["pe"] += 1; stats[lid]["pts"] += pts_e
                stats[lid]["opponents"].append((vid, 'D'))
            if vid in stats: 
                stats[vid]["pe"] += 1; stats[vid]["pts"] += pts_e
                stats[vid]["opponents"].append((lid, 'D'))

    # Calculate Buchholz and Sonneborn-Berger
    for eid in equipos:
        bh = 0
        sb = 0.0
        for opp_id, result in stats[eid]["opponents"]:
            if opp_id in stats:
                opp_pts = stats[opp_id]["pts"]
                bh += opp_pts
                if result == 'W':
                    sb += opp_pts
                elif result == 'D':
                    sb += opp_pts * 0.5
        stats[eid]["bh"] = bh
        stats[eid]["sb"] = sb

    # Obtener fair play por equipo
    fp_res = await session.execute(
        text("""
            SELECT tt.equipo_id, SUM(tt.pts_fair_play)
            FROM torneos.tarjetas tt
            JOIN torneos.partidos tp ON tt.partido_id = tp.id
            WHERE tp.torneo_id = :tid
            GROUP BY tt.equipo_id
        """),
        {"tid": torneo_id}
    )
    fp_map = {str(r[0]): int(r[1]) for r in fp_res.fetchall()}

    # Ordenar: pts DESC, bh DESC, sb DESC, dg DESC, gf DESC, fair_play ASC
    sorted_equipos = sorted(
        equipos,
        key=lambda e: (
            -stats[e]["pts"],
            -stats[e]["bh"],
            -stats[e]["sb"],
            -(stats[e]["gf"] - stats[e]["gc"]),
            -stats[e]["gf"],
            fp_map.get(e, 0)
        )
    )

    # UPSERT posiciones
    for pos, eid in enumerate(sorted_equipos, start=1):
        s = stats[eid]
        dg = s["gf"] - s["gc"]
        fp = fp_map.get(eid, 0)
        await session.execute(text("""
            INSERT INTO torneos.posiciones
                (id, torneo_id, torneo_equipo_id, posicion, pj, pg, pe, pp, gf, gc, dg, pts, pts_fair_play_neg, actualizado_en)
            VALUES
                (:uuid, :tid, :eid, :pos, :pj, :pg, :pe, :pp, :gf, :gc, :dg, :pts, :fp, CURRENT_TIMESTAMP)
            ON CONFLICT (torneo_id, torneo_equipo_id)
            DO UPDATE SET
                posicion=:pos, pj=:pj, pg=:pg, pe=:pe, pp=:pp,
                gf=:gf, gc=:gc, dg=:dg, pts=:pts, pts_fair_play_neg=:fp, actualizado_en=CURRENT_TIMESTAMP
        """), {"uuid": str(uuid.uuid4()), "tid": torneo_id, "eid": eid, "pos": pos,
               "pj": s["pj"], "pg": s["pg"], "pe": s["pe"], "pp": s["pp"],
               "gf": s["gf"], "gc": s["gc"], "dg": dg, "pts": s["pts"], "fp": fp})


async def _aplicar_tarjeta_logica(player_id: str, tipo: str, torneo_id: str, session: AsyncSession) -> dict:
    """Calcula pts fair play, detecta si genera suspensión y la aplica."""
    pts_fp = {"amarilla": 1, "roja_segunda": 3, "roja_directa": 4}.get(tipo, 0)
    genera = False
    partidos_susp = 0

    # Contar amarillas acumuladas del jugador en el torneo
    if tipo == "amarilla":
        am_res = await session.execute(text("""
            SELECT COUNT(*) FROM torneos.tarjetas tt
            JOIN torneos.partidos tp ON tt.partido_id = tp.id
            WHERE tp.torneo_id = :tid AND tt.player_id = :pid AND tt.tipo = 'amarilla'
        """), {"tid": torneo_id, "pid": player_id})
        total_amarillas = (am_res.scalar() or 0) + 1  # +1 la actual
        if total_amarillas % 3 == 0:  # Cada 3 amarillas = 1 fecha
            genera = True
            partidos_susp = 1

    elif tipo in ("roja_directa", "roja_segunda"):
        genera = True
        partidos_susp = 1 if tipo == "roja_segunda" else 2

    # Actualizar acumulados en tournament_players
    if tipo == "amarilla":
        await session.execute(text("""
            UPDATE cancha.tournament_players
            SET amarillas_acum = amarillas_acum + 1,
                estado = CASE WHEN amarillas_acum + 1 >= 3 THEN 'suspendido' ELSE estado END,
                actualizado_en = NOW()
            WHERE id = :pid
        """), {"pid": player_id})
    elif tipo in ("roja_directa", "roja_segunda"):
        await session.execute(text("""
            UPDATE cancha.tournament_players
            SET rojas_acum = rojas_acum + 1,
                estado = 'suspendido',
                actualizado_en = NOW()
            WHERE id = :pid
        """), {"pid": player_id})

    return {"pts_fair_play": pts_fp, "genera_suspension": genera, "partidos_suspension": partidos_susp}


# ============================================================
# ENDPOINTS — EVENTOS Y TORNEOS
# ============================================================

@router.get("/eventos", summary="Listar eventos")
async def get_eventos(
    complejo_id: Optional[str] = None,
    estado: Optional[str] = None,
    session: AsyncSession = Depends(get_session)
):
    sql = """
        SELECT e.id, e.complejo_id, e.nombre, e.descripcion, e.fecha_inicio, e.fecha_fin, e.estado,
               c.nombre AS complejo_nombre,
               (SELECT json_agg(json_build_object('id', t.id, 'categoria', t.categoria, 'formato', t.formato)) 
                FROM torneos.torneos t WHERE t.evento_id = e.id) as categorias
        FROM torneos.eventos e
        JOIN cancha.complejos c ON e.complejo_id = c.id
        WHERE 1=1
    """
    params: dict = {}
    if complejo_id:
        sql += " AND e.complejo_id = :complejo_id"
        params["complejo_id"] = complejo_id
    if estado:
        sql += " AND e.estado = :estado"
        params["estado"] = estado
    sql += " ORDER BY e.fecha_inicio DESC"

    result = await session.execute(text(sql), params)
    rows = result.fetchall()
    
    events = []
    for r in rows:
        d = _row_to_dict(["id","complejo_id","nombre","descripcion","fecha_inicio","fecha_fin","estado","complejo_nombre","categorias"], r)
        # Parse JSON if needed
        cats = d.get("categorias")
        if isinstance(cats, str):
            try:
                d["categorias"] = json.loads(cats)
            except:
                d["categorias"] = []
        events.append(d)
    return events

@router.get("", summary="Listar torneos")
async def get_torneos(
    complejo_id: Optional[str] = None,
    organizador_id: Optional[int] = None,
    estado: Optional[str] = None,
    deporte: Optional[str] = None,
    session: AsyncSession = Depends(get_session)
):
    sql = """
        SELECT t.id, t.complejo_id, t.organizador_id, t.nombre, t.descripcion, t.deporte,
               t.fecha_inicio, t.fecha_fin, t.estado,
               c.nombre AS complejo_nombre, org.nombre AS organizador_nombre, t.formato,
               t.max_equipos, t.costo_inscripcion,
               t.evento_id, t.categoria,
               (SELECT COUNT(*) FROM torneos.equipos te
                WHERE te.torneo_id = t.id AND te.estado_inscripcion = 'confirmado') AS equipos_confirmados
        FROM torneos.torneos t
        LEFT JOIN cancha.complejos c ON t.complejo_id = c.id
        LEFT JOIN cancha.organizadores org ON t.organizador_id = org.id
        WHERE 1=1
    """
    params: dict = {}
    if complejo_id:
        sql += " AND t.complejo_id = :complejo_id"
        params["complejo_id"] = complejo_id
    if organizador_id:
        sql += " AND t.organizador_id = :organizador_id"
        params["organizador_id"] = organizador_id
    if estado:
        sql += " AND t.estado = :estado"
        params["estado"] = estado
    if deporte:
        sql += " AND t.deporte ILIKE :deporte"
        params["deporte"] = f"%{deporte}%"
    sql += " ORDER BY t.fecha_inicio DESC"

    result = await session.execute(text(sql), params)
    rows = result.fetchall()
    keys = ["id","complejo_id","organizador_id","nombre","descripcion","deporte","fecha_inicio",
            "fecha_fin","estado","complejo_nombre","organizador_nombre","formato","max_equipos",
            "costo_inscripcion","evento_id","categoria","equipos_confirmados"]
    return [_row_to_dict(keys, r) for r in rows]


@router.post("", summary="Crear evento y categorías")
async def create_evento_y_categorias(payload: EventoCreate, session: AsyncSession = Depends(get_session)):
    try:
        fecha_ini = date.fromisoformat(payload.fecha_inicio)
        fecha_f = date.fromisoformat(payload.fecha_fin) if payload.fecha_fin else None

        # Insert Event
        evento_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.eventos
                (id, complejo_id, organizador_id, nombre, descripcion, fecha_inicio, fecha_fin, estado)
            VALUES
                (:id, :complejo_id, :organizador_id, :nombre, :descripcion, :fecha_inicio, :fecha_fin, 'abierto')
        """), {
            "id": evento_id, "complejo_id": payload.complejo_id, "organizador_id": payload.organizador_id, 
            "nombre": payload.nombre, "descripcion": payload.descripcion, "fecha_inicio": fecha_ini, "fecha_fin": fecha_f
        })

        # Insert Categories (Torneos)
        categorias_creadas = []
        for cat in payload.categorias:
            cat_id = str(uuid.uuid4())
            await session.execute(text("""
                INSERT INTO torneos.torneos
                    (id, evento_id, complejo_id, organizador_id, nombre, descripcion, deporte, formato,
                     fecha_inicio, max_equipos, costo_inscripcion, estado,
                     puntos_victoria, puntos_empate, puntos_derrota, reglas, premios, categoria, configuracion)
                VALUES
                    (:id, :evento_id, :complejo_id, :organizador_id, :nombre, :descripcion, :deporte, :formato,
                     :fecha_inicio, :max_equipos, :costo_inscripcion, 'abierto',
                     :pts_v, :pts_e, :pts_d, :reglas, :premios, :categoria, :configuracion)
            """), {
                "id": cat_id, "evento_id": evento_id, "complejo_id": payload.complejo_id, "organizador_id": payload.organizador_id,
                "nombre": f"{payload.nombre} - {cat.nombre}", "descripcion": payload.descripcion,
                "deporte": payload.deporte, "formato": cat.formato, "fecha_inicio": fecha_ini,
                "max_equipos": cat.max_equipos, "costo_inscripcion": cat.costo_inscripcion,
                "pts_v": cat.pts_victoria, "pts_e": cat.pts_empate, "pts_d": cat.pts_derrota,
                "reglas": json.dumps(payload.reglas), "premios": json.dumps(payload.premios),
                "categoria": cat.nombre, "configuracion": json.dumps(cat.configuracion)
            })
            categorias_creadas.append({"id": cat_id, "nombre": cat.nombre, "formato": cat.formato})

        await session.commit()
        return {"evento_id": evento_id, "nombre": payload.nombre, "categorias": categorias_creadas}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ── ORGANIZADORES INDEPENDIENTES (Sin complejo físico) ──────

@router.get("/organizadores", summary="Listar organizadores independientes")
async def get_organizadores(session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("""
            SELECT o.id, o.usuario_id, o.nombre, o.habilitado, o.plan, o.max_torneos, o.creado_en,
                   u.nombre_completo AS usuario_nombre, u.email AS usuario_email
            FROM cancha.organizadores o
            JOIN sistema.usuarios u ON u.id = o.usuario_id
            ORDER BY o.nombre
        """))
        rows = result.fetchall()
        cols = ["id", "usuario_id", "nombre", "habilitado", "plan", "max_torneos", "creado_en", "usuario_nombre", "usuario_email"]
        return [_row_to_dict(cols, r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/organizadores", summary="Crear o actualizar organizador independiente")
async def create_organizador(data: OrganizadorCreate, session: AsyncSession = Depends(get_session)):
    try:
        # Verificar si ya existe un organizador para este usuario
        chk = await session.execute(
            text("SELECT id FROM cancha.organizadores WHERE usuario_id = :uid"),
            {"uid": data.usuario_id}
        )
        row = chk.fetchone()
        if row:
            # Actualizar
            await session.execute(text("""
                UPDATE cancha.organizadores
                SET nombre = :nombre, plan = :plan, max_torneos = :max_torneos
                WHERE usuario_id = :uid
            """), {"uid": data.usuario_id, "nombre": data.nombre, "plan": data.plan, "max_torneos": data.max_torneos})
            await session.commit()
            return {"status": "ok", "message": "Organizador actualizado exitosamente"}
        else:
            # Insertar
            await session.execute(text("""
                INSERT INTO cancha.organizadores (usuario_id, nombre, plan, max_torneos)
                VALUES (:uid, :nombre, :plan, :max_torneos)
            """), {"uid": data.usuario_id, "nombre": data.nombre, "plan": data.plan, "max_torneos": data.max_torneos})
            await session.commit()
            return {"status": "ok", "message": "Organizador independiente creado exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/organizadores/usuario/{usuario_id}", summary="Obtener organizador por usuario")
async def get_organizador_por_usuario(usuario_id: int, session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("""
            SELECT id, usuario_id, nombre, habilitado, plan, max_torneos, creado_en
            FROM cancha.organizadores
            WHERE usuario_id = :uid AND habilitado = TRUE
        """), {"uid": usuario_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Organizador no habilitado para este usuario")
        cols = ["id", "usuario_id", "nombre", "habilitado", "plan", "max_torneos", "creado_en"]
        return _row_to_dict(cols, row)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/organizadores/{organizador_id}", summary="Eliminar organizador independiente")
async def delete_organizador(organizador_id: int, session: AsyncSession = Depends(get_session)):
    try:
        # Check if it has any associated tournaments
        # Just try to delete, if foreign key constraints fail it will raise an error
        await session.execute(
            text("DELETE FROM cancha.organizador_deporte WHERE organizador_id = :id"),
            {"id": organizador_id}
        )
        await session.execute(
            text("DELETE FROM cancha.organizadores WHERE id = :id"),
            {"id": organizador_id}
        )
        await session.commit()
        return {"status": "ok", "message": "Organizador eliminado exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail="No se pudo eliminar el organizador. Asegúrate de que no tenga datos asociados (como torneos).")



@router.get("/{torneo_id}", summary="Detalle de torneo")
async def get_torneo(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT t.id, t.complejo_id, t.organizador_id, t.nombre, t.descripcion, t.deporte,
               t.fecha_inicio, t.fecha_fin, t.estado, c.nombre AS complejo_nombre,
               org.nombre AS organizador_nombre,
               t.formato, t.max_equipos, t.costo_inscripcion,
               t.puntos_victoria, t.puntos_empate, t.puntos_derrota,
               t.reglas, t.premios, t.configuracion,
               t.imagen_portada, t.imagen_banner, t.tipo_ubicacion, t.privacidad
        FROM torneos.torneos t
        LEFT JOIN cancha.complejos c ON t.complejo_id = c.id
        LEFT JOIN cancha.organizadores org ON t.organizador_id = org.id
        WHERE t.id = CAST(:tid AS UUID)
    """), {"tid": torneo_id})
    row = result.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Torneo no encontrado")
    keys = ["id","complejo_id","organizador_id","nombre","descripcion","deporte","fecha_inicio",
            "fecha_fin","estado","complejo_nombre","organizador_nombre","formato","max_equipos",
            "costo_inscripcion","pts_victoria","pts_empate","pts_derrota",
            "reglas","premios","configuracion","imagen_portada","imagen_banner","tipo_ubicacion","privacidad"]
    d = _row_to_dict(keys, row)
    import json
    d["reglas"] = json.loads(d["reglas"]) if isinstance(d["reglas"], str) else (d["reglas"] or [])
    d["premios"] = json.loads(d["premios"]) if isinstance(d["premios"], str) else (d["premios"] or [])
    d["configuracion"] = json.loads(d["configuracion"]) if isinstance(d["configuracion"], str) else (d["configuracion"] or {})
    
    # Check for patrocinadores
    patrocinadores = []
    if d["organizador_id"]:
        res_org = await session.execute(text("""
            SELECT po.opcion_publicidad, po.posicion_banner, po.usuario_id, po.opcion_chat
            FROM sistema.perfil_organizador po
            JOIN cancha.organizadores org ON org.usuario_id = po.usuario_id
            WHERE org.id = :oid
        """), {"oid": d["organizador_id"]})
        org_data = res_org.fetchone()
        
        if org_data:
            opcion_publicidad, posicion_banner, uid, opcion_chat = org_data
            d["opcion_publicidad"] = opcion_publicidad
            d["posicion_banner"] = posicion_banner
            d["opcion_chat"] = opcion_chat
            
            if opcion_publicidad in ['torneo', 'ambos']:
                res_patroc = await session.execute(text("""
                    SELECT titulo, logo_url, banner_app_url, banner_sitio_url, tiempo_banner, sitio_web, telefono 
                    FROM sistema.patrocinadores 
                    WHERE usuario_id = :uid
                """), {"uid": uid})
                patrocinadores = [
                    {
                        "titulo": p[0], "logo_url": p[1], "banner_app_url": p[2], 
                        "banner_sitio_url": p[3], "tiempo_banner": p[4], "sitio_web": p[5], "telefono": p[6]
                    } for p in res_patroc.fetchall()
                ]
    
    d["patrocinadores"] = patrocinadores

    return d


# ============================================================
# ENDPOINTS — EQUIPOS
# ============================================================

@router.get("/{torneo_id}/equipos", summary="Equipos del torneo")
async def get_equipos(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT * FROM (
            SELECT DISTINCT ON (nombre) id, nombre, capitan_nombre, capitan_telefono, capitan_email,
                   estado_inscripcion, semilla, logo_url, color_principal, color_secundario, creado_en,
                   foto_equipo_url, token_jugadores, inscripcion_confirmada
            FROM torneos.equipos
            WHERE torneo_id = :tid 
            ORDER BY nombre ASC, creado_en ASC
        ) t
        ORDER BY t.creado_en ASC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["id","nombre","capitan_nombre","capitan_telefono","capitan_email",
            "estado_inscripcion","semilla","logo_url","color_principal","color_secundario", "creado_en", "foto_equipo_url", "token_jugadores", "inscripcion_confirmada"]
    return [_row_to_dict(keys, r) for r in rows]

@router.get("/{torneo_id}/noticias", summary="Noticias del torneo")
async def get_noticias_torneo_alias(torneo_id: str, session: AsyncSession = Depends(get_session)):
    q = text("""
        SELECT id, titulo, contenido, autor, fecha_publicacion, es_ia 
        FROM torneos.noticias 
        WHERE torneo_id = :torneo_id
        ORDER BY fecha_publicacion DESC
    """)
    res = await session.execute(q, {"torneo_id": torneo_id})
    return [{"id": r.id, "titulo": r.titulo, "contenido": r.contenido, "autor": r.autor, "fecha_publicacion": r.fecha_publicacion, "es_ia": r.es_ia} for r in res.fetchall()]

@router.get("/{torneo_id}/posiciones", summary="Tabla de posiciones")
async def get_posiciones(torneo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        t_res = await session.execute(
            text("SELECT puntos_victoria, puntos_empate, puntos_derrota FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        
        pts_v = t_row[0] if t_row[0] is not None else 3
        pts_e = t_row[1] if t_row[1] is not None else 1
        pts_d = t_row[2] if t_row[2] is not None else 0

        sql = """
            WITH partidos AS (
                SELECT 
                    equipo_local_id AS equipo_id,
                    CAST(goles_local AS INTEGER) AS gf,
                    CAST(goles_visitante AS INTEGER) AS gc,
                    CAST(CASE 
                        WHEN goles_local > goles_visitante THEN CAST(:pts_v AS INTEGER)
                        WHEN goles_local = goles_visitante THEN CAST(:pts_e AS INTEGER)
                        ELSE CAST(:pts_d AS INTEGER)
                    END AS INTEGER) AS pts,
                    CAST(CASE WHEN goles_local > goles_visitante THEN 1 ELSE 0 END AS INTEGER) AS pg,
                    CAST(CASE WHEN goles_local = goles_visitante THEN 1 ELSE 0 END AS INTEGER) AS pe,
                    CAST(CASE WHEN goles_local < goles_visitante THEN 1 ELSE 0 END AS INTEGER) AS pp
                FROM torneos.partidos
                WHERE torneo_id = :tid AND estado = 'finalizado'

                UNION ALL

                SELECT 
                    equipo_visitante_id AS equipo_id,
                    CAST(goles_visitante AS INTEGER) AS gf,
                    CAST(goles_local AS INTEGER) AS gc,
                    CAST(CASE 
                        WHEN goles_visitante > goles_local THEN CAST(:pts_v AS INTEGER)
                        WHEN goles_visitante = goles_local THEN CAST(:pts_e AS INTEGER)
                        ELSE CAST(:pts_d AS INTEGER)
                    END AS INTEGER) AS pts,
                    CAST(CASE WHEN goles_visitante > goles_local THEN 1 ELSE 0 END AS INTEGER) AS pg,
                    CAST(CASE WHEN goles_visitante = goles_local THEN 1 ELSE 0 END AS INTEGER) AS pe,
                    CAST(CASE WHEN goles_visitante < goles_local THEN 1 ELSE 0 END AS INTEGER) AS pp
                FROM torneos.partidos
                WHERE torneo_id = :tid AND estado = 'finalizado'
            ),
            equipos_distinct AS (
                SELECT DISTINCT ON (nombre) id, nombre, logo_url
                FROM torneos.equipos
                WHERE torneo_id = :tid
                ORDER BY nombre ASC, creado_en ASC
            )
            SELECT 
                e.id, 
                e.nombre,
                e.logo_url,
                COUNT(p.equipo_id) AS pj,
                COALESCE(SUM(p.pg), 0) AS pg,
                COALESCE(SUM(p.pe), 0) AS pe,
                COALESCE(SUM(p.pp), 0) AS pp,
                COALESCE(SUM(p.gf), 0) AS gf,
                COALESCE(SUM(p.gc), 0) AS gc,
                COALESCE(SUM(p.gf) - SUM(p.gc), 0) AS dif,
                COALESCE(SUM(p.pts), 0) AS pts
            FROM equipos_distinct e
            LEFT JOIN partidos p ON e.id = p.equipo_id
            GROUP BY e.id, e.nombre, e.logo_url
            ORDER BY pts DESC, dif DESC, gf DESC, e.nombre ASC
        """
        result = await session.execute(text(sql), {
            "tid": torneo_id, 
            "pts_v": pts_v, 
            "pts_e": pts_e, 
            "pts_d": pts_d
        })
        
        keys = ["id", "nombre", "logo_url", "pj", "pg", "pe", "pp", "gf", "gc", "dif", "pts"]
        return [_row_to_dict(keys, r) for r in result.fetchall()]

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# ENDPOINTS — FLUJO DE INSCRIPCIONES PUBLICO
# ============================================================

class InscripcionPublicaCreate(BaseModel):
    """Payload para que un delegado inscriba su equipo."""
    nombre_equipo: str
    nombre_academia: Optional[str] = None
    capitan_nombre: str
    capitan_email: str
    capitan_telefono: Optional[str] = None
    color_principal: Optional[str] = "#1e3a8a"
    color_secundario: Optional[str] = "#93c5fd"
    categoria_id: Optional[str] = None  # solo cuando competicion_por_atleta=false

class JugadorAutoRegistroUpdate(BaseModel):
    """Payload para que un jugador complete sus propios datos."""
    nombre: Optional[str] = None
    dni: Optional[str] = None
    fecha_nacimiento: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    numero_camiseta: Optional[int] = None
    posicion: Optional[str] = None
    peso_declarado: Optional[float] = None
    estatura_declarada: Optional[float] = None
    categoria_id: Optional[str] = None


@router.get("/{torneo_id}/info-publica", summary="Info publica del torneo para pagina de inscripcion")
async def get_torneo_info_publica(torneo_id: str, session: AsyncSession = Depends(get_session)):
    """Endpoint publico sin autenticacion para que los delegados vean el torneo."""
    try:
        t_res = await session.execute(text("""
            SELECT t.id, t.nombre, t.descripcion, t.deporte, t.estado,
                   t.fecha_inicio, t.fecha_fin, t.costo_inscripcion,
                   t.max_equipos, t.imagen_portada, t.imagen_banner,
                   t.configuracion, t.reglas,
                   org.nombre AS organizador_nombre
            FROM torneos.torneos t
            LEFT JOIN cancha.organizadores org ON t.organizador_id = org.id
            WHERE t.id = CAST(:tid AS UUID)
        """), {"tid": torneo_id})
        row = t_res.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")

        configuracion = row.configuracion or {}
        if isinstance(configuracion, str):
            try: configuracion = json.loads(configuracion)
            except: configuracion = {}

        por_atleta = configuracion.get("competicion_por_atleta", False)

        # Traer categorias del torneo
        cat_res = await session.execute(text("""
            SELECT id, nombre, formato, descripcion, configuracion
            FROM torneos.categorias
            WHERE torneo_id = CAST(:tid AS UUID)
            ORDER BY nombre ASC
        """), {"tid": torneo_id})
        categorias = [dict(r._mapping) for r in cat_res.fetchall()]

        # Contar equipos ya inscritos
        eq_res = await session.execute(text("""
            SELECT COUNT(*) FROM torneos.equipos
            WHERE torneo_id = CAST(:tid AS UUID) AND estado_inscripcion != 'eliminado'
        """), {"tid": torneo_id})
        equipos_count = eq_res.scalar() or 0

        reglas = row.reglas
        if isinstance(reglas, str):
            try: reglas = json.loads(reglas)
            except: reglas = []

        return {
            "id": str(row.id),
            "nombre": row.nombre,
            "descripcion": row.descripcion,
            "deporte": row.deporte,
            "estado": row.estado,
            "fecha_inicio": str(row.fecha_inicio) if row.fecha_inicio else None,
            "fecha_fin": str(row.fecha_fin) if row.fecha_fin else None,
            "costo_inscripcion": float(row.costo_inscripcion or 0),
            "max_equipos": row.max_equipos,
            "imagen_portada": row.imagen_portada,
            "imagen_banner": row.imagen_banner,
            "organizador_nombre": row.organizador_nombre,
            "competicion_por_atleta": por_atleta,
            "categorias": categorias,
            "equipos_inscritos": int(equipos_count),
            "reglas": reglas or [],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{torneo_id}/inscripcion-publica", summary="Inscripcion publica de equipo (delegado)")
async def inscripcion_publica_equipo(
    torneo_id: str,
    payload: InscripcionPublicaCreate,
    session: AsyncSession = Depends(get_session)
):
    """Delegado inscribe su equipo/academia al torneo. Devuelve token_delegado para gestionar."""
    try:
        # Verificar que el torneo existe y esta abierto
        t_res = await session.execute(text("""
            SELECT id, nombre, costo_inscripcion, estado, configuracion
            FROM torneos.torneos
            WHERE id = CAST(:tid AS UUID)
        """), {"tid": torneo_id})
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        if t_row.estado not in ("abierto", "en_curso", "preparacion"):
            raise HTTPException(status_code=400, detail="El torneo no esta aceptando inscripciones en este momento")

        configuracion = t_row.configuracion or {}
        if isinstance(configuracion, str):
            try: configuracion = json.loads(configuracion)
            except: configuracion = {}
        por_atleta = configuracion.get("competicion_por_atleta", False)

        # Si no es por_atleta, categoria_id es requerida
        if not por_atleta and not payload.categoria_id:
            # Intentar tomar la primera categoria si hay una sola
            cat_res = await session.execute(text("""
                SELECT id FROM torneos.categorias WHERE torneo_id = CAST(:tid AS UUID) LIMIT 1
            """), {"tid": torneo_id})
            cat_row = cat_res.fetchone()
            if cat_row:
                payload.categoria_id = str(cat_row.id)

        costo = float(t_row.costo_inscripcion or 0)
        estado_insc = "confirmado" if costo <= 0 else "pendiente"

        # Verificar que el email no ya esta registrado en este torneo
        dup_res = await session.execute(text("""
            SELECT id FROM torneos.equipos
            WHERE torneo_id = CAST(:tid AS UUID) AND capitan_email = :email AND estado_inscripcion != 'eliminado'
        """), {"tid": torneo_id, "email": payload.capitan_email.strip().lower()})
        if dup_res.fetchone():
            raise HTTPException(status_code=400, detail="Ya existe una inscripcion con este email en este torneo")

        # Insertar el equipo
        ins_res = await session.execute(text("""
            INSERT INTO torneos.equipos (
                id, torneo_id, nombre, nombre_academia, capitan_nombre, capitan_email,
                capitan_telefono, color_principal, color_secundario,
                estado_inscripcion, categoria_id
            )
            VALUES (
                gen_random_uuid(), CAST(:tid AS UUID), :nombre, :academia, :cap_nombre, :cap_email,
                :cap_tel, :color_p, :color_s, :estado, :cat_id
            )
            RETURNING id, token_delegado, token_jugadores
        """), {
            "tid": torneo_id,
            "nombre": payload.nombre_equipo,
            "academia": payload.nombre_academia or payload.nombre_equipo,
            "cap_nombre": payload.capitan_nombre,
            "cap_email": payload.capitan_email.strip().lower(),
            "cap_tel": payload.capitan_telefono,
            "color_p": payload.color_principal,
            "color_s": payload.color_secundario,
            "estado": estado_insc,
            "cat_id": payload.categoria_id,
        })
        row = ins_res.fetchone()
        await session.commit()

        equipo_id = str(row.id)
        token_delegado = str(row.token_delegado)
        token_jugadores = str(row.token_jugadores)

        # Crear usuario delegado si no existe
        if payload.capitan_email:
            email_del = payload.capitan_email.strip().lower()
            u_res = await session.execute(
                text("SELECT id FROM sistema.usuarios WHERE email = :email"),
                {"email": email_del}
            )
            if not u_res.fetchone():
                usr_base = email_del.split('@')[0]
                usr_check = await session.execute(
                    text("SELECT id FROM sistema.usuarios WHERE username = :username"),
                    {"username": usr_base}
                )
                username_final = usr_base if not usr_check.fetchone() else f"{usr_base}_{secrets.token_hex(2)}"
                pass_temp = generate_random_password(10)
                pass_hash = get_password_hash(pass_temp)
                await session.execute(text("""
                    INSERT INTO sistema.usuarios (username, email, hashed_password, nombre_completo, rol, activo, fecha_creacion)
                    VALUES (:username, :email, :pass_hash, :nombre, 'delegado', true, NOW())
                """), {
                    "username": username_final,
                    "email": email_del,
                    "pass_hash": pass_hash,
                    "nombre": payload.capitan_nombre
                })
                await session.commit()
                try:
                    email_service.send_welcome_email(
                        to_email=email_del,
                        username=username_final,
                        password=pass_temp,
                        role="Delegado de Equipo"
                    )
                except Exception as mail_err:
                    print("Error al enviar email a delegado:", str(mail_err))

        frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
        enlace_delegado = f"{frontend_url}/delegados/{token_delegado}"
        enlace_jugadores = f"{frontend_url}/jugadores/registro/{token_jugadores}"

        # Enviar email con los enlaces al delegado
        if payload.capitan_email:
            try:
                email_service.send_delegado_link_email(
                    to_email=payload.capitan_email.strip().lower(),
                    equipo_nombre=payload.nombre_equipo,
                    torneo_nombre=t_row.nombre if hasattr(t_row, 'nombre') else "Torneo",
                    enlace_delegado=enlace_delegado,
                    enlace_jugadores=enlace_jugadores,
                    estado_inscripcion=estado_insc,
                    costo_inscripcion=costo
                )
            except Exception as mail_err:
                print("Error al enviar email de enlaces al delegado:", str(mail_err))

        return {
            "equipo_id": equipo_id,
            "token_delegado": token_delegado,
            "token_jugadores": token_jugadores,
            "estado_inscripcion": estado_insc,
            "enlace_delegado": enlace_delegado,
            "enlace_jugadores": enlace_jugadores,
            "mensaje": "Inscripcion realizada con exito. Guarda el enlace de delegado para gestionar tu equipo.",
        }
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/delegados/{token}", summary="Panel del delegado por token")
async def get_panel_delegado(token: str, session: AsyncSession = Depends(get_session)):
    """Devuelve toda la info del equipo, torneo, jugadores y categorias disponibles."""
    try:
        eq_res = await session.execute(text("""
            SELECT te.id, te.nombre, te.nombre_academia, te.logo_url, te.foto_equipo_url,
                   te.color_principal, te.color_secundario,
                   te.capitan_nombre, te.capitan_telefono, te.capitan_email,
                   te.estado_inscripcion, te.token_jugadores, te.categoria_id,
                   t.id AS torneo_id, t.nombre AS torneo_nombre, t.deporte, t.formato,
                   t.estado AS torneo_estado, t.configuracion AS torneo_config,
                   t.imagen_portada, t.costo_inscripcion
            FROM torneos.equipos te
            JOIN torneos.torneos t ON te.torneo_id = t.id
            WHERE te.token_delegado = CAST(:token AS UUID)
        """), {"token": token})
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Enlace de delegado no valido o expirado")

        torneo_config = eq_row.torneo_config or {}
        if isinstance(torneo_config, str):
            try: torneo_config = json.loads(torneo_config)
            except: torneo_config = {}
        por_atleta = torneo_config.get("competicion_por_atleta", False)

        # Jugadores del equipo
        pl_res = await session.execute(text("""
            SELECT tp.id, tp.nombre, tp.dni, tp.fecha_nacimiento, tp.numero_camiseta,
                   tp.posicion, tp.foto_url, tp.estado, tp.email,
                   tp.token_jugador, tp.categoria_id,
                   tp.peso_verificado, tp.estatura_verificada, tp.biometria_aprobada,
                   c.nombre AS categoria_nombre
            FROM torneos.tournament_players tp
            LEFT JOIN torneos.categorias c ON tp.categoria_id = c.id
            WHERE tp.torneo_equipo_id = :eid
            ORDER BY tp.numero_camiseta ASC NULLS LAST, tp.nombre ASC
        """), {"eid": str(eq_row.id)})
        pl_rows = pl_res.fetchall()
        jugadores = []
        for r in pl_rows:
            jugadores.append({
                "id": str(r.id),
                "nombre": r.nombre,
                "dni": r.dni,
                "fecha_nacimiento": str(r.fecha_nacimiento) if r.fecha_nacimiento else None,
                "numero_camiseta": r.numero_camiseta,
                "posicion": r.posicion,
                "foto_url": r.foto_url,
                "estado": r.estado,
                "email": r.email,
                "token_jugador": str(r.token_jugador) if r.token_jugador else None,
                "categoria_id": str(r.categoria_id) if r.categoria_id else None,
                "categoria_nombre": r.categoria_nombre,
                "peso_verificado": float(r.peso_verificado) if r.peso_verificado else None,
                "estatura_verificada": float(r.estatura_verificada) if r.estatura_verificada else None,
                "biometria_aprobada": r.biometria_aprobada,
            })

        # Categorias disponibles en el torneo (para asignar por atleta)
        cat_res = await session.execute(text("""
            SELECT id, nombre, descripcion, configuracion
            FROM torneos.categorias
            WHERE torneo_id = :tid
            ORDER BY nombre ASC
        """), {"tid": str(eq_row.torneo_id)})
        categorias = [dict(r._mapping) for r in cat_res.fetchall()]

        frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
        return {
            "equipo": {
                "id": str(eq_row.id),
                "nombre": eq_row.nombre,
                "nombre_academia": eq_row.nombre_academia,
                "logo_url": eq_row.logo_url,
                "foto_equipo_url": eq_row.foto_equipo_url,
                "color_principal": eq_row.color_principal,
                "color_secundario": eq_row.color_secundario,
                "capitan_nombre": eq_row.capitan_nombre,
                "capitan_telefono": eq_row.capitan_telefono,
                "capitan_email": eq_row.capitan_email,
                "estado_inscripcion": eq_row.estado_inscripcion,
                "token_jugadores": str(eq_row.token_jugadores) if eq_row.token_jugadores else None,
                "categoria_id": str(eq_row.categoria_id) if eq_row.categoria_id else None,
            },
            "torneo": {
                "id": str(eq_row.torneo_id),
                "nombre": eq_row.torneo_nombre,
                "deporte": eq_row.deporte,
                "formato": eq_row.formato,
                "estado": eq_row.torneo_estado,
                "imagen_portada": eq_row.imagen_portada,
                "costo_inscripcion": float(eq_row.costo_inscripcion or 0),
                "competicion_por_atleta": por_atleta,
            },
            "jugadores": jugadores,
            "categorias": categorias,
            "enlace_jugadores": f"{frontend_url}/jugadores/registro/{eq_row.token_jugadores}",
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/delegados/{token}/jugadores", summary="Delegado agrega jugador al equipo")
async def delegado_agregar_jugador(
    token: str,
    payload: JugadorCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        eq_res = await session.execute(text("""
            SELECT id, torneo_id FROM torneos.equipos WHERE token_delegado = CAST(:token AS UUID)
        """), {"token": token})
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token de delegado no valido")

        ins_res = await session.execute(text("""
            INSERT INTO torneos.tournament_players (
                id, torneo_equipo_id, nombre, dni, fecha_nacimiento, numero_camiseta,
                posicion, foto_url, estado, email
            )
            VALUES (
                gen_random_uuid(), :eid, :nombre, :dni, :fnac, :num,
                :pos, :foto, 'habilitado', :email
            )
            RETURNING id, token_jugador
        """), {
            "eid": str(eq_row.id),
            "nombre": payload.nombre,
            "dni": payload.dni,
            "fnac": payload.fecha_nacimiento,
            "num": payload.numero_camiseta,
            "pos": payload.posicion,
            "foto": payload.foto_url,
            "email": payload.email,
        })
        row = ins_res.fetchone()
        await session.commit()

        frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
        return {
            "id": str(row.id),
            "token_jugador": str(row.token_jugador) if row.token_jugador else None,
            "enlace_autoregistro": f"{frontend_url}/jugadores/registro/{row.token_jugador}",
            "mensaje": "Jugador agregado correctamente"
        }
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/delegados/{token}/jugadores/{jugador_id}/categoria", summary="Asignar categoria a un jugador (por atleta)")
async def delegado_asignar_categoria_jugador(
    token: str,
    jugador_id: str,
    categoria_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        eq_res = await session.execute(text("""
            SELECT id FROM torneos.equipos WHERE token_delegado = CAST(:token AS UUID)
        """), {"token": token})
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token de delegado no valido")

        upd = await session.execute(text("""
            UPDATE torneos.tournament_players
            SET categoria_id = CAST(:cat_id AS UUID)
            WHERE id = CAST(:jid AS UUID) AND torneo_equipo_id = :eid
        """), {"cat_id": categoria_id, "jid": jugador_id, "eid": str(eq_row.id)})
        await session.commit()

        if upd.rowcount == 0:
            raise HTTPException(status_code=404, detail="Jugador no encontrado en este equipo")
        return {"mensaje": "Categoria asignada correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/jugadores/token/{token}", summary="Info del jugador por su token de auto-registro")
async def get_jugador_por_token(token: str, session: AsyncSession = Depends(get_session)):
    try:
        res = await session.execute(text("""
            SELECT tp.id, tp.nombre, tp.dni, tp.fecha_nacimiento, tp.numero_camiseta,
                   tp.posicion, tp.foto_url, tp.estado, tp.email,
                   tp.peso_verificado, tp.estatura_verificada, tp.categoria_id,
                   te.nombre AS equipo_nombre, te.nombre_academia, te.color_principal,
                   te.torneo_id, t.nombre AS torneo_nombre, t.deporte,
                   t.imagen_portada, t.configuracion AS torneo_config
            FROM torneos.tournament_players tp
            JOIN torneos.equipos te ON tp.torneo_equipo_id = te.id
            JOIN torneos.torneos t ON te.torneo_id = t.id
            WHERE tp.token_jugador = CAST(:token AS UUID)
        """), {"token": token})
        row = res.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Token de jugador no valido o expirado")

        torneo_config = row.torneo_config or {}
        if isinstance(torneo_config, str):
            try: torneo_config = json.loads(torneo_config)
            except: torneo_config = {}

        # Categorias disponibles
        cat_res = await session.execute(text("""
            SELECT id, nombre, descripcion, configuracion
            FROM torneos.categorias
            WHERE torneo_id = :tid
            ORDER BY nombre ASC
        """), {"tid": str(row.torneo_id)})
        categorias = [dict(r._mapping) for r in cat_res.fetchall()]

        return {
            "jugador": {
                "id": str(row.id),
                "nombre": row.nombre,
                "dni": row.dni,
                "fecha_nacimiento": str(row.fecha_nacimiento) if row.fecha_nacimiento else None,
                "numero_camiseta": row.numero_camiseta,
                "posicion": row.posicion,
                "foto_url": row.foto_url,
                "estado": row.estado,
                "email": row.email,
                "peso_verificado": float(row.peso_verificado) if row.peso_verificado else None,
                "estatura_verificada": float(row.estatura_verificada) if row.estatura_verificada else None,
                "categoria_id": str(row.categoria_id) if row.categoria_id else None,
            },
            "equipo": {
                "nombre": row.equipo_nombre,
                "nombre_academia": row.nombre_academia,
                "color_principal": row.color_principal,
            },
            "torneo": {
                "id": str(row.torneo_id),
                "nombre": row.torneo_nombre,
                "deporte": row.deporte,
                "imagen_portada": row.imagen_portada,
                "competicion_por_atleta": torneo_config.get("competicion_por_atleta", False),
            },
            "categorias": categorias,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/jugadores/token/{token}", summary="Jugador completa sus datos de auto-registro")
async def jugador_autoregistro(
    token: str,
    payload: JugadorAutoRegistroUpdate,
    session: AsyncSession = Depends(get_session)
):
    try:
        # Verificar que el token existe
        check = await session.execute(text("""
            SELECT tp.id FROM torneos.tournament_players tp
            WHERE tp.token_jugador = CAST(:token AS UUID)
        """), {"token": token})
        if not check.fetchone():
            raise HTTPException(status_code=404, detail="Token de jugador no valido")

        updates = []
        params = {"token": token}

        if payload.nombre: updates.append("nombre = :nombre"); params["nombre"] = payload.nombre
        if payload.dni: updates.append("dni = :dni"); params["dni"] = payload.dni
        if payload.fecha_nacimiento: updates.append("fecha_nacimiento = :fnac"); params["fnac"] = payload.fecha_nacimiento
        if payload.email: updates.append("email = :email"); params["email"] = payload.email
        if payload.numero_camiseta is not None: updates.append("numero_camiseta = :num"); params["num"] = payload.numero_camiseta
        if payload.posicion: updates.append("posicion = :pos"); params["pos"] = payload.posicion
        if payload.peso_declarado is not None: updates.append("peso_verificado = :peso"); params["peso"] = payload.peso_declarado
        if payload.estatura_declarada is not None: updates.append("estatura_verificada = :est"); params["est"] = payload.estatura_declarada
        if payload.categoria_id: updates.append("categoria_id = CAST(:cat_id AS UUID)"); params["cat_id"] = payload.categoria_id

        if updates:
            await session.execute(text(f"""
                UPDATE torneos.tournament_players
                SET {', '.join(updates)}, actualizado_en = NOW()
                WHERE token_jugador = CAST(:token AS UUID)
            """), params)
            await session.commit()

        return {"mensaje": "Datos actualizados correctamente. Tu inscripcion esta confirmada."}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{torneo_id}/equipos", summary="Inscribir equipo")
async def create_equipo(torneo_id: str, payload: EquipoCreate, session: AsyncSession = Depends(get_session)):
    try:
        t_res = await session.execute(
            text("SELECT costo_inscripcion FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")

        costo = float(t_row[0]) if t_row[0] else 0.0
        estado_insc = "confirmado" if costo <= 0 else "pendiente"

        result = await session.execute(text("""
            INSERT INTO torneos.equipos
                (id, torneo_id, nombre, capitan_nombre, capitan_telefono, capitan_email,
                 estado_inscripcion, logo_url, color_principal, color_secundario, promocion)
            VALUES
                (gen_random_uuid(), :tid, :nombre, :capitan_nombre, :capitan_telefono,
                 :capitan_email, :estado, :logo_url, :color_p, :color_s, :promocion)
            RETURNING id, torneo_id, nombre, estado_inscripcion, promocion, token_delegado
        """), {
            "tid": torneo_id, "nombre": payload.nombre,
            "capitan_nombre": payload.capitan_nombre,
            "capitan_telefono": payload.capitan_telefono,
            "capitan_email": payload.capitan_email,
            "estado": estado_insc,
            "logo_url": payload.logo_url,
            "color_p": payload.color_principal,
            "color_s": payload.color_secundario,
            "promocion": payload.promocion
        })
        await session.commit()
        row = result.fetchone()
        
        # Crear usuario para el delegado (si tiene email cargado)
        if payload.capitan_email:
            email_del = payload.capitan_email.strip()
            # Verificar si existe el usuario
            u_res = await session.execute(
                text("SELECT id FROM sistema.usuarios WHERE email = :email"),
                {"email": email_del}
            )
            if not u_res.fetchone():
                # Generar username
                usr_base = email_del.split('@')[0]
                usr_check = await session.execute(
                    text("SELECT id FROM sistema.usuarios WHERE username = :username"),
                    {"username": usr_base}
                )
                username_final = usr_base
                if usr_check.fetchone():
                    username_final = f"{usr_base}_{secrets.token_hex(2)}"
                
                # Crear contraseña temporal
                pass_temp = generate_random_password(10)
                pass_hash = get_password_hash(pass_temp)
                
                # Insertar usuario
                await session.execute(text("""
                    INSERT INTO sistema.usuarios (username, email, hashed_password, nombre_completo, rol, activo, fecha_creacion)
                    VALUES (:username, :email, :pass_hash, :nombre, 'delegado', true, NOW())
                """), {
                    "username": username_final,
                    "email": email_del,
                    "pass_hash": pass_hash,
                    "nombre": payload.capitan_nombre or payload.nombre
                })
                await session.commit()
                
                # Enviar correo de bienvenida
                try:
                    email_service.send_welcome_email(
                        to_email=email_del,
                        username=username_final,
                        password=pass_temp,
                        role="Delegado de Equipo"
                    )
                except Exception as mail_err:
                    print("Error al enviar email a delegado:", str(mail_err))
                    
        return _row_to_dict(["id","torneo_id","nombre","estado_inscripcion","promocion","token_delegado"], row)
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{torneo_id}/equipos/{equipo_id}", summary="Eliminar equipo")
async def delete_equipo(
    torneo_id: str,
    equipo_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        # Check if the team exists
        res = await session.execute(text("SELECT id FROM torneos.equipos WHERE id = CAST(:eid AS UUID) AND torneo_id = CAST(:tid AS UUID)"), {"eid": equipo_id, "tid": torneo_id})
        if not res.scalar():
            raise HTTPException(status_code=404, detail="Equipo no encontrado")
            
        # Eliminar técnicos
        await session.execute(text("DELETE FROM torneos.equipo_tecnico WHERE equipo_id = CAST(:eid AS UUID)"), {"eid": equipo_id})
        # Eliminar jugadores
        await session.execute(text("DELETE FROM torneos.tournament_players WHERE torneo_equipo_id = CAST(:eid AS UUID)"), {"eid": equipo_id})
        # Eliminar el equipo
        await session.execute(text("DELETE FROM torneos.equipos WHERE id = CAST(:eid AS UUID) AND torneo_id = CAST(:tid AS UUID)"), {"eid": equipo_id, "tid": torneo_id})
        await session.commit()
        return {"mensaje": "Equipo eliminado exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.patch("/{torneo_id}/equipos/{equipo_id}", summary="Actualizar equipo (incluyendo delegado y confirmación)")
async def update_equipo(
    torneo_id: str,
    equipo_id: str,
    payload: EquipoUpdate,
    session: AsyncSession = Depends(get_session)
):
    try:
        # Verificar existencia y traer datos necesarios para el email
        res = await session.execute(
            text("""
                SELECT e.id, e.nombre, e.capitan_email, e.capitan_nombre,
                       e.inscripcion_confirmada, e.token_delegado, e.token_jugadores,
                       t.nombre AS torneo_nombre
                FROM torneos.equipos e
                JOIN torneos.torneos t ON e.torneo_id = t.id
                WHERE e.id = CAST(:eid AS UUID) AND e.torneo_id = CAST(:tid AS UUID)
            """),
            {"eid": equipo_id, "tid": torneo_id}
        )
        row = res.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Equipo no encontrado")

        was_confirmed = row.inscripcion_confirmada  # previous state

        updates = []
        params = {"eid": equipo_id, "tid": torneo_id}

        if payload.nombre is not None:
            updates.append("nombre = :nombre")
            params["nombre"] = payload.nombre
        if payload.capitan_nombre is not None:
            updates.append("capitan_nombre = :cn")
            params["cn"] = payload.capitan_nombre
        if payload.capitan_telefono is not None:
            updates.append("capitan_telefono = :ct")
            params["ct"] = payload.capitan_telefono
        if payload.capitan_email is not None:
            updates.append("capitan_email = :ce")
            params["ce"] = payload.capitan_email
        if payload.inscripcion_confirmada is not None:
            updates.append("inscripcion_confirmada = :ic")
            params["ic"] = payload.inscripcion_confirmada
            # Sync estado_inscripcion with the boolean
            if payload.inscripcion_confirmada is True:
                updates.append("estado_inscripcion = 'confirmado'")
            elif payload.inscripcion_confirmada is False:
                updates.append("estado_inscripcion = 'eliminado'")
            else:
                updates.append("estado_inscripcion = 'pendiente'")

        if not updates:
            return {"status": "ok", "message": "Sin cambios"}

        sql = f"UPDATE torneos.equipos SET {', '.join(updates)} WHERE id = CAST(:eid AS UUID) AND torneo_id = CAST(:tid AS UUID)"
        await session.execute(text(sql), params)
        await session.commit()

        # Send approval email if just confirmed (was not confirmed before)
        if payload.inscripcion_confirmada is True and not was_confirmed:
            capitan_email = row.capitan_email
            if capitan_email:
                try:
                    # Look up the username created during inscription
                    u_res = await session.execute(
                        text("SELECT username FROM sistema.usuarios WHERE email = :email"),
                        {"email": capitan_email.strip()}
                    )
                    u_row = u_res.fetchone()
                    delegado_username = u_row[0] if u_row else capitan_email

                    email_service.send_inscription_approved_email(
                        to_email=capitan_email,
                        equipo_nombre=row.nombre,
                        torneo_nombre=row.torneo_nombre,
                        username=delegado_username,
                        token_jugadores=str(row.token_jugadores) if row.token_jugadores else ""
                    )
                except Exception as mail_err:
                    print("Error al enviar email de aprobación:", str(mail_err))

        return {"status": "ok", "message": "Equipo actualizado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{torneo_id}/equipos/{equipo_id}/logo", summary="Subir logo del equipo")
async def upload_equipo_logo(
    torneo_id: str,
    equipo_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "logos")
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
        filename = f"logo_equipo_{equipo_id}_{uuid.uuid4().hex[:8]}.{ext}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        logo_url = f"/static/uploads/logos/{filename}"

        sql = "UPDATE torneos.equipos SET logo_url = :logo_url WHERE id = :eid AND torneo_id = :tid RETURNING id"
        res = await session.execute(text(sql), {"logo_url": logo_url, "eid": equipo_id, "tid": torneo_id})
        await session.commit()

        if not res.fetchone():
            raise HTTPException(status_code=404, detail="Equipo no encontrado")

        return {"status": "ok", "logo_url": logo_url}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{torneo_id}/equipos/{equipo_id}/foto", summary="Subir foto del equipo")
async def upload_equipo_foto(
    torneo_id: str,
    equipo_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "fotos_equipos")
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
        filename = f"foto_equipo_{equipo_id}_{uuid.uuid4().hex[:8]}.{ext}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        foto_url = f"/static/uploads/fotos_equipos/{filename}"

        sql = "UPDATE torneos.equipos SET foto_equipo_url = :foto_url WHERE id = :eid AND torneo_id = :tid RETURNING id"
        res = await session.execute(text(sql), {"foto_url": foto_url, "eid": equipo_id, "tid": torneo_id})
        await session.commit()

        if not res.fetchone():
            raise HTTPException(status_code=404, detail="Equipo no encontrado")

        return {"status": "ok", "foto_equipo_url": foto_url}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/equipos/token/{token}", summary="Validar token de delegado")
async def get_equipo_by_token(token: str, session: AsyncSession = Depends(get_session)):
    try:
        # 1. Buscar equipo y torneo correspondiente
        eq_res = await session.execute(text("""
            SELECT te.id AS equipo_id, te.nombre AS equipo_nombre, te.logo_url, te.foto_equipo_url, te.color_principal, te.color_secundario,
                   te.capitan_nombre, te.capitan_telefono, te.capitan_email, te.promocion, te.token_jugadores,
                   t.id AS torneo_id, t.nombre AS torneo_nombre, t.deporte, t.formato, t.estado AS torneo_estado
            FROM torneos.equipos te
            JOIN torneos.torneos t ON te.torneo_id = t.id
            WHERE te.token_delegado = :token
        """), {"token": token})
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Enlace de delegado no válido o expirado")

        keys = ["equipo_id", "equipo_nombre", "logo_url", "foto_equipo_url", "color_principal", "color_secundario",
                "capitan_nombre", "capitan_telefono", "capitan_email", "promocion", "token_jugadores",
                "torneo_id", "torneo_nombre", "deporte", "formato", "torneo_estado"]
        data = _row_to_dict(keys, eq_row)

        # 2. Buscar plantilla de jugadores del equipo
        pl_res = await session.execute(text("""
            SELECT id, nombre, dni, fecha_nacimiento, numero_camiseta, posicion, foto_url, estado,
                   documento_firmado_url, cedula_anverso_url, cedula_reverso_url, egreso_ano, es_exalumno
            FROM cancha.tournament_players
            WHERE tournament_team_id = :eid
            ORDER BY numero_camiseta ASC NULLS LAST, nombre ASC
        """), {"eid": data["equipo_id"]})
        pl_rows = pl_res.fetchall()
        pl_keys = ["id", "nombre", "dni", "fecha_nacimiento", "numero_camiseta", "posicion", "foto_url", "estado",
                   "documento_firmado_url", "cedula_anverso_url", "cedula_reverso_url", "egreso_ano", "es_exalumno"]
        
        jugadores = [_row_to_dict(pl_keys, r) for r in pl_rows]

        return {
            "equipo": {
                "id": data["equipo_id"],
                "nombre": data["equipo_nombre"],
                "logo_url": data["logo_url"],
                "color_principal": data["color_principal"],
                "color_secundario": data["color_secundario"],
                "capitan_nombre": data["capitan_nombre"],
                "capitan_telefono": data["capitan_telefono"],
                "capitan_email": data["capitan_email"],
                "promocion": data["promocion"]
            },
            "torneo": {
                "id": data["torneo_id"],
                "nombre": data["torneo_nombre"],
                "deporte": data["deporte"],
                "formato": data["formato"],
                "estado": data["torneo_estado"]
            },
            "jugadores": jugadores
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/equipos/token-jugadores/{token_jugadores}", summary="Info del equipo por token de jugadores")
async def get_equipo_by_token_jugadores(token_jugadores: str, session: AsyncSession = Depends(get_session)):
    """Endpoint público para que la página de auto-registro cargue el nombre del equipo y torneo."""
    try:
        res = await session.execute(text("""
            SELECT te.nombre AS equipo_nombre, te.token_jugadores,
                   t.nombre AS torneo_nombre, t.deporte
            FROM torneos.equipos te
            JOIN torneos.torneos t ON te.torneo_id = t.id
            WHERE te.token_jugadores = :token
        """), {"token": token_jugadores})
        row = res.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Token de jugadores no válido")
        return {
            "nombre": row.equipo_nombre,
            "torneo_nombre": row.torneo_nombre,
            "deporte": row.deporte,
            "token_jugadores": str(row.token_jugadores)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




@router.get("/equipos/buscar-token", summary="Buscar token de delegado por email")
async def buscar_token_delegado(
    email: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        res = await session.execute(
            text("SELECT token_delegado FROM torneos.equipos WHERE capitan_email = :email"),
            {"email": email.strip()}
        )
        row = res.fetchone()
        if not row or not row[0]:
            raise HTTPException(status_code=404, detail="No se encontró equipo o token para este correo de delegado")
        return {"token_delegado": str(row[0])}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/equipos/token/{token}/logo", summary="Subir logo del equipo con token de delegado")
async def upload_equipo_logo_by_token(
    token: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

        # Validar token
        eq_res = await session.execute(
            text("SELECT id, torneo_id FROM torneos.equipos WHERE token_delegado = :token"),
            {"token": token}
        )
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token no válido")

        equipo_id, torneo_id = eq_row

        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "logos")
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
        filename = f"logo_equipo_{equipo_id}_{uuid.uuid4().hex[:8]}.{ext}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        logo_url = f"/static/uploads/logos/{filename}"

        await session.execute(
            text("UPDATE torneos.equipos SET logo_url = :logo_url WHERE id = :eid"),
            {"logo_url": logo_url, "eid": equipo_id}
        )
        await session.commit()

        return {"status": "ok", "logo_url": logo_url}
    except HTTPException:
        raise
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/equipos/token/{token}/foto", summary="Subir foto del equipo (Delegado)")
async def upload_equipo_foto_by_token(
    token: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")

        # Validar token
        eq_res = await session.execute(
            text("SELECT id, torneo_id FROM torneos.equipos WHERE token_delegado = :token"),
            {"token": token}
        )
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token no válido")

        equipo_id, torneo_id = eq_row

        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "fotos_equipos")
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
        filename = f"foto_equipo_{equipo_id}_{uuid.uuid4().hex[:8]}.{ext}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        foto_url = f"/static/uploads/fotos_equipos/{filename}"

        await session.execute(
            text("UPDATE torneos.equipos SET foto_equipo_url = :foto_url WHERE id = :eid"),
            {"foto_url": foto_url, "eid": equipo_id}
        )
        await session.commit()

        return {"status": "ok", "foto_equipo_url": foto_url}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/equipos/token/{token}/jugadores", summary="Agregar jugador al plantel usando token de delegado")
async def add_jugador_by_token(
    token: str,
    payload: JugadorCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        # Validar token
        eq_res = await session.execute(
            text("SELECT id, torneo_id FROM torneos.equipos WHERE token_delegado = :token"),
            {"token": token}
        )
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token de delegado no válido")

        equipo_id, torneo_id = eq_row
        return await _add_jugador_logic(torneo_id, equipo_id, payload, session)
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/documentacion/upload", summary="Subir archivos de documentación de jugadores")
async def upload_documentacion(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        # Validar tipo de archivo
        allowed_types = ["image/jpeg", "image/png", "image/webp", "application/pdf"]
        if file.content_type not in allowed_types:
            raise HTTPException(status_code=400, detail="Tipo de archivo no permitido. Solo se permiten imágenes (PNG/JPG/WEBP) y PDFs.")

        # Guardar en static/uploads/documentacion
        upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "documentacion")
        os.makedirs(upload_dir, exist_ok=True)

        ext = file.filename.split('.')[-1] if '.' in file.filename else 'dat'
        filename = f"doc_{uuid.uuid4().hex}_{int(time.time())}.{ext}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        file_url = f"/static/uploads/documentacion/{filename}"
        return {"status": "ok", "url": file_url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/jugadores/self-register/{token_jugadores}", summary="Auto-registro de jugador por token")
async def self_register_jugador(
    token_jugadores: str,
    nombre: str = Form(...),
    dni: str = Form(...),
    fecha_nacimiento: str = Form(None),
    numero_camiseta: int = Form(None),
    posicion: str = Form(None),
    file: UploadFile = File(None),
    session: AsyncSession = Depends(get_session)
):
    try:
        # Validar token y obtener torneo_equipo_id
        eq_res = await session.execute(
            text("SELECT id, torneo_id FROM torneos.equipos WHERE token_jugadores = :token"),
            {"token": token_jugadores}
        )
        eq_row = eq_res.fetchone()
        if not eq_row:
            raise HTTPException(status_code=404, detail="Token de registro no válido")

        equipo_id, torneo_id = eq_row

        # Validar si el DNI o camiseta ya existen en el equipo
        dupl_res = await session.execute(
            text("SELECT dni, numero_camiseta FROM cancha.tournament_players WHERE torneo_equipo_id = :eid AND (dni = :dni OR numero_camiseta = :cam)"),
            {"eid": equipo_id, "dni": dni, "cam": numero_camiseta if numero_camiseta is not None else -1}
        )
        for r in dupl_res.fetchall():
            if r.dni == dni:
                raise HTTPException(status_code=400, detail="El DNI ya está registrado en este equipo")
            if r.numero_camiseta == numero_camiseta:
                raise HTTPException(status_code=400, detail="El número de camiseta ya está en uso")

        # Procesar la foto (si se envió)
        foto_url = None
        face_encoding_json = None
        if file:
            if not file.content_type.startswith("image/"):
                raise HTTPException(status_code=400, detail="El archivo de foto debe ser una imagen")

            file_bytes = await file.read()
            if len(file_bytes) > 5 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="La foto es demasiado grande (máximo 5MB)")

            # Intentar extraer vector facial si hay rostro
            try:
                encoding = FacialRecognitionService.extract_encoding(file_bytes)
                face_encoding_json = json.dumps(encoding)
            except ValueError:
                # Si falla, no cortamos el flujo, solo no guardamos vector facial
                pass

            upload_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "fotos_perfil")
            os.makedirs(upload_dir, exist_ok=True)
            ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
            filename = f"perfil_{uuid.uuid4().hex[:12]}.{ext}"
            filepath = os.path.join(upload_dir, filename)

            with open(filepath, "wb") as buffer:
                buffer.write(file_bytes)

            foto_url = f"/static/uploads/fotos_perfil/{filename}"

        fnac = date.fromisoformat(fecha_nacimiento) if fecha_nacimiento else None

        # Insertar jugador
        sql = """
            INSERT INTO cancha.tournament_players
            (torneo_equipo_id, nombre, dni, fecha_nacimiento, numero_camiseta, posicion, foto_url, face_encoding, estado)
            VALUES (:eid, :nombre, :dni, :fnac, :cam, :pos, :foto, :face, 'en_revision')
            RETURNING id
        """
        ins_res = await session.execute(text(sql), {
            "eid": equipo_id, "nombre": nombre, "dni": dni, "fnac": fnac,
            "cam": numero_camiseta, "pos": posicion, "foto": foto_url, "face": face_encoding_json
        })
        await session.commit()
        jugador_id = ins_res.fetchone()[0]

        return {"status": "ok", "message": "Jugador registrado y enviado a revisión", "id": jugador_id}

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# ENDPOINTS — JUGADORES / PLANTEL
# ============================================================

@router.get("/{torneo_id}/equipos/{equipo_id}/jugadores", summary="Plantel del equipo")
async def get_jugadores(torneo_id: str, equipo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT tp.id, tp.nombre, tp.dni, tp.fecha_nacimiento, tp.numero_camiseta,
               tp.posicion, tp.foto_url, tp.estado, tp.partidos_jugados,
               tp.amarillas_acum, tp.rojas_acum, tp.email, tp.telefono,
               tp.genero, tp.modalidad, tp.peso_verificado as peso, tp.estatura_verificada as estatura
        FROM torneos.tournament_players tp
        WHERE tp.torneo_equipo_id = CAST(:eid AS UUID)
        ORDER BY tp.numero_camiseta ASC NULLS LAST, tp.nombre ASC
    """), {"eid": equipo_id})
    rows = result.fetchall()
    keys = ["id","nombre","dni","fecha_nacimiento","numero_camiseta",
            "posicion","foto_url","estado","partidos_jugados","amarillas_acum","rojas_acum",
            "email", "telefono", "genero", "modalidad", "peso", "estatura"]
    return [_row_to_dict(keys, r) for r in rows]


async def _get_config_param(param_key: str, default_value, torneo_id: str, session: AsyncSession):
    """Retrieves a configuration parameter hierarchically: Torneo -> Complejo -> Default."""
    t_res = await session.execute(
        text("SELECT config, configuracion, complejo_id FROM torneos.torneos WHERE id = :tid"),
        {"tid": torneo_id}
    )
    t_row = t_res.fetchone()
    if not t_row:
        return default_value
        
    t_config, t_configuracion, complejo_id = t_row
    
    # Try t_config
    if t_config:
        try:
            if isinstance(t_config, str):
                t_config = json.loads(t_config)
            if isinstance(t_config, dict) and param_key in t_config:
                return t_config[param_key]
        except Exception:
            pass
            
    # Try t_configuracion
    if t_configuracion:
        try:
            if isinstance(t_configuracion, str):
                t_configuracion = json.loads(t_configuracion)
            if isinstance(t_configuracion, dict) and param_key in t_configuracion:
                return t_configuracion[param_key]
        except Exception:
            pass

    # 2. Check Complejo's configuracion
    if complejo_id:
        c_res = await session.execute(
            text("SELECT configuracion FROM cancha.complejos WHERE id = :cid"),
            {"cid": complejo_id}
        )
        c_row = c_res.fetchone()
        if c_row and c_row[0]:
            c_config = c_row[0]
            try:
                if isinstance(c_config, str):
                    c_config = json.loads(c_config)
                if isinstance(c_config, dict) and param_key in c_config:
                    return c_config[param_key]
            except Exception:
                pass

    return default_value


async def _add_jugador_logic(
    torneo_id: str, equipo_id: str,
    payload: JugadorCreate,
    session: AsyncSession
) -> dict:
    # 1. Obtener datos del equipo y del torneo
    team_res = await session.execute(
        text("""
            SELECT te.id, t.estado
            FROM torneos.equipos te
            JOIN torneos.torneos t ON te.torneo_id = t.id
            WHERE te.id = CAST(:eid AS UUID) AND te.torneo_id = CAST(:tid AS UUID)
        """),
        {"eid": equipo_id, "tid": torneo_id}
    )
    team_row = team_res.fetchone()
    if not team_row:
        raise HTTPException(status_code=404, detail="Equipo no encontrado en este torneo")
    
    torneo_estado = team_row[1]

    # 2. Bloqueo de adición de jugadores en fases finales (playoffs)
    if torneo_estado in ('playoffs', 'finalizado'):
         raise HTTPException(
             status_code=400,
             detail="Está prohibido incorporar jugadores en fases de eliminación directa / playoffs o cuando el torneo está finalizado."
         )

    # 3. Validar DNI no repetido en otro equipo del mismo torneo
    dup = await session.execute(text("""
        SELECT tp.id FROM torneos.tournament_players tp
        JOIN torneos.equipos te ON tp.torneo_equipo_id = te.id
        WHERE te.torneo_id = CAST(:tid AS UUID) AND tp.dni = :dni
    """), {"tid": torneo_id, "dni": payload.dni})
    if dup.fetchone():
        raise HTTPException(status_code=409, detail=f"El jugador con DNI {payload.dni} ya está inscripto en otro equipo de este torneo")

    # 4. Validar edad y fecha de nacimiento
    fnac = None
    if payload.fecha_nacimiento:
        try:
            fnac = date.fromisoformat(payload.fecha_nacimiento)
        except Exception:
            fnac = None

    # 5. Insertar jugador
    player_uuid = str(uuid.uuid4())
    result = await session.execute(text("""
        INSERT INTO torneos.tournament_players
            (id, torneo_equipo_id, nombre, dni, fecha_nacimiento, numero_camiseta, posicion, foto_url, modalidad, genero, peso_verificado, estatura_verificada)
        VALUES
            (:id, CAST(:eid AS UUID), :nombre, :dni, :fnac, :camiseta, :posicion, :foto_url, :modalidad, :genero, :peso, :estatura)
        RETURNING id, nombre, dni, numero_camiseta, posicion, estado
    """), {
        "id": player_uuid,
        "eid": equipo_id, "nombre": payload.nombre, "dni": payload.dni,
        "fnac": fnac, "camiseta": payload.numero_camiseta,
        "posicion": payload.posicion, "foto_url": payload.foto_url,
        "modalidad": payload.modalidad,
        "genero": payload.genero,
        "peso": payload.peso,
        "estatura": payload.estatura
    })
    await session.commit()
    row = result.fetchone()

    return {"status": "ok", "message": "Jugador agregado exitosamente"}


@router.post("/{torneo_id}/equipos/{equipo_id}/jugadores", summary="Agregar jugador al plantel")
async def add_jugador(
    torneo_id: str, equipo_id: str,
    payload: JugadorCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        return await _add_jugador_logic(torneo_id, equipo_id, payload, session)
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{torneo_id}/equipos/{equipo_id}/jugadores/{jugador_id}", summary="Actualizar jugador")
async def update_jugador(
    torneo_id: str, equipo_id: str, jugador_id: str,
    payload: JugadorUpdate,
    session: AsyncSession = Depends(get_session)
):
    try:
        updates = []
        params: dict = {"pid": jugador_id}
        if payload.nombre is not None:
            updates.append("nombre = :nombre"); params["nombre"] = payload.nombre
        if payload.dni is not None:
            updates.append("dni = :dni"); params["dni"] = payload.dni
        if payload.email is not None:
            updates.append("email = :email"); params["email"] = payload.email
        if payload.telefono is not None:
            updates.append("telefono = :telefono"); params["telefono"] = payload.telefono
        if payload.fecha_nacimiento is not None:
            fnac = None
            if payload.fecha_nacimiento:
                try: fnac = date.fromisoformat(payload.fecha_nacimiento)
                except Exception: pass
            updates.append("fecha_nacimiento = :fnac"); params["fnac"] = fnac
        if payload.numero_camiseta is not None:
            updates.append("numero_camiseta = :camiseta"); params["camiseta"] = payload.numero_camiseta
        if payload.posicion is not None:
            updates.append("posicion = :posicion"); params["posicion"] = payload.posicion
        if payload.estado is not None:
            updates.append("estado = :estado"); params["estado"] = payload.estado
        if payload.modalidad is not None:
            updates.append("modalidad = :modalidad"); params["modalidad"] = payload.modalidad
        if payload.genero is not None:
            updates.append("genero = :genero"); params["genero"] = payload.genero
        if payload.peso is not None:
            updates.append("peso_verificado = :peso"); params["peso"] = payload.peso
        if payload.estatura is not None:
            updates.append("estatura_verificada = :estatura"); params["estatura"] = payload.estatura
        if not updates:
            raise HTTPException(status_code=400, detail="Sin campos para actualizar")

        sql = f"UPDATE torneos.tournament_players SET {', '.join(updates)}, actualizado_en=NOW() WHERE id = :pid RETURNING id, nombre, estado"
        result = await session.execute(text(sql), params)
        await session.commit()
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Jugador no encontrado")
        return _row_to_dict(["id","nombre","estado"], row)
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/jugadores/{jugador_id}/upload-face", summary="Registrar rostro del jugador")
async def upload_face(
    jugador_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        # Validación de tipo de archivo y tamaño aproximado en memoria (seguridad)
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")
        
        file_bytes = await file.read()
        
        # Limitar a ~5MB en memoria para prevenir DoS
        if len(file_bytes) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="El archivo es demasiado grande (máximo 5MB)")

        # Extraer el vector facial
        try:
            encoding = FacialRecognitionService.extract_encoding(file_bytes)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))

        # Serializar y guardar en base de datos
        encoding_json = json.dumps(encoding)
        sql = "UPDATE cancha.tournament_players SET face_encoding = :encoding, actualizado_en=NOW() WHERE id = :pid RETURNING id"
        result = await session.execute(text(sql), {"encoding": encoding_json, "pid": jugador_id})
        await session.commit()
        
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Jugador no encontrado")

        return {"status": "ok", "message": "Identidad verificada y registrada correctamente", "id": jugador_id}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=f"Error procesando el rostro: {str(e)}")

@router.delete("/{torneo_id}/equipos/{equipo_id}/jugadores/{jugador_id}", summary="Eliminar jugador")
async def delete_jugador(
    torneo_id: str, equipo_id: str, jugador_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        await session.execute(text("DELETE FROM torneos.tournament_players WHERE id = :pid AND torneo_equipo_id = :eid"), {"pid": jugador_id, "eid": equipo_id})
        await session.commit()
        return {"status": "ok", "message": "Jugador eliminado"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# ENDPOINTS — EQUIPO TÉCNICO
# ============================================================

@router.get("/{torneo_id}/equipos/{equipo_id}/tecnicos", summary="Cuerpo técnico del equipo")
async def get_tecnicos(torneo_id: str, equipo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT id, nombre, dni, rol, foto_url
        FROM torneos.equipo_tecnico
        WHERE equipo_id = CAST(:eid AS UUID)
        ORDER BY nombre ASC
    """), {"eid": equipo_id})
    return [dict(row._mapping) for row in result.fetchall()]

@router.post("/{torneo_id}/equipos/{equipo_id}/tecnicos", summary="Agregar miembro técnico")
async def add_tecnico(
    torneo_id: str, equipo_id: str,
    payload: EquipoTecnicoCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        tecnico_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.equipo_tecnico (id, equipo_id, nombre, dni, rol, foto_url)
            VALUES (:id, CAST(:eid AS UUID), :nombre, :dni, :rol, :foto_url)
        """), {
            "id": tecnico_id, "eid": equipo_id, "nombre": payload.nombre,
            "dni": payload.dni, "rol": payload.rol, "foto_url": payload.foto_url
        })
        await session.commit()
        return {"status": "ok", "message": "Cuerpo técnico agregado", "id": tecnico_id}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.patch("/{torneo_id}/equipos/{equipo_id}/tecnicos/{tecnico_id}", summary="Actualizar miembro técnico")
async def update_tecnico(
    torneo_id: str, equipo_id: str, tecnico_id: str,
    payload: EquipoTecnicoUpdate,
    session: AsyncSession = Depends(get_session)
):
    try:
        updates = []
        params = {"tid": tecnico_id, "eid": equipo_id}
        if payload.nombre is not None:
            updates.append("nombre = :nombre"); params["nombre"] = payload.nombre
        if payload.dni is not None:
            updates.append("dni = :dni"); params["dni"] = payload.dni
        if payload.rol is not None:
            updates.append("rol = :rol"); params["rol"] = payload.rol
        if payload.foto_url is not None:
            updates.append("foto_url = :foto_url"); params["foto_url"] = payload.foto_url
            
        if not updates:
            raise HTTPException(status_code=400, detail="Sin campos para actualizar")
            
        sql = f"UPDATE torneos.equipo_tecnico SET {', '.join(updates)} WHERE id = CAST(:tid AS UUID) AND equipo_id = CAST(:eid AS UUID)"
        await session.execute(text(sql), params)
        await session.commit()
        return {"status": "ok"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{torneo_id}/equipos/{equipo_id}/tecnicos/{tecnico_id}", summary="Eliminar miembro técnico")
async def delete_tecnico(
    torneo_id: str, equipo_id: str, tecnico_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        await session.execute(text("DELETE FROM torneos.equipo_tecnico WHERE id = CAST(:tid AS UUID) AND equipo_id = CAST(:eid AS UUID)"), {"tid": tecnico_id, "eid": equipo_id})
        await session.commit()
        return {"status": "ok"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/jugadores/test-face", summary="Probar reconocimiento facial libremente")
async def test_face(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    try:
        # Validación de tipo de archivo y tamaño aproximado en memoria
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")
        
        file_bytes = await file.read()
        if len(file_bytes) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="El archivo es demasiado grande (máximo 5MB)")

        # Extraer el vector facial
        try:
            encoding_test = FacialRecognitionService.extract_encoding(file_bytes)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))

        # Consultar todos los jugadores que tengan un encoding registrado
        result = await session.execute(text("""
            SELECT id, nombre, face_encoding 
            FROM cancha.tournament_players 
            WHERE face_encoding IS NOT NULL
        """))
        players = result.fetchall()

        # Búsqueda secuencial
        recognized_player = None
        for row in players:
            try:
                db_encoding = row.face_encoding if isinstance(row.face_encoding, list) else json.loads(row.face_encoding)
                if FacialRecognitionService.compare_encodings(db_encoding, encoding_test):
                    recognized_player = {"id": str(row.id), "nombre": row.nombre}
                    break
            except Exception:
                continue

        if recognized_player:
            return {"status": "ok", "match": True, "message": "¡Rostro reconocido exitosamente!", "jugador": recognized_player}
        else:
            return {"status": "ok", "match": False, "message": "No se encontró coincidencia con ningún jugador registrado"}

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================
# ENDPOINTS — PARTIDOS
# ============================================================

@router.get("/{torneo_id}/partidos", summary="Partidos del torneo")
async def get_partidos(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT p.id, p.equipo_local_id, p.equipo_visitante_id,
               p.goles_local, p.goles_visitante,
               p.fecha_hora, p.estado, p.jornada, p.fase, p.es_wo,
               el.nombre AS local_nombre, el.logo_url AS local_logo,
               ev.nombre AS visitante_nombre, ev.logo_url AS visitante_logo,
               c.nombre AS cancha_nombre,
               p.jugador_local_id, p.jugador_visitante_id, p.estadisticas,
               jl.nombre AS jugador_local_nombre, jv.nombre AS jugador_visitante_nombre
        FROM torneos.partidos p
        JOIN torneos.equipos el ON p.equipo_local_id = el.id
        LEFT JOIN torneos.equipos ev ON p.equipo_visitante_id = ev.id
        LEFT JOIN cancha.canchas c ON p.cancha_id = c.id
        LEFT JOIN torneos.tournament_players jl ON p.jugador_local_id = jl.id
        LEFT JOIN torneos.tournament_players jv ON p.jugador_visitante_id = jv.id
        WHERE p.torneo_id = :tid
        ORDER BY p.jornada ASC NULLS LAST, p.fecha_hora ASC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["id","equipo_local_id","equipo_visitante_id","goles_local","goles_visitante",
            "fecha_hora","estado","jornada","fase","es_wo",
            "local_nombre","local_logo","visitante_nombre","visitante_logo","cancha_nombre",
            "jugador_local_id", "jugador_visitante_id", "estadisticas", 
            "jugador_local_nombre", "jugador_visitante_nombre"]
    return [_row_to_dict(keys, r) for r in rows]

class PartidoManualCreate(BaseModel):
    equipo_local_id: str
    equipo_visitante_id: str
    jugador_local_id: Optional[str] = None
    jugador_visitante_id: Optional[str] = None
    fase: str
    fecha_hora: Optional[str] = None

@router.post("/{torneo_id}/partidos", summary="Crear partido manualmente")
async def create_partido(torneo_id: str, payload: PartidoManualCreate, session: AsyncSession = Depends(get_session)):
    try:
        await session.execute(text("""
            INSERT INTO torneos.partidos 
            (torneo_id, equipo_local_id, equipo_visitante_id, jugador_local_id, jugador_visitante_id, fase, fecha_hora, estado)
            VALUES (:tid, :el, :ev, :jl, :jv, :fase, :fecha, 'programado')
        """), {
            "tid": torneo_id,
            "el": payload.equipo_local_id,
            "ev": payload.equipo_visitante_id,
            "jl": payload.jugador_local_id,
            "jv": payload.jugador_visitante_id,
            "fase": payload.fase,
            "fecha": payload.fecha_hora if payload.fecha_hora else None
        })
        await session.commit()
        return {"status": "ok"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/{torneo_id}/asignaciones", summary="Obtener asignaciones de jugadores por fase")
async def get_asignaciones(torneo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        # Obtener todos los jugadores
        result = await session.execute(text("""
            SELECT tp.id, tp.torneo_equipo_id, tp.nombre, tp.genero, tp.fecha_nacimiento, tp.fase_asignada, tp.modalidad, e.logo_url
            FROM torneos.tournament_players tp
            JOIN torneos.equipos e ON tp.torneo_equipo_id = e.id
            WHERE e.torneo_id = :tid AND tp.estado = 'habilitado'
        """), {"tid": torneo_id})
        players = result.fetchall()

        # Obtener deporte y categorías para inicializar grupos y asignar
        torneo_res = await session.execute(text("SELECT deporte FROM torneos.torneos WHERE id = CAST(:tid AS UUID)"), {"tid": torneo_id})
        deporte = torneo_res.scalar()

        import re
        cat_res = await session.execute(
            text("SELECT id, nombre FROM torneos.categorias WHERE torneo_id = CAST(:tid AS UUID)"),
            {"tid": torneo_id}
        )
        categorias_db = cat_res.fetchall()
        
        parsed_cats = []
        for c in categorias_db:
            name_lower = c.nombre.lower()
            years = re.findall(r'(\d{4})', name_lower)
            min_y = min(int(y) for y in years) if len(years) >= 2 else (int(years[0]) if len(years) == 1 else None)
            max_y = max(int(y) for y in years) if len(years) >= 2 else (int(years[0]) if len(years) == 1 else None)
            gender = None
            if any(x in name_lower for x in ["fem", "mujer", "niña"]): gender = "Femenino"
            elif any(x in name_lower for x in ["masc", "hom", "niño", "varon"]): gender = "Masculino"
            parsed_cats.append({"nombre": c.nombre, "min_y": min_y, "max_y": max_y, "gender": gender})

        # Inicializar todos los grupos posibles en base a las categorías
        grupos = {}
        for c in parsed_cats:
            for g in ["Masculino", "Femenino"]:
                if c["gender"] and c["gender"] != g: continue
                
                fase_name_base = c['nombre']
                name_lower = c['nombre'].lower()
                if g == "Masculino" and not any(x in name_lower for x in ["masc", "hom", "niño", "varon"]):
                    fase_name_base = f"Masculino {fase_name_base}"
                elif g == "Femenino" and not any(x in name_lower for x in ["fem", "mujer", "niña"]):
                    fase_name_base = f"Femenino {fase_name_base}"
                    
                if deporte == "Artes Marciales Mixtas":
                    grupos[f"{fase_name_base} Combate - 1º Fase"] = []
                    grupos[f"{fase_name_base} Formas - 1º Fase"] = []
                else:
                    grupos[f"{fase_name_base} - 1º Fase"] = []

        # Si hay jugadores sin fase_asignada, calcular su fase sugerida
        unassigned = [p for p in players if not p.fase_asignada]
        if unassigned:

            from datetime import date
            for p in unassigned:
                year = p.fecha_nacimiento.year if p.fecha_nacimiento else date.today().year
                g_raw = p.genero or "Masculino"
                genero = "Femenino" if g_raw.lower().startswith("f") else "Masculino"
                
                assigned_fase = "Sin Asignar"
                for c in parsed_cats:
                    if c["gender"] and c["gender"] != genero: continue
                    if c["min_y"] and c["max_y"] and not (c["min_y"] <= year <= c["max_y"]): continue
                    
                    # Generar nombre de fase
                    fase_name_base = c['nombre']
                    name_lower = c['nombre'].lower()
                    if genero == "Masculino" and not any(x in name_lower for x in ["masc", "hom", "niño", "varon"]):
                        fase_name_base = f"Masculino {fase_name_base}"
                    elif genero == "Femenino" and not any(x in name_lower for x in ["fem", "mujer", "niña"]):
                        fase_name_base = f"Femenino {fase_name_base}"
                        
                    if p.modalidad:
                        mod = p.modalidad.strip().capitalize()
                        fase_name_base = f"{fase_name_base} {mod}"
                        
                    assigned_fase = f"{fase_name_base} - 1º Fase"
                    break
                
                # Actualizar DB
                await session.execute(text("""
                    UPDATE torneos.tournament_players SET fase_asignada = :fase WHERE id = :pid
                """), {"fase": assigned_fase, "pid": p.id})

            await session.commit()
            
            # Recargar jugadores
            result = await session.execute(text("""
                SELECT tp.id, tp.torneo_equipo_id, tp.nombre, tp.genero, tp.fecha_nacimiento, tp.fase_asignada, tp.modalidad, e.logo_url
                FROM torneos.tournament_players tp
                JOIN torneos.equipos e ON tp.torneo_equipo_id = e.id
                WHERE e.torneo_id = :tid AND tp.estado = 'habilitado'
            """), {"tid": torneo_id})
            players = result.fetchall()

        # Agrupar por fase
        for p in players:
            fase = p.fase_asignada or "Sin Asignar"
            if fase not in grupos: grupos[fase] = []
            grupos[fase].append({
                "id": str(p.id), "nombre": p.nombre, "genero": p.genero, 
                "fecha_nacimiento": str(p.fecha_nacimiento) if p.fecha_nacimiento else None,
                "torneo_equipo_id": str(p.torneo_equipo_id),
                "logo_url": p.logo_url,
                "modalidad": p.modalidad
            })

        return {"status": "ok", "grupos": grupos}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.patch("/{torneo_id}/asignaciones/{jugador_id}")
async def update_asignacion(torneo_id: str, jugador_id: str, payload: dict, session: AsyncSession = Depends(get_session)):
    try:
        nueva_fase = payload.get("fase_asignada")
        if not nueva_fase:
            raise HTTPException(status_code=400, detail="fase_asignada es requerida")
            
        await session.execute(text("""
            UPDATE torneos.tournament_players SET fase_asignada = :fase WHERE id = :pid
        """), {"fase": nueva_fase, "pid": jugador_id})
        await session.commit()
        return {"status": "ok", "message": "Asignación actualizada"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/{torneo_id}/autoalineacion", summary="Autoalineación de competencias (MMA)")
async def autoalineacion(torneo_id: str, payload: dict, session: AsyncSession = Depends(get_session)):
    try:
        # 1. Obtener todos los jugadores con su fase asignada
        result = await session.execute(text("""
            SELECT tp.id, tp.torneo_equipo_id, tp.nombre, tp.fase_asignada
            FROM torneos.tournament_players tp
            JOIN torneos.equipos e ON tp.torneo_equipo_id = e.id
            WHERE e.torneo_id = :tid AND tp.estado = 'habilitado'
        """), {"tid": torneo_id})
        players = result.fetchall()

        if not players:
            return {"status": "ok", "message": "No hay jugadores habilitados"}

        # Agrupar por fase_asignada
        from collections import defaultdict
        by_fase = defaultdict(list)
        for p in players:
            fase = p.fase_asignada
            if fase and fase != "Sin Asignar":
                by_fase[fase].append(p)

        if not by_fase:
            return {"status": "error", "message": "No hay jugadores asignados a ninguna fase. Primero asigna las categorías."}

        # Obtener los tipos de categoría
        cat_res = await session.execute(text("SELECT nombre, tipo_categoria FROM torneos.categorias WHERE torneo_id = CAST(:tid AS UUID)"), {"tid": torneo_id})
        cat_types = {row.nombre: row.tipo_categoria for row in cat_res.fetchall()}

        # Obtener partidos existentes para no duplicar
        existing_partidos_res = await session.execute(
            text("SELECT id, fase, jugador_local_id FROM torneos.partidos WHERE torneo_id = CAST(:tid AS UUID)"), 
            {"tid": torneo_id}
        )
        existing_partidos = existing_partidos_res.fetchall()
        
        fases_con_partidos = {r.fase for r in existing_partidos}
        formas_jugadores_con_partido = {str(r.jugador_local_id) for r in existing_partidos if r.jugador_local_id}

        import random
        matches_created = 0

        for fase_name, p_list in by_fase.items():
            cat_name = fase_name.split(' - ')[0] if ' - ' in fase_name else fase_name
            tipo = cat_types.get(cat_name, 'combate')

            if tipo == 'formas':
                for p1 in p_list:
                    if str(p1.id) in formas_jugadores_con_partido:
                        continue
                    await session.execute(text("""
                        INSERT INTO torneos.partidos 
                        (torneo_id, equipo_local_id, equipo_visitante_id, jugador_local_id, jugador_visitante_id, fase)
                        VALUES (CAST(:tid AS UUID), CAST(:el AS UUID), NULL, CAST(:jl AS UUID), NULL, :fase)
                    """), {
                        "tid": torneo_id, 
                        "el": p1.torneo_equipo_id, 
                        "jl": p1.id, 
                        "fase": fase_name
                    })
                    matches_created += 1
            else:
                if fase_name in fases_con_partidos:
                    continue
                if len(p_list) < 2:
                    continue
                
                random.shuffle(p_list)
                
                # Generar partidos 1v1
                for i in range(0, len(p_list) - 1, 2):
                    p1 = p_list[i]
                    p2 = p_list[i+1]
                    await session.execute(text("""
                        INSERT INTO torneos.partidos 
                        (torneo_id, equipo_local_id, equipo_visitante_id, jugador_local_id, jugador_visitante_id, fase)
                        VALUES (CAST(:tid AS UUID), CAST(:el AS UUID), CAST(:ev AS UUID), CAST(:jl AS UUID), CAST(:jv AS UUID), :fase)
                    """), {
                        "tid": torneo_id, 
                        "el": p1.torneo_equipo_id, 
                        "ev": p2.torneo_equipo_id,
                        "jl": p1.id, 
                        "jv": p2.id,
                        "fase": fase_name
                    })
                    matches_created += 1

        await session.commit()
        return {"status": "ok", "message": f"Se generaron {matches_created} partidos."}
        
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{torneo_id}/autoalineacion/reset", summary="Eliminar todas las alineaciones y partidos de un torneo")
async def reset_alineacion(torneo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        # Primero eliminar dependencias si existen (eventos, alineaciones, etc.)
        # Hacemos cascada manual en caso de que falte ON DELETE CASCADE
        await session.execute(text("""
            DELETE FROM torneos.eventos_partido 
            WHERE partido_id IN (SELECT id FROM torneos.partidos WHERE torneo_id = :tid)
        """), {"tid": torneo_id})
        
        # Ya no intentamos borrar de alineaciones porque no existe en la base de datos

        await session.execute(text("DELETE FROM torneos.partidos WHERE torneo_id = :tid"), {"tid": torneo_id})
        await session.commit()
        return {"status": "ok", "message": "Todos los partidos han sido eliminados correctamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/partidos/{partido_id}/iniciar", summary="Iniciar partido")
async def iniciar_partido(partido_id: str, session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("""
            UPDATE torneos.partidos
            SET estado = 'en_curso', fecha_hora_inicio_real = NOW()
            WHERE id = :pid AND estado = 'programado'
            RETURNING id, estado
        """), {"pid": partido_id})
        await session.commit()
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Partido no encontrado o ya iniciado")
        return {"id": str(row[0]), "estado": row[1]}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/partidos/{partido_id}", summary="Actualizar resultado de partido")
async def update_partido(partido_id: str, payload: PartidoUpdate, session: AsyncSession = Depends(get_session)):
    try:
        # Obtener IDs para determinar ganador
        p_res = await session.execute(
            text("SELECT torneo_id, equipo_local_id, equipo_visitante_id FROM torneos.partidos WHERE id = :pid"),
            {"pid": partido_id}
        )
        p_row = p_res.fetchone()
        if not p_row:
            raise HTTPException(status_code=404, detail="Partido no encontrado")

        torneo_id = str(p_row[0])
        el_id, ev_id = str(p_row[1]), str(p_row[2])
        ganador_id = None
        if payload.goles_local > payload.goles_visitante:
            ganador_id = el_id
        elif payload.goles_visitante > payload.goles_local:
            ganador_id = ev_id

        import json
        updates = []
        params = {"pid": partido_id}

        if payload.goles_local is not None:
            updates.append("goles_local = :goles_local")
            params["goles_local"] = payload.goles_local
        if payload.goles_visitante is not None:
            updates.append("goles_visitante = :goles_visitante")
            params["goles_visitante"] = payload.goles_visitante
        if payload.estado is not None:
            updates.append("estado = :estado")
            params["estado"] = payload.estado
        if payload.observaciones is not None:
            updates.append("observaciones = :observaciones")
            params["observaciones"] = payload.observaciones
        if payload.cancha is not None:
            updates.append("cancha = :cancha")
            params["cancha"] = payload.cancha
        if payload.fecha_hora is not None:
            updates.append("fecha_hora = :fecha_hora")
            params["fecha_hora"] = payload.fecha_hora
        if payload.equipo_local_id is not None:
            updates.append("equipo_local_id = :equipo_local_id")
            params["equipo_local_id"] = payload.equipo_local_id
        if payload.equipo_visitante_id is not None:
            updates.append("equipo_visitante_id = :equipo_visitante_id")
            params["equipo_visitante_id"] = payload.equipo_visitante_id
        if payload.estadisticas is not None:
            updates.append("estadisticas = :estadisticas")
            params["estadisticas"] = json.dumps(payload.estadisticas)

        # Ganador
        if payload.goles_local is not None and payload.goles_visitante is not None:
            if payload.goles_local > payload.goles_visitante:
                updates.append("ganador_id = :ganador_id")
                params["ganador_id"] = el_id
            elif payload.goles_visitante > payload.goles_local:
                updates.append("ganador_id = :ganador_id")
                params["ganador_id"] = ev_id

        if updates:
            sql_set = ", ".join(updates)
            await session.execute(text(f"""
                UPDATE torneos.partidos
                SET {sql_set}
                WHERE id = :pid
            """), params)
            await session.commit()
        
        # Broadcast via WebSockets si se requiere
        try:
            from main import active_connections
            import asyncio
            msg = json.dumps({
                "type": "SCORE_UPDATE", 
                "partido_id": partido_id, 
                "goles_local": payload.goles_local, 
                "goles_visitante": payload.goles_visitante,
                "estado": payload.estado,
                "estadisticas": payload.estadisticas
            })
            for ws in list(active_connections):
                asyncio.create_task(ws.send_text(msg))
        except Exception as e:
            print("Error broadcasting WebSocket:", e)        # Sincronizar goles con la tabla de goles si el acta fue cargada
        # (solo si no hay goles registrados individualmente)
        goles_registrados = await session.execute(
            text("SELECT COUNT(*) FROM torneos.goles WHERE partido_id = :pid AND NOT anulado"),
            {"pid": partido_id}
        )
        total_goles = goles_registrados.scalar() or 0

        await session.commit()

        # Recalcular posiciones si el partido quedó finalizado
        if payload.estado in ("finalizado", "wo"):
            await _recalcular_posiciones(torneo_id, session)
            await session.commit()
            
            await _avanzar_ronda_eliminatoria(torneo_id, session)
            await _verificar_fin_fase_grupos(torneo_id, session)
            await session.commit()

        return {"id": partido_id, "goles_local": payload.goles_local,
                "goles_visitante": payload.goles_visitante, "estado": payload.estado}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


async def _chequear_descalificacion_wo(equipo_id: str, torneo_id: str, session: AsyncSession):
    """Monitorea los W.O. y descalifica automáticamente si acumula 3 consecutivos o 4 alternados."""
    res = await session.execute(
        text("""
            SELECT id, estado, equipo_wo_id
            FROM torneos.partidos
            WHERE torneo_id = :tid 
              AND (equipo_local_id = :eid OR equipo_visitante_id = :eid)
              AND estado IN ('finalizado', 'wo')
            ORDER BY fecha_hora ASC
        """),
        {"tid": torneo_id, "eid": equipo_id}
    )
    partidos = res.fetchall()
    
    # 1. Contar W.O.s alternados (totales)
    wo_totales = 0
    for pid, estado, eq_wo_id in partidos:
        if estado == 'wo' and eq_wo_id and str(eq_wo_id) == equipo_id:
            wo_totales += 1
            
    # 2. Contar W.O.s consecutivos
    wo_consecutivos = 0
    max_consecutivos = 0
    for pid, estado, eq_wo_id in partidos:
        if estado == 'wo' and eq_wo_id and str(eq_wo_id) == equipo_id:
            wo_consecutivos += 1
            if wo_consecutivos > max_consecutivos:
                max_consecutivos = wo_consecutivos
        else:
            wo_consecutivos = 0
            
    # Disparador de Expulsión: parametrizable por tenant
    consecutivos_limit = await _get_config_param("consecutivos_wo_descalificacion", 3, torneo_id, session)
    alternados_limit = await _get_config_param("alternados_wo_descalificacion", 4, torneo_id, session)

    if max_consecutivos >= consecutivos_limit or wo_totales >= alternados_limit:
        # Cambiar estado del equipo a 'eliminado'
        await session.execute(
            text("""
                UPDATE torneos.equipos
                SET estado_inscripcion = 'eliminado', updated_at = CURRENT_TIMESTAMP
                WHERE id = :eid
            """),
            {"eid": equipo_id}
        )
        
        # Sancionar al equipo dando por perdidos todos sus partidos restantes por 2-0 (marcador WO)
        partidos_restantes_res = await session.execute(
            text("""
                SELECT id, equipo_local_id, equipo_visitante_id
                FROM torneos.partidos
                WHERE torneo_id = :tid
                  AND (equipo_local_id = :eid OR equipo_visitante_id = :eid)
                  AND estado NOT IN ('finalizado', 'wo')
            """),
            {"tid": torneo_id, "eid": equipo_id}
        )
        partidos_restantes = partidos_restantes_res.fetchall()
        
        for p_id, loc_id, vis_id in partidos_restantes:
            p_id = str(p_id)
            loc_id = str(loc_id)
            vis_id = str(vis_id)
            
            if loc_id == equipo_id:
                gl, gv = 0, 2
                ganador_id = vis_id
            else:
                gl, gv = 2, 0
                ganador_id = loc_id
                
            await session.execute(
                text("""
                    UPDATE torneos.partidos
                    SET estado = 'wo', es_wo = true,
                        equipo_wo_id = :infractor_id,
                        goles_local = :gl, goles_visitante = :gv,
                        ganador_id = :ganador_id,
                        fecha_hora_fin_real = CURRENT_TIMESTAMP, acta_cerrada_en = CURRENT_TIMESTAMP
                    WHERE id = :pid
                """),
                {
                    "pid": p_id, "infractor_id": equipo_id,
                    "gl": gl, "gv": gv, "ganador_id": ganador_id
                }
            )
        return True
    return False


@router.post("/partidos/{partido_id}/wo", summary="Declarar W.O.")
async def declarar_wo(partido_id: str, payload: WORequest, session: AsyncSession = Depends(get_session)):
    """Declara walkover: equipo infractor pierde, se aplica marcador estándar."""
    try:
        p_res = await session.execute(
            text("SELECT torneo_id, equipo_local_id, equipo_visitante_id FROM torneos.partidos WHERE id = :pid"),
            {"pid": partido_id}
        )
        p_row = p_res.fetchone()
        if not p_row:
            raise HTTPException(status_code=404, detail="Partido no encontrado")
 
        torneo_id, local_id, visitante_id = str(p_row[0]), str(p_row[1]), str(p_row[2])
 
        # Determinar quién gana
        if str(local_id) == payload.equipo_infractor_id:
            gl, gv = payload.marcador_perdedor, payload.marcador_ganador
            ganador_id = str(visitante_id)
        else:
            gl, gv = payload.marcador_ganador, payload.marcador_perdedor
            ganador_id = str(local_id)
 
        await session.execute(text("""
            UPDATE torneos.partidos
            SET estado = 'wo', es_wo = true,
                equipo_wo_id = :infractor_id,
                goles_local = :gl, goles_visitante = :gv,
                ganador_id = :ganador_id,
                fecha_hora_fin_real = CURRENT_TIMESTAMP, acta_cerrada_en = CURRENT_TIMESTAMP
            WHERE id = :pid
        """), {
            "pid": partido_id, "infractor_id": payload.equipo_infractor_id,
            "gl": gl, "gv": gv, "ganador_id": ganador_id
        })
        await session.commit()
        
        # Chequear si corresponde la descalificación
        fue_descalificado = await _chequear_descalificacion_wo(payload.equipo_infractor_id, torneo_id, session)
        await session.commit()

        await _recalcular_posiciones(torneo_id, session)
        await session.commit()

        await _avanzar_ronda_eliminatoria(torneo_id, session)
        await _verificar_fin_fase_grupos(torneo_id, session)
        await session.commit()
 
        msg = "W.O. declarado correctamente"
        if fue_descalificado:
            msg += ". El equipo infractor acumuló el límite de W.O.s y fue ELIMINADO DEL TORNEO. Se otorgaron victorias automáticas para sus partidos restantes."

        return {"status": "ok", "message": msg,
                "goles_local": gl, "goles_visitante": gv, "ganador_id": ganador_id,
                "eliminado": fue_descalificado}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================
# ENDPOINTS — GOLES
# ============================================================

@router.get("/partidos/{partido_id}/goles", summary="Goles de un partido")
async def get_goles(partido_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT g.id, g.player_id, g.equipo_id, g.minuto, g.tipo, g.anulado,
               tp.nombre AS jugador_nombre, te.nombre AS equipo_nombre
        FROM torneos.goles g
        LEFT JOIN cancha.tournament_players tp ON g.player_id = tp.id
        JOIN torneos.equipos te ON g.equipo_id = te.id
        WHERE g.partido_id = :pid
        ORDER BY g.minuto ASC NULLS LAST, g.creado_en ASC
    """), {"pid": partido_id})
    rows = result.fetchall()
    keys = ["id","player_id","equipo_id","minuto","tipo","anulado","jugador_nombre","equipo_nombre"]
    return [_row_to_dict(keys, r) for r in rows]


@router.post("/partidos/{partido_id}/goles", summary="Registrar gol")
async def add_gol(partido_id: str, payload: GolCreate, session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("""
            INSERT INTO torneos.goles
                (id, partido_id, player_id, equipo_id, minuto, tipo)
            VALUES
                (gen_random_uuid(), :pid, :player_id, :equipo_id, :minuto, :tipo)
            RETURNING id, player_id, equipo_id, minuto, tipo
        """), {
            "pid": partido_id, "player_id": payload.player_id,
            "equipo_id": payload.equipo_id, "minuto": payload.minuto,
            "tipo": payload.tipo
        })

        # Actualizar marcador en el partido
        gol_row = result.fetchone()
        if payload.tipo == "autogol":
            # autogol suma al rival
            await session.execute(text("""
                UPDATE torneos.partidos
                SET goles_local = CASE WHEN equipo_visitante_id = :eid THEN goles_local + 1 ELSE goles_local END,
                    goles_visitante = CASE WHEN equipo_local_id = :eid THEN goles_visitante + 1 ELSE goles_visitante END
                WHERE id = :pid
            """), {"pid": partido_id, "eid": payload.equipo_id})
        else:
            await session.execute(text("""
                UPDATE torneos.partidos
                SET goles_local = CASE WHEN equipo_local_id = :eid THEN goles_local + 1 ELSE goles_local END,
                    goles_visitante = CASE WHEN equipo_visitante_id = :eid THEN goles_visitante + 1 ELSE goles_visitante END
                WHERE id = :pid
            """), {"pid": partido_id, "eid": payload.equipo_id})

        await session.commit()
        return _row_to_dict(["id","player_id","equipo_id","minuto","tipo"], gol_row)
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/partidos/{partido_id}/goles/{gol_id}", summary="Anular gol")
async def anular_gol(partido_id: str, gol_id: str, session: AsyncSession = Depends(get_session)):
    try:
        gol_res = await session.execute(
            text("SELECT equipo_id, tipo FROM torneos.goles WHERE id = :gid AND partido_id = :pid AND NOT anulado"),
            {"gid": gol_id, "pid": partido_id}
        )
        gol_row = gol_res.fetchone()
        if not gol_row:
            raise HTTPException(status_code=404, detail="Gol no encontrado o ya anulado")

        equipo_id, tipo = str(gol_row[0]), gol_row[1]
        await session.execute(
            text("UPDATE torneos.goles SET anulado = true WHERE id = :gid"),
            {"gid": gol_id}
        )
        # Restar del marcador
        if tipo == "autogol":
            await session.execute(text("""
                UPDATE torneos.partidos
                SET goles_local = CASE WHEN equipo_visitante_id = :eid THEN GREATEST(goles_local - 1, 0) ELSE goles_local END,
                    goles_visitante = CASE WHEN equipo_local_id = :eid THEN GREATEST(goles_visitante - 1, 0) ELSE goles_visitante END
                WHERE id = :pid
            """), {"pid": partido_id, "eid": equipo_id})
        else:
            await session.execute(text("""
                UPDATE torneos.partidos
                SET goles_local = CASE WHEN equipo_local_id = :eid THEN GREATEST(goles_local - 1, 0) ELSE goles_local END,
                    goles_visitante = CASE WHEN equipo_visitante_id = :eid THEN GREATEST(goles_visitante - 1, 0) ELSE goles_visitante END
                WHERE id = :pid
            """), {"pid": partido_id, "eid": equipo_id})
        await session.commit()
        return {"status": "ok", "message": "Gol anulado correctamente"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================
# ENDPOINTS — TARJETAS
# ============================================================

@router.get("/partidos/{partido_id}/tarjetas", summary="Tarjetas de un partido")
async def get_tarjetas(partido_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT t.id, t.player_id, t.equipo_id, t.minuto, t.tipo,
               t.pts_fair_play, t.genera_suspension, t.partidos_suspension,
               tp.nombre AS jugador_nombre, te.nombre AS equipo_nombre
        FROM torneos.tarjetas t
        JOIN cancha.tournament_players tp ON t.player_id = tp.id
        JOIN torneos.equipos te ON t.equipo_id = te.id
        WHERE t.partido_id = :pid
        ORDER BY t.minuto ASC NULLS LAST
    """), {"pid": partido_id})
    rows = result.fetchall()
    keys = ["id","player_id","equipo_id","minuto","tipo","pts_fair_play",
            "genera_suspension","partidos_suspension","jugador_nombre","equipo_nombre"]
    return [_row_to_dict(keys, r) for r in rows]


@router.post("/partidos/{partido_id}/tarjetas", summary="Registrar tarjeta")
async def add_tarjeta(partido_id: str, payload: TarjetaCreate, session: AsyncSession = Depends(get_session)):
    try:
        # Obtener torneo_id
        tid_res = await session.execute(
            text("SELECT torneo_id FROM torneos.partidos WHERE id = :pid"),
            {"pid": partido_id}
        )
        tid_row = tid_res.fetchone()
        if not tid_row:
            raise HTTPException(status_code=404, detail="Partido no encontrado")
        torneo_id = str(tid_row[0])

        # Calcular pts fair play y suspensión
        logica = await _aplicar_tarjeta_logica(payload.player_id, payload.tipo, torneo_id, session)

        result = await session.execute(text("""
            INSERT INTO torneos.tarjetas
                (id, partido_id, player_id, equipo_id, minuto, tipo,
                 pts_fair_play, genera_suspension, partidos_suspension)
            VALUES
                (gen_random_uuid(), :pid, :player_id, :equipo_id, :minuto, :tipo,
                 :pts_fp, :genera, :susp)
            RETURNING id, tipo, pts_fair_play, genera_suspension, partidos_suspension
        """), {
            "pid": partido_id, "player_id": payload.player_id,
            "equipo_id": payload.equipo_id, "minuto": payload.minuto,
            "tipo": payload.tipo, "pts_fp": logica["pts_fair_play"],
            "genera": logica["genera_suspension"],
            "susp": logica["partidos_suspension"]
        })
        tarjeta_row = result.fetchone()
        tarjeta_id = str(tarjeta_row[0])

        # Crear sanción si corresponde
        if logica["genera_suspension"]:
            await session.execute(text("""
                INSERT INTO torneos.sanciones
                    (id, torneo_id, player_id, tarjeta_id, tipo, descripcion,
                     partidos_suspension, estado)
                VALUES
                    (gen_random_uuid(), :tid, :player_id, :tarjeta_id, 'suspension',
                     :desc, :susp, 'vigente')
            """), {
                "tid": torneo_id, "player_id": payload.player_id,
                "tarjeta_id": tarjeta_id,
                "desc": f"Suspensión automática por {payload.tipo}",
                "susp": logica["partidos_suspension"]
            })

        # B.7 Integrar cargo automático
        cfg_res = await session.execute(
            text("SELECT multa_amarilla_monto, multa_roja_monto FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        cfg_row = cfg_res.fetchone()
        if cfg_row:
            monto_multa = 0.0
            if payload.tipo == 'amarilla':
                monto_multa = float(cfg_row[0] or 0)
            elif payload.tipo in ['roja', 'doble_amarilla']:
                monto_multa = float(cfg_row[1] or 0)
            
            if monto_multa > 0:
                await session.execute(text("""
                    INSERT INTO cancha.cuenta_corriente_equipos
                    (id, torneo_id, equipo_id, concepto, monto, estado, partido_id)
                    VALUES (gen_random_uuid(), :tid, :eid, :concepto, :monto, 'pendiente', :pid)
                """), {
                    "tid": torneo_id, "eid": payload.equipo_id,
                    "concepto": f"Multa automática por tarjeta {payload.tipo}",
                    "monto": monto_multa, "pid": partido_id
                })

        await session.commit()
        return {
            "id": tarjeta_id, "tipo": payload.tipo,
            "pts_fair_play": logica["pts_fair_play"],
            "genera_suspension": logica["genera_suspension"],
            "partidos_suspension": logica["partidos_suspension"]
        }
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ============================================================
# ENDPOINTS — ESTADÍSTICAS
# ============================================================

@router.get("/{torneo_id}/posiciones", summary="Tabla de posiciones")
async def get_posiciones(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT tp.posicion, te.id AS equipo_id, te.nombre, te.logo_url, te.color_principal,
               tp.pj, tp.pg, tp.pe, tp.pp, tp.gf, tp.gc, tp.dg, tp.pts, tp.pts_fair_play_neg
        FROM torneos.posiciones tp
        JOIN torneos.equipos te ON tp.torneo_equipo_id = te.id
        WHERE tp.torneo_id = :tid
        ORDER BY tp.posicion ASC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["posicion","equipo_id","nombre","logo_url","color_principal",
            "pj","pg","pe","pp","gf","gc","dg","pts","pts_fair_play_neg"]
    return [_row_to_dict(keys, r) for r in rows]


@router.get("/{torneo_id}/goleadores", summary="Tabla de goleadores")
async def get_goleadores(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT tp.id AS player_id, tp.nombre, tp.foto_url, tp.numero_camiseta,
               te.nombre AS equipo_nombre, te.color_principal,
               COUNT(CASE WHEN g.tipo != 'autogol' AND NOT g.anulado THEN 1 END) AS goles,
               COUNT(CASE WHEN g.tipo = 'penal' AND NOT g.anulado THEN 1 END) AS penales,
               COUNT(CASE WHEN g.tipo = 'autogol' AND NOT g.anulado THEN 1 END) AS autogoles
        FROM cancha.tournament_players tp
        JOIN torneos.equipos te ON tp.tournament_team_id = te.id
        LEFT JOIN torneos.goles g ON g.player_id = tp.id
        LEFT JOIN torneos.partidos p ON g.partido_id = p.id
        WHERE te.torneo_id = :tid
        GROUP BY tp.id, tp.nombre, tp.foto_url, tp.numero_camiseta, te.nombre, te.color_principal
        HAVING COUNT(CASE WHEN g.tipo != 'autogol' AND NOT g.anulado THEN 1 END) > 0
        ORDER BY goles DESC, tp.nombre ASC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["player_id","nombre","foto_url","numero_camiseta",
            "equipo_nombre","color_principal","goles","penales","autogoles"]
    return [_row_to_dict(keys, r) for r in rows]


@router.get("/{torneo_id}/fair-play", summary="Tabla de fair play (disciplina)")
async def get_fair_play(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT te.id AS equipo_id, te.nombre, te.logo_url,
               COALESCE(SUM(tt.pts_fair_play), 0) AS pts_negativos,
               COUNT(CASE WHEN tt.tipo = 'amarilla' THEN 1 END) AS amarillas,
               COUNT(CASE WHEN tt.tipo IN ('roja_directa','roja_segunda') THEN 1 END) AS rojas
        FROM torneos.equipos te
        LEFT JOIN torneos.tarjetas tt ON tt.equipo_id = te.id
        LEFT JOIN torneos.partidos p ON tt.partido_id = p.id AND p.torneo_id = :tid
        WHERE te.torneo_id = :tid AND te.estado_inscripcion = 'confirmado'
        GROUP BY te.id, te.nombre, te.logo_url
        ORDER BY pts_negativos ASC, te.nombre ASC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["equipo_id","nombre","logo_url","pts_negativos","amarillas","rojas"]
    return [_row_to_dict(keys, r) for r in rows]


@router.get("/{torneo_id}/sanciones", summary="Suspensiones vigentes del torneo")
async def get_sanciones(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT s.id, s.player_id, tp.nombre AS jugador_nombre,
               tp.numero_camiseta, te.nombre AS equipo_nombre,
               s.tipo, s.descripcion, s.partidos_suspension, s.partidos_cumplidos,
               s.estado, s.creado_en
        FROM torneos.sanciones s
        JOIN cancha.tournament_players tp ON s.player_id = tp.id
        JOIN torneos.equipos te ON tp.tournament_team_id = te.id
        WHERE s.torneo_id = :tid
        ORDER BY s.estado ASC, s.creado_en DESC
    """), {"tid": torneo_id})
    rows = result.fetchall()
    keys = ["id","player_id","jugador_nombre","numero_camiseta","equipo_nombre",
            "tipo","descripcion","partidos_suspension","partidos_cumplidos","estado","creado_en"]
    return [_row_to_dict(keys, r) for r in rows]


@router.post("/sanciones/{sancion_id}/levantar-por-multa", summary="Levantar una sanción registrando pago de multa")
async def levantar_sancion_por_multa(
    sancion_id: str,
    monto: float,
    metodo_pago: str = "Efectivo",
    session: AsyncSession = Depends(get_session)
):
    try:
        # Buscar la sanción
        s_res = await session.execute(
            text("""
                SELECT s.torneo_id, s.player_id, tp.tournament_team_id
                FROM torneos.sanciones s
                JOIN cancha.tournament_players tp ON s.player_id = tp.id
                WHERE s.id = :sid AND s.estado = 'vigente'
            """),
            {"sid": sancion_id}
        )
        s_row = s_res.fetchone()
        if not s_row:
            raise HTTPException(status_code=404, detail="Sanción no encontrada o no está vigente")
        
        torneo_id, player_id, equipo_id = s_row

        # 1. Registrar el cargo como pagado en cuenta corriente
        import uuid
        cargo_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO cancha.cuenta_corriente_equipos
            (id, torneo_id, equipo_id, concepto, monto, estado, creado_en)
            VALUES
            (:id, :tid, :eid, 'Levantamiento de sanción por pago', :monto, 'pagado', NOW())
        """), {
            "id": cargo_id, "tid": str(torneo_id), "eid": str(equipo_id), "monto": monto
        })

        # 2. Levantar la sanción
        await session.execute(text("""
            UPDATE torneos.sanciones
            SET estado = 'levantada_por_pago', partidos_suspension = 0
            WHERE id = :sid
        """), {"sid": sancion_id})

        await session.commit()
        return {"status": "ok", "message": "Sanción levantada correctamente y pago registrado."}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# FIXTURE (mantenido desde main.py pero mejorado)
# ============================================================

def _generar_primera_ronda_eliminatoria(equipos: list) -> tuple[str, list, any]:
    """Genera la primera ronda de bracket de eliminación directa según el número de equipos."""
    m = len(equipos)
    if m <= 2:
        fase = "Final"
    elif m <= 4:
        fase = "Semifinal"
    elif m <= 8:
        fase = "Cuartos de Final"
    elif m <= 16:
        fase = "Octavos de Final"
    else:
        fase = "Dieciseisavos de Final"
        
    partidos = []
    for i in range(m // 2):
        partidos.append((equipos[i], equipos[m - 1 - i]))
        
    bye_team = equipos[m // 2] if m % 2 != 0 else None
    return fase, partidos, bye_team


async def _avanzar_ronda_eliminatoria(torneo_id: str, session: AsyncSession):
    """Verifica si todos los partidos de la fase eliminatoria actual terminaron y genera la siguiente fase."""
    t_res = await session.execute(
        text("SELECT formato, estado, fecha_inicio FROM torneos.torneos WHERE id = :tid"),
        {"tid": torneo_id}
    )
    t_row = t_res.fetchone()
    if not t_row:
        return
    formato, torneo_estado, fecha_inicio = t_row
    if isinstance(fecha_inicio, str):
        fecha_inicio = date.fromisoformat(fecha_inicio)
    
    if formato not in ('eliminatoria', 'mixta'):
        return

    # Buscar partidos activos en fase eliminatoria
    f_res = await session.execute(
        text("""
            SELECT fase, MAX(jornada) 
            FROM torneos.partidos 
            WHERE torneo_id = :tid 
              AND (fase LIKE '%Final%' OR fase LIKE '%Semifinal%')
            GROUP BY fase
        """),
        {"tid": torneo_id}
    )
    fases = f_res.fetchall()
    if not fases:
        return
        
    jerarquia = ["Dieciseisavos de Final", "Octavos de Final", "Cuartos de Final", "Semifinal", "Final"]
    
    fase_actual = None
    max_idx = -1
    max_jornada = 1
    for f, j in fases:
        for idx, name in enumerate(jerarquia):
            if f == name and idx > max_idx:
                max_idx = idx
                fase_actual = name
                max_jornada = j or 1
                
    if not fase_actual:
        return
        
    p_res = await session.execute(
        text("""
            SELECT id, estado, ganador_id, equipo_local_id, equipo_visitante_id, es_wo
            FROM torneos.partidos
            WHERE torneo_id = :tid AND fase = :fase
        """),
        {"tid": torneo_id, "fase": fase_actual}
    )
    partidos = p_res.fetchall()
    if not partidos:
        return
        
    for pid, estado, win_id, el_id, ev_id, es_wo in partidos:
        if estado not in ('finalizado', 'wo'):
            return # Aún quedan partidos por jugar en esta fase
            
    ganadores = []
    for pid, estado, win_id, el_id, ev_id, es_wo in partidos:
        if win_id:
            ganadores.append(str(win_id))
            
    # ¿Hay algún equipo que tuvo BYE?
    eq_res = await session.execute(
        text("""
            SELECT id FROM torneos.equipos 
            WHERE torneo_id = :tid AND estado_inscripcion = 'confirmado'
        """),
        {"tid": torneo_id}
    )
    todos_equipos = [str(r[0]) for r in eq_res.fetchall()]
    
    jugaron = set()
    for pid, estado, win_id, el_id, ev_id, es_wo in partidos:
        if el_id: jugaron.add(str(el_id))
        if ev_id: jugaron.add(str(ev_id))
        
    for eq_id in todos_equipos:
        if eq_id not in jugaron:
            ganadores.append(eq_id)

    idx_actual = jerarquia.index(fase_actual)
    if idx_actual == len(jerarquia) - 1:
        await session.execute(
            text("UPDATE torneos.torneos SET estado = 'finalizado' WHERE id = :tid"),
            {"tid": torneo_id}
        )
        return
        
    siguiente_fase = jerarquia[idx_actual + 1]
    
    from datetime import timedelta, time as dtime
    siguiente_jornada = max_jornada + 1
    fecha_partido = fecha_inicio + timedelta(days=(siguiente_jornada - 1) * 7)
    
    partidos_siguiente = []
    m = len(ganadores)
    for i in range(m // 2):
        partidos_siguiente.append((ganadores[i], ganadores[m - 1 - i]))
        
    total_creados = 0
    for i, (local, visitante) in enumerate(partidos_siguiente):
        hora = dtime(18 + i, 0)
        await session.execute(text("""
            INSERT INTO torneos.partidos
                (id, torneo_id, equipo_local_id, equipo_visitante_id,
                 fase, jornada, numero_partido, fecha_hora, estado)
            VALUES
                (:uuid, :tid, :local, :visitante,
                 :fase, :jornada, :num, :fecha_hora, 'programado')
        """), {
            "uuid": str(uuid.uuid4()),
            "tid": torneo_id, "local": local, "visitante": visitante,
            "fase": siguiente_fase, "jornada": siguiente_jornada,
            "num": i + 1,
            "fecha_hora": datetime.combine(fecha_partido, hora)
        })
        total_creados += 1
        
    if m % 2 != 0:
        bye_team = ganadores[m // 2]
        hora = dtime(18 + total_creados, 0)
        await session.execute(text("""
            INSERT INTO torneos.partidos
                (id, torneo_id, equipo_local_id, equipo_visitante_id,
                 fase, jornada, numero_partido, fecha_hora, estado, es_wo, ganador_id, goles_local, goles_visitante, fecha_hora_fin_real, acta_cerrada_en)
            VALUES
                (:uuid, :tid, :local, NULL,
                 :fase, :jornada, :num, :fecha_hora, 'wo', true, :local, 2, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """), {
            "uuid": str(uuid.uuid4()),
            "tid": torneo_id, "local": bye_team,
            "fase": siguiente_fase, "jornada": siguiente_jornada,
            "num": total_creados + 1,
            "fecha_hora": datetime.combine(fecha_partido, hora)
        })


async def _verificar_fin_fase_grupos(torneo_id: str, session: AsyncSession):
    """Para torneos mixtos, verifica si terminó la fase de grupos y genera las semifinales de playoffs."""
    t_res = await session.execute(
        text("SELECT formato, estado, fecha_inicio FROM torneos.torneos WHERE id = :tid"),
        {"tid": torneo_id}
    )
    t_row = t_res.fetchone()
    if not t_row or t_row[0] != 'mixta':
        return
        
    fecha_inicio = t_row[2]
    if isinstance(fecha_inicio, str):
        fecha_inicio = date.fromisoformat(fecha_inicio)

    playoffs_res = await session.execute(
        text("SELECT COUNT(*) FROM torneos.partidos WHERE torneo_id = :tid AND (fase = 'Semifinal' OR fase = 'Final')"),
        {"tid": torneo_id}
    )
    if playoffs_res.scalar() > 0:
        return

    p_res = await session.execute(
        text("SELECT id, estado, fase FROM torneos.partidos WHERE torneo_id = :tid AND fase LIKE 'Grupo%'"),
        {"tid": torneo_id}
    )
    group_partidos = p_res.fetchall()
    if not group_partidos:
        return
        
    for pid, estado, fase in group_partidos:
        if estado not in ('finalizado', 'wo'):
            return

    eq_res = await session.execute(
        text("SELECT id, nombre FROM torneos.equipos WHERE torneo_id = :tid AND estado_inscripcion = 'confirmado'"),
        {"tid": torneo_id}
    )
    equipos = [str(r[0]) for r in eq_res.fetchall()]
    
    grupo_a_teams = set()
    grupo_b_teams = set()
    for pid, estado, fase in group_partidos:
        eqs_res = await session.execute(
            text("SELECT equipo_local_id, equipo_visitante_id FROM torneos.partidos WHERE id = :pid"),
            {"pid": pid}
        )
        el, ev = eqs_res.fetchone()
        if 'Grupo A' in fase:
            if el: grupo_a_teams.add(str(el))
            if ev: grupo_a_teams.add(str(ev))
        elif 'Grupo B' in fase:
            if el: grupo_b_teams.add(str(el))
            if ev: grupo_b_teams.add(str(ev))

    pos_res = await session.execute(
        text("SELECT torneo_equipo_id, pts, dg, gf FROM torneos.posiciones WHERE torneo_id = :tid"),
        {"tid": torneo_id}
    )
    pos_map = {str(r[0]): (r[1] or 0, r[2] or 0, r[3] or 0) for r in pos_res.fetchall()}

    sorted_a = sorted(
        list(grupo_a_teams),
        key=lambda e: pos_map.get(e, (0, 0, 0)),
        reverse=True
    )
    sorted_b = sorted(
        list(grupo_b_teams),
        key=lambda e: pos_map.get(e, (0, 0, 0)),
        reverse=True
    )

    if len(sorted_a) < 2 or len(sorted_b) < 2:
        return

    a1, a2 = sorted_a[0], sorted_a[1]
    b1, b2 = sorted_b[0], sorted_b[1]

    jornada_max_res = await session.execute(
        text("SELECT MAX(jornada) FROM torneos.partidos WHERE torneo_id = :tid"),
        {"tid": torneo_id}
    )
    jornada_max = (jornada_max_res.scalar() or 0) + 1

    from datetime import timedelta, time as dtime
    fecha_partido = fecha_inicio + timedelta(days=(jornada_max - 1) * 7)
    
    cruces = [(a1, b2), (b1, a2)]
    for i, (local, visitante) in enumerate(cruces):
        hora = dtime(18 + i, 0)
        await session.execute(text("""
            INSERT INTO torneos.partidos
                (id, torneo_id, equipo_local_id, equipo_visitante_id,
                 fase, jornada, numero_partido, fecha_hora, estado)
            VALUES
                (:uuid, :tid, :local, :visitante,
                 'Semifinal', :jornada, :num, :fecha_hora, 'programado')
        """), {
            "uuid": str(uuid.uuid4()),
            "tid": torneo_id, "local": local, "visitante": visitante,
            "jornada": jornada_max, "num": i + 1,
            "fecha_hora": datetime.combine(fecha_partido, hora)
        })
    
    await session.execute(
        text("UPDATE torneos.torneos SET estado = 'playoffs' WHERE id = :tid"),
        {"tid": torneo_id}
    )


def _generar_round_robin(equipos_ids: list, a_dos_vueltas: bool = False) -> list:
    """Algoritmo de Berger — maneja número impar con BYE, e Ida/Vuelta."""
    n = len(equipos_ids)
    if n % 2 != 0:
        equipos_ids = equipos_ids + [None]
        n += 1
    rondas = []
    ids = equipos_ids[:]
    for r in range(n - 1):
        partidos = []
        for i in range(n // 2):
            local, visitante = ids[i], ids[n - 1 - i]
            if local is not None and visitante is not None:
                if r % 2 == 0:
                    partidos.append((local, visitante))
                else:
                    partidos.append((visitante, local))
        ids = [ids[0]] + [ids[-1]] + ids[1:-1]
        rondas.append(partidos)

    if a_dos_vueltas:
        rondas_vuelta = []
        for ronda in rondas:
            partidos_vuelta = [(v, l) for l, v in ronda]
            rondas_vuelta.append(partidos_vuelta)
        rondas.extend(rondas_vuelta)

    return rondas


@router.post("/{torneo_id}/fixture", summary="Generar fixture automático")
async def generar_fixture(torneo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        from datetime import timedelta, time as dtime

        torneo_res = await session.execute(
            text("SELECT id, formato, fecha_inicio, configuracion FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        torneo = torneo_res.fetchone()
        if not torneo:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")

        formato = torneo[1] or 'liga'
        fecha_base = torneo[2]
        if isinstance(fecha_base, str):
            fecha_base = date.fromisoformat(fecha_base)
            
        config_str = torneo[3]
        configuracion = {}
        if config_str:
            if isinstance(config_str, str):
                try: configuracion = json.loads(config_str)
                except: pass
            else: configuracion = config_str
        
        a_dos_vueltas = configuracion.get("a_dos_vueltas", False)
        tipo_sorteo = configuracion.get("tipo_sorteo_playoffs", "random")

        equipos_res = await session.execute(
            text("SELECT id, semilla FROM torneos.equipos WHERE torneo_id = :tid AND estado_inscripcion = 'confirmado'"),
            {"tid": torneo_id}
        )
        equipos_db = equipos_res.fetchall()
        
        if tipo_sorteo == 'random':
            import random
            random.shuffle(equipos_db)
        else:
            equipos_db = sorted(equipos_db, key=lambda x: x[1] or 9999)
            
        equipos = [str(r[0]) for r in equipos_db]

        if len(equipos) < 2:
            raise HTTPException(status_code=400, detail="Se necesitan al menos 2 equipos confirmados")

        # Limpiar fixture anterior
        await session.execute(
            text("DELETE FROM torneos.partidos WHERE torneo_id = :tid"),
            {"tid": torneo_id}
        )

        total = 0
        jornadas_totales = 0

        if formato == 'liga':
            rondas = _generar_round_robin(equipos, a_dos_vueltas=a_dos_vueltas)
            jornadas_totales = len(rondas)
            for round_num, partidos_ronda in enumerate(rondas, start=1):
                fecha_partido = fecha_base + timedelta(days=(round_num - 1) * 7)
                for i, (local, visitante) in enumerate(partidos_ronda):
                    hora = dtime(18 + i, 0)
                    await session.execute(text("""
                        INSERT INTO torneos.partidos
                            (id, torneo_id, equipo_local_id, equipo_visitante_id,
                             fase, jornada, numero_partido, fecha_hora, estado)
                        VALUES
                            (:uuid, :tid, :local, :visitante,
                             :fase, :jornada, :num, :fecha_hora, 'programado')
                    """), {
                        "uuid": str(uuid.uuid4()),
                        "tid": torneo_id, "local": local, "visitante": visitante,
                        "fase": f"Fecha {round_num}", "jornada": round_num,
                        "num": total + 1,
                        "fecha_hora": datetime.combine(fecha_partido, hora)
                    })
                    total += 1

        elif formato == 'eliminatoria':
            fase, partidos_ronda, bye_team = _generar_primera_ronda_eliminatoria(equipos)
            jornadas_totales = 1
            fecha_partido = fecha_base
            for i, (local, visitante) in enumerate(partidos_ronda):
                hora = dtime(18 + i, 0)
                await session.execute(text("""
                    INSERT INTO torneos.partidos
                        (id, torneo_id, equipo_local_id, equipo_visitante_id,
                         fase, jornada, numero_partido, fecha_hora, estado)
                    VALUES
                        (:uuid, :tid, :local, :visitante,
                         :fase, 1, :num, :fecha_hora, 'programado')
                """), {
                    "uuid": str(uuid.uuid4()),
                    "tid": torneo_id, "local": local, "visitante": visitante,
                    "fase": fase, "num": total + 1,
                    "fecha_hora": datetime.combine(fecha_partido, hora)
                })
                total += 1
            
            if bye_team:
                hora = dtime(18 + total, 0)
                await session.execute(text("""
                    INSERT INTO torneos.partidos
                        (id, torneo_id, equipo_local_id, equipo_visitante_id,
                         fase, jornada, numero_partido, fecha_hora, estado, es_wo, ganador_id, goles_local, goles_visitante, fecha_hora_fin_real, acta_cerrada_en)
                    VALUES
                        (:uuid, :tid, :local, NULL,
                         :fase, 1, :num, :fecha_hora, 'wo', true, :local, 2, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """), {
                    "uuid": str(uuid.uuid4()),
                    "tid": torneo_id, "local": bye_team,
                    "fase": fase, "num": total + 1,
                    "fecha_hora": datetime.combine(fecha_partido, hora)
                })
                total += 1

        elif formato == 'mixta':
            grupo_a = [eq for idx, eq in enumerate(equipos) if idx % 2 == 0]
            grupo_b = [eq for idx, eq in enumerate(equipos) if idx % 2 != 0]

            rondas_a = _generar_round_robin(grupo_a, a_dos_vueltas=a_dos_vueltas)
            rondas_b = _generar_round_robin(grupo_b, a_dos_vueltas=a_dos_vueltas)
            
            max_rondas = max(len(rondas_a), len(rondas_b))
            jornadas_totales = max_rondas

            for r_num in range(max_rondas):
                fecha_partido = fecha_base + timedelta(days=r_num * 7)
                i_partido = 0
                
                # Partidos Grupo A
                if r_num < len(rondas_a):
                    for local, visitante in rondas_a[r_num]:
                        hora = dtime(18 + i_partido, 0)
                        await session.execute(text("""
                            INSERT INTO torneos.partidos
                                (id, torneo_id, equipo_local_id, equipo_visitante_id,
                                 fase, jornada, numero_partido, fecha_hora, estado)
                            VALUES
                                (:uuid, :tid, :local, :visitante,
                                 :fase, :jornada, :num, :fecha_hora, 'programado')
                        """), {
                            "uuid": str(uuid.uuid4()),
                            "tid": torneo_id, "local": local, "visitante": visitante,
                            "fase": f"Grupo A - Fecha {r_num + 1}", "jornada": r_num + 1,
                            "num": total + 1,
                            "fecha_hora": datetime.combine(fecha_partido, hora)
                        })
                        total += 1
                        i_partido += 1
                        
                # Partidos Grupo B
                if r_num < len(rondas_b):
                    for local, visitante in rondas_b[r_num]:
                        hora = dtime(18 + i_partido, 0)
                        await session.execute(text("""
                            INSERT INTO torneos.partidos
                                (id, torneo_id, equipo_local_id, equipo_visitante_id,
                                 fase, jornada, numero_partido, fecha_hora, estado)
                            VALUES
                                (:uuid, :tid, :local, :visitante,
                                 :fase, :jornada, :num, :fecha_hora, 'programado')
                        """), {
                            "uuid": str(uuid.uuid4()),
                            "tid": torneo_id, "local": local, "visitante": visitante,
                            "fase": f"Grupo B - Fecha {r_num + 1}", "jornada": r_num + 1,
                            "num": total + 1,
                            "fecha_hora": datetime.combine(fecha_partido, hora)
                        })
                        total += 1
                        i_partido += 1

        elif formato == 'suizo':
            jornadas_totales = 1
            fecha_partido = fecha_base
            m = len(equipos)
            
            partidos_ronda = []
            for i in range(m // 2):
                partidos_ronda.append((equipos[i], equipos[m - 1 - i]))
                
            bye_team = equipos[m // 2] if m % 2 != 0 else None
            
            for i, (local, visitante) in enumerate(partidos_ronda):
                hora = dtime(18 + i, 0)
                await session.execute(text("""
                    INSERT INTO torneos.partidos
                        (id, torneo_id, equipo_local_id, equipo_visitante_id,
                         fase, jornada, numero_partido, fecha_hora, estado)
                    VALUES
                        (:uuid, :tid, :local, :visitante,
                         'Suizo - Ronda 1', 1, :num, :fecha_hora, 'programado')
                """), {
                    "uuid": str(uuid.uuid4()),
                    "tid": torneo_id, "local": local, "visitante": visitante,
                    "num": total + 1,
                    "fecha_hora": datetime.combine(fecha_partido, hora)
                })
                total += 1
                
            if bye_team:
                hora = dtime(18 + total, 0)
                await session.execute(text("""
                    INSERT INTO torneos.partidos
                        (id, torneo_id, equipo_local_id, equipo_visitante_id,
                         fase, jornada, numero_partido, fecha_hora, estado, es_wo, ganador_id, goles_local, goles_visitante, fecha_hora_fin_real, acta_cerrada_en)
                    VALUES
                        (:uuid, :tid, :local, NULL,
                         'Suizo - Ronda 1', 1, :num, :fecha_hora, 'wo', true, :local, 2, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """), {
                    "uuid": str(uuid.uuid4()),
                    "tid": torneo_id, "local": bye_team,
                    "num": total + 1,
                    "fecha_hora": datetime.combine(fecha_partido, hora)
                })
                total += 1

        # Inicializar tabla de posiciones vacía para todos los equipos
        for eid in equipos:
            await session.execute(text("""
                INSERT INTO torneos.posiciones
                    (id, torneo_id, torneo_equipo_id, posicion)
                VALUES
                    (:uuid, :tid, :eid, 0)
                ON CONFLICT (torneo_id, torneo_equipo_id) DO NOTHING
            """), {"uuid": str(uuid.uuid4()), "tid": torneo_id, "eid": eid})

        await session.commit()
        return {"status": "ok", "message": f"Fixture generado: {total} partidos en {jornadas_totales} jornadas"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{torneo_id}/fixture/suizo/siguiente", summary="Generar siguiente ronda del sistema Suizo")
async def generar_siguiente_ronda_suizo(torneo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        from datetime import timedelta, time as dtime

        torneo_res = await session.execute(
            text("SELECT id, formato, fecha_inicio FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        torneo = torneo_res.fetchone()
        if not torneo:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")

        formato, fecha_base = torneo[1], torneo[2]
        if isinstance(fecha_base, str):
            fecha_base = date.fromisoformat(fecha_base)
        if formato != 'suizo':
            raise HTTPException(status_code=400, detail="Este torneo no es de formato suizo")

        ronda_res = await session.execute(
            text("SELECT MAX(jornada) FROM torneos.partidos WHERE torneo_id = :tid"),
            {"tid": torneo_id}
        )
        ronda_actual = ronda_res.scalar()
        if not ronda_actual:
            raise HTTPException(status_code=400, detail="No se ha generado la primera ronda aún")

        activos_res = await session.execute(
            text("""
                SELECT COUNT(*) FROM torneos.partidos 
                WHERE torneo_id = :tid AND jornada = :ronda AND estado NOT IN ('finalizado', 'wo')
            """),
            {"tid": torneo_id, "ronda": ronda_actual}
        )
        if activos_res.scalar() > 0:
            raise HTTPException(status_code=400, detail=f"No se puede avanzar: existen partidos pendientes en la Ronda {ronda_actual}")

        enfrentamientos_res = await session.execute(
            text("""
                SELECT equipo_local_id, equipo_visitante_id FROM torneos.partidos
                WHERE torneo_id = :tid AND equipo_local_id IS NOT NULL AND equipo_visitante_id IS NOT NULL
            """),
            {"tid": torneo_id}
        )
        enfrentamientos = set()
        for loc, vis in enfrentamientos_res.fetchall():
            enfrentamientos.add((str(loc), str(vis)))
            enfrentamientos.add((str(vis), str(loc)))

        equipos_res = await session.execute(
            text("""
                SELECT tp.torneo_equipo_id, tp.pts, tp.dg, tp.gf 
                FROM torneos.posiciones tp
                JOIN torneos.equipos te ON tp.torneo_equipo_id = te.id
                WHERE tp.torneo_id = :tid AND te.estado_inscripcion = 'confirmado'
            """),
            {"tid": torneo_id}
        )
        posiciones = sorted(
            [{"id": str(r[0]), "pts": r[1] or 0, "dg": r[2] or 0, "gf": r[3] or 0} for r in equipos_res.fetchall()],
            key=lambda e: (e["pts"], e["dg"], e["gf"]),
            reverse=True
        )

        n_ronda = ronda_actual + 1
        fecha_partido = fecha_base + timedelta(days=(n_ronda - 1) * 7)

        paired = set()
        partidos_ronda = []
        bye_team = None

        if len(posiciones) % 2 != 0:
            byes_res = await session.execute(
                text("SELECT equipo_local_id FROM torneos.partidos WHERE torneo_id = :tid AND equipo_visitante_id IS NULL"),
                {"tid": torneo_id}
            )
            equipos_con_bye = {str(r[0]) for r in byes_res.fetchall()}
            
            idx_bye = -1
            for idx in range(len(posiciones) - 1, -1, -1):
                tid = posiciones[idx]["id"]
                if tid not in equipos_con_bye:
                    idx_bye = idx
                    break
            if idx_bye == -1:
                idx_bye = len(posiciones) - 1
                
            bye_team = posiciones.pop(idx_bye)["id"]
            paired.add(bye_team)

        for i in range(len(posiciones)):
            t1_id = posiciones[i]["id"]
            if t1_id in paired:
                continue

            t2_id = None
            for j in range(i + 1, len(posiciones)):
                cand_id = posiciones[j]["id"]
                if cand_id in paired:
                    continue
                if (t1_id, cand_id) not in enfrentamientos:
                    t2_id = cand_id
                    break
                    
            if not t2_id:
                for j in range(i + 1, len(posiciones)):
                    cand_id = posiciones[j]["id"]
                    if cand_id not in paired:
                        t2_id = cand_id
                        break

            if t2_id:
                partidos_ronda.append((t1_id, t2_id))
                paired.add(t1_id)
                paired.add(t2_id)

        total_creados = 0
        for idx, (local, visitante) in enumerate(partidos_ronda):
            hora = dtime(18 + idx, 0)
            await session.execute(text("""
                INSERT INTO torneos.partidos
                    (id, torneo_id, equipo_local_id, equipo_visitante_id,
                     fase, jornada, numero_partido, fecha_hora, estado)
                VALUES
                    (:uuid, :tid, :local, :visitante,
                     :fase, :jornada, :num, :fecha_hora, 'programado')
            """), {
                "uuid": str(uuid.uuid4()),
                "tid": torneo_id, "local": local, "visitante": visitante,
                "fase": f"Suizo - Ronda {n_ronda}", "jornada": n_ronda,
                "num": total_creados + 1,
                "fecha_hora": datetime.combine(fecha_partido, hora)
            })
            total_creados += 1

        if bye_team:
            hora = dtime(18 + total_creados, 0)
            await session.execute(text("""
                INSERT INTO torneos.partidos
                    (id, torneo_id, equipo_local_id, equipo_visitante_id,
                     fase, jornada, numero_partido, fecha_hora, estado, es_wo, ganador_id, goles_local, goles_visitante, fecha_hora_fin_real, acta_cerrada_en)
                VALUES
                    (:uuid, :tid, :local, NULL,
                     :fase, :jornada, :num, :fecha_hora, 'wo', true, :local, 2, 0, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """), {
                "uuid": str(uuid.uuid4()),
                "tid": torneo_id, "local": bye_team,
                "fase": f"Suizo - Ronda {n_ronda}", "jornada": n_ronda,
                "num": total_creados + 1,
                "fecha_hora": datetime.combine(fecha_partido, hora)
            })
            total_creados += 1

        await session.commit()
        return {"status": "ok", "message": f"Ronda {n_ronda} del sistema suizo generada con éxito."}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# ENDPOINTS MULTITENANCY 007 — Catálogos y Roles por Complejo
# ============================================================

# ── Schemas nuevos ──────────────────────────────────────────

class ModalidadOut(BaseModel):
    id: int
    codigo: str
    nombre: str
    descripcion: Optional[str] = None

class CategoriaOut(BaseModel):
    id: int
    codigo: str
    nombre: str
    edad_minima: Optional[int] = None
    edad_maxima: Optional[int] = None

class RolComplejoCreate(BaseModel):
    usuario_id: int
    complejo_id: str
    rol: str = "organizador"  # admin_complejo | organizador | veedor

class EventoPartidoCreate(BaseModel):
    player_id: Optional[str] = None
    equipo_id: str
    tipo: str  # GOL | GOL_PENAL | AUTOGOL | AMARILLA | ROJA | ROJA_DIRECTA | DOBLE_AMARILLA | LESION | SUSTITUCION
    tipo_evento_id: Optional[int] = None  # FK a cancha.tipos_evento (se resuelve automáticamente si no se envía)
    minuto: int
    periodo: int = 1
    es_tiempo_adicional: bool = False
    observaciones: Optional[str] = None


# ── CATALOGOS ───────────────────────────────────────────────

@router.get("/catalogos/modalidades", summary="Listar modalidades de torneo")
async def get_modalidades(session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("SELECT id, codigo, nombre, descripcion FROM cancha.modalidades ORDER BY id"))
        rows = result.fetchall()
        return [{"id": r[0], "codigo": r[1], "nombre": r[2], "descripcion": r[3]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo modalidades: {str(e)}")


@router.get("/catalogos/categorias", summary="Listar categorías de torneo")
async def get_categorias(session: AsyncSession = Depends(get_session)):
    try:
        result = await session.execute(text("SELECT id, codigo, nombre, edad_minima, edad_maxima FROM cancha.categorias ORDER BY id"))
        rows = result.fetchall()
        return [{"id": r[0], "codigo": r[1], "nombre": r[2], "edad_minima": r[3], "edad_maxima": r[4]} for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo categorías: {str(e)}")


@router.get("/catalogos/tipos-evento", summary="Listar tipos de evento de partido")
async def get_tipos_evento(session: AsyncSession = Depends(get_session)):
    """Devuelve el catálogo normalizado de tipos de evento (gol, tarjeta, sustitución, etc.)."""
    try:
        result = await session.execute(text("""
            SELECT id, codigo, nombre, descripcion, aplica_a, afecta_marcador, afecta_disciplina
            FROM cancha.tipos_evento
            WHERE activo = TRUE
            ORDER BY id
        """))
        rows = result.fetchall()
        return [
            {
                "id": r[0], "codigo": r[1], "nombre": r[2], "descripcion": r[3],
                "aplica_a": r[4], "afecta_marcador": r[5], "afecta_disciplina": r[6]
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error obteniendo tipos de evento: {str(e)}")


# ── EVENTOS DE PARTIDO (tabla unificada) ─────────────────────────

@router.get("/{torneo_id}/partidos/{partido_id}/eventos", summary="Listar eventos de un partido")
async def get_eventos_partido(
    torneo_id: str,
    partido_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        result = await session.execute(text("""
            SELECT ep.id, ep.tipo, ep.minuto, ep.periodo, ep.es_tiempo_adicional,
                   ep.observaciones, ep.registrado_en,
                   ep.equipo_id,
                   ep.tipo_evento_id,
                   tp_in.nombre  AS jugador_nombre,
                   tp_out.nombre AS jugador_sale_nombre
            FROM torneos.eventos_partido ep
            LEFT JOIN cancha.tournament_players tp_in  ON tp_in.id  = ep.player_id
            LEFT JOIN cancha.tournament_players tp_out ON tp_out.id = ep.player_out_id
            WHERE ep.partido_id = :pid
            ORDER BY ep.minuto, ep.periodo
        """), {"pid": partido_id})
        rows = result.fetchall()
        cols = ["id","tipo","minuto","periodo","es_tiempo_adicional","observaciones",
                "registrado_en","equipo_id","tipo_evento_id","jugador_nombre","jugador_sale_nombre"]
        return [_row_to_dict(cols, r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{torneo_id}/partidos/{partido_id}/eventos", summary="Registrar evento en partido")
async def create_evento_partido(
    torneo_id: str,
    partido_id: str,
    data: EventoPartidoCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        # Verificar que el partido pertenece al torneo
        chk = await session.execute(
            text("SELECT estado FROM torneos.partidos WHERE id = :pid AND torneo_id = :tid"),
            {"pid": partido_id, "tid": torneo_id}
        )
        row = chk.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Partido no encontrado en este torneo")
        if row[0] == "finalizado":
            raise HTTPException(status_code=400, detail="No se pueden agregar eventos a un partido finalizado")

        # Resolver tipo_evento_id si no viene en el payload
        tipo_evento_id = data.tipo_evento_id
        if tipo_evento_id is None:
            te_res = await session.execute(
                text("SELECT id FROM cancha.tipos_evento WHERE codigo = :cod"),
                {"cod": data.tipo.upper()}
            )
            te_row = te_res.fetchone()
            if te_row:
                tipo_evento_id = te_row[0]

        new_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.eventos_partido
                (id, partido_id, player_id, equipo_id, tipo, tipo_evento_id,
                 minuto, periodo, es_tiempo_adicional, observaciones)
            VALUES
                (:id, :pid, :pid_ref, :eid, :tipo, :tipo_evento_id,
                 :min, :per, :ta, :obs)
        """), {
            "id": new_id, "pid": partido_id,
            "pid_ref": data.player_id,
            "eid": data.equipo_id, "tipo": data.tipo,
            "tipo_evento_id": tipo_evento_id,
            "min": data.minuto, "per": data.periodo,
            "ta": data.es_tiempo_adicional, "obs": data.observaciones
        })
        await session.commit()
        return {"status": "ok", "id": new_id, "message": f"Evento '{data.tipo}' registrado en minuto {data.minuto}"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{torneo_id}/partidos/{partido_id}/sustituciones", summary="Registrar sustitución")
async def add_sustitucion(
    torneo_id: str,
    partido_id: str,
    payload: SustitucionCreate,
    session: AsyncSession = Depends(get_session)
):
    """Registra el cambio de un jugador: quién entra (player_id) y quién sale (player_out_id)."""
    try:
        chk = await session.execute(
            text("SELECT estado FROM torneos.partidos WHERE id = :pid AND torneo_id = :tid"),
            {"pid": partido_id, "tid": torneo_id}
        )
        row = chk.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Partido no encontrado en este torneo")
        if row[0] == "finalizado":
            raise HTTPException(status_code=400, detail="No se pueden agregar eventos a un partido finalizado")

        # Obtener id del tipo SUSTITUCION del catálogo
        te_res = await session.execute(
            text("SELECT id FROM cancha.tipos_evento WHERE codigo = 'SUSTITUCION'"),
        )
        te_row = te_res.fetchone()
        tipo_evento_id = te_row[0] if te_row else None

        new_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.eventos_partido
                (id, partido_id, player_id, player_out_id, equipo_id,
                 tipo, tipo_evento_id, minuto, observaciones)
            VALUES
                (:id, :pid, :player_in, :player_out, :eid,
                 'SUSTITUCION', :tipo_evento_id, :min, :obs)
        """), {
            "id": new_id,
            "pid": partido_id,
            "player_in": payload.player_id,
            "player_out": payload.player_out_id,
            "eid": payload.equipo_id,
            "tipo_evento_id": tipo_evento_id,
            "min": payload.minuto,
            "obs": payload.observaciones
        })
        await session.commit()
        return {
            "status": "ok",
            "id": new_id,
            "message": f"Sustitución registrada en minuto {payload.minuto}",
            "player_entra": payload.player_id,
            "player_sale": payload.player_out_id
        }
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{torneo_id}/partidos/{partido_id}/eventos/{evento_id}", summary="Eliminar evento de partido")
@router.delete("/partidos/{partido_id}/eventos/{evento_id}", summary="Eliminar evento de partido (alt)")
async def delete_evento_partido(
    partido_id: str,
    evento_id: str,
    torneo_id: Optional[str] = None,
    session: AsyncSession = Depends(get_session)
):
    try:
        await session.execute(
            text("DELETE FROM torneos.eventos_partido WHERE id = :eid AND partido_id = :pid"),
            {"eid": evento_id, "pid": partido_id}
        )
        await session.commit()
        return {"status": "ok", "message": "Evento eliminado"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


# ── ROLES POR COMPLEJO (Multitenancy) ───────────────────────

@router.get("/complejos/{complejo_id}/roles", summary="Listar usuarios y roles de un complejo")
async def get_roles_complejo(
    complejo_id: str,
    session: AsyncSession = Depends(get_session)
):
    try:
        result = await session.execute(text("""
            SELECT rc.id, rc.usuario_id, rc.rol, rc.activo, rc.creado_en,
                   u.nombre_completo AS usuario_nombre,
                   u.email AS usuario_email
            FROM cancha.roles_complejo rc
            JOIN sistema.usuarios u ON u.id = rc.usuario_id
            WHERE rc.complejo_id = :cid
            ORDER BY rc.rol, u.nombre_completo
        """), {"cid": complejo_id})
        rows = result.fetchall()
        cols = ["id","usuario_id","rol","activo","creado_en","usuario_nombre","usuario_email"]
        return [_row_to_dict(cols, r) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/complejos/{complejo_id}/roles", summary="Asignar rol a usuario en un complejo")
async def create_rol_complejo(
    complejo_id: str,
    data: RolComplejoCreate,
    session: AsyncSession = Depends(get_session)
):
    try:
        new_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO cancha.roles_complejo (id, complejo_id, usuario_id, rol)
            VALUES (:id, :cid, :uid, :rol)
            ON CONFLICT (complejo_id, usuario_id)
            DO UPDATE SET rol = EXCLUDED.rol, activo = TRUE
        """), {"id": new_id, "cid": complejo_id, "uid": data.usuario_id, "rol": data.rol})
        await session.commit()
        return {"status": "ok", "message": f"Rol '{data.rol}' asignado exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/complejos/{complejo_id}/roles/{usuario_id}", summary="Revocar rol de usuario en un complejo")
async def delete_rol_complejo(
    complejo_id: str,
    usuario_id: int,
    session: AsyncSession = Depends(get_session)
):
    try:
        await session.execute(
            text("UPDATE cancha.roles_complejo SET activo = FALSE WHERE complejo_id = :cid AND usuario_id = :uid"),
            {"cid": complejo_id, "uid": usuario_id}
        )
        await session.commit()
        return {"status": "ok", "message": "Rol revocado exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================
# MÓDULO A — CLONACIÓN DE TORNEOS
# ============================================================

class ClonarTorneoRequest(BaseModel):
    nuevo_nombre: Optional[str] = None          # Si None, usa "[nombre original] [COPIA]"
    incluir_equipos: bool = False               # Si True, copia también los equipos (sin jugadores ni pagos)


@router.post("/{torneo_id}/clonar", summary="Clonar un torneo existente (deep copy de configuración)")
async def clonar_torneo(torneo_id: str, payload: ClonarTorneoRequest, session: AsyncSession = Depends(get_session)):
    """
    Duplica la configuración de un torneo pasado para una nueva temporada.
    Copia: nombre, estado inicial (borrador), parámetros de puntos, costo de inscripción,
    max_equipos, modalidad, categoría, reglas, premios y configuración de multas/deuda.
    Opcionalmente copia los equipos (sin jugadores ni historial financiero).
    NO copia: fixture, partidos, goles, tarjetas, sanciones, pagos ni posiciones.
    """
    try:
        # 1. Obtener torneo origen
        res = await session.execute(
            text("""
                SELECT id, complejo_id, organizador_id, modalidad_id, categoria_id, creado_por,
                       nombre, deporte, descripcion, estado,
                       puntos_victoria, puntos_empate, puntos_derrota,
                       max_equipos, costo_inscripcion, es_publico,
                       reglas, premios, configuracion
                FROM torneos.torneos WHERE id = :tid
            """),
            {"tid": torneo_id}
        )
        row = res.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Torneo origen no encontrado")

        cols = ["id", "complejo_id", "organizador_id", "modalidad_id", "categoria_id", "creado_por",
                "nombre", "deporte", "descripcion", "estado",
                "puntos_victoria", "puntos_empate", "puntos_derrota",
                "max_equipos", "costo_inscripcion", "es_publico",
                "reglas", "premios", "configuracion"]
        origen = _row_to_dict(cols, row)

        # 2. Generar nombre del clon
        nombre_clon = payload.nuevo_nombre or f"{origen['nombre']} [COPIA]"
        nuevo_id = str(uuid.uuid4())

        # 3. Insertar torneo clonado (estado siempre 'borrador')
        await session.execute(text("""
            INSERT INTO torneos.torneos
                (id, complejo_id, organizador_id, modalidad_id, categoria_id, creado_por,
                 nombre, deporte, descripcion, estado,
                 puntos_victoria, puntos_empate, puntos_derrota,
                 max_equipos, costo_inscripcion, es_publico,
                 reglas, premios, configuracion)
            VALUES
                (:id, :complejo_id, :organizador_id, :modalidad_id, :categoria_id, :creado_por,
                 :nombre, :deporte, :descripcion, 'borrador',
                 :pts_v, :pts_e, :pts_d,
                 :max_equipos, :costo_inscripcion, :es_publico,
                 :reglas, :premios, :configuracion)
        """), {
            "id": nuevo_id,
            "complejo_id": origen["complejo_id"],
            "organizador_id": origen["organizador_id"],
            "modalidad_id": origen["modalidad_id"],
            "categoria_id": origen["categoria_id"],
            "creado_por": origen["creado_por"],
            "nombre": nombre_clon,
            "deporte": origen.get("deporte", "Fútbol 5"),
            "descripcion": origen.get("descripcion"),
            "pts_v": origen["puntos_victoria"],
            "pts_e": origen["puntos_empate"],
            "pts_d": origen["puntos_derrota"],
            "max_equipos": origen["max_equipos"],
            "costo_inscripcion": origen["costo_inscripcion"],
            "es_publico": origen["es_publico"],
            "reglas": json.dumps(origen["reglas"]) if origen.get("reglas") else None,
            "premios": json.dumps(origen["premios"]) if origen.get("premios") else None,
            "configuracion": json.dumps(origen["configuracion"]) if origen.get("configuracion") else None,
        })

        equipos_copiados = 0

        # 4. (Opcional) Copiar equipos sin jugadores ni pagos
        if payload.incluir_equipos:
            eq_res = await session.execute(
                text("""
                    SELECT nombre, logo_url, color_principal, color_secundario,
                           capitan_nombre, capitan_telefono, capitan_email, promocion
                    FROM torneos.equipos
                    WHERE torneo_id = :tid AND estado_inscripcion != 'descalificado'
                """),
                {"tid": torneo_id}
            )
            equipos_origen = eq_res.fetchall()
            for eq in equipos_origen:
                eq_id = str(uuid.uuid4())
                await session.execute(text("""
                    INSERT INTO torneos.equipos
                        (id, torneo_id, nombre, logo_url, color_principal, color_secundario,
                         capitan_nombre, capitan_telefono, capitan_email, promocion,
                         estado_inscripcion, payment_status)
                    VALUES
                        (:id, :torneo_id, :nombre, :logo_url, :cp, :cs,
                         :cap_n, :cap_t, :cap_e, :promo,
                         'pendiente', 'pending')
                """), {
                    "id": eq_id, "torneo_id": nuevo_id,
                    "nombre": eq[0], "logo_url": eq[1], "cp": eq[2], "cs": eq[3],
                    "cap_n": eq[4], "cap_t": eq[5], "cap_e": eq[6], "promo": eq[7] or 0,
                })
                equipos_copiados += 1

        await session.commit()

        return {
            "status": "ok",
            "torneo_id": nuevo_id,
            "nombre": nombre_clon,
            "equipos_copiados": equipos_copiados,
            "message": f"Torneo clonado exitosamente como '{nombre_clon}'"
        }

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# MÓDULO A — EXPORTACIÓN XLSX
# ============================================================

@router.get("/{torneo_id}/exportar/xlsx", summary="Exportar torneo completo a Excel (.xlsx)")
async def exportar_torneo_xlsx(torneo_id: str, session: AsyncSession = Depends(get_session)):
    """
    Genera y descarga un archivo .xlsx con 5 hojas:
    - Hoja 1: Equipos (nombre, capitán, estado inscripción, pago)
    - Hoja 2: Planteles / Jugadores (equipo, jugador, DNI, camiseta, estado)
    - Hoja 3: Fixture (jornada, fecha, local vs visitante, marcador, estado)
    - Hoja 4: Tabla de Posiciones (pos, equipo, PJ, PG, PE, PP, GF, GC, DG, PTS)
    - Hoja 5: Fair Play (equipo, amarillas, rojas, pts_disciplina)
    """
    try:
        import io
        from fastapi.responses import StreamingResponse
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no está instalado. Ejecute: pip install openpyxl")

        # Verificar que el torneo existe
        t_res = await session.execute(
            text("SELECT nombre FROM torneos.torneos WHERE id = :tid"),
            {"tid": torneo_id}
        )
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        torneo_nombre = t_row[0]

        # ── Helpers de estilo ──
        def _header_row(ws, headers: list, fill_hex: str = "1e3a5f"):
            fill = PatternFill("solid", fgColor=fill_hex)
            font = Font(bold=True, color="FFFFFF", size=11)
            border_side = Side(style="thin", color="2d4a6e")
            border = Border(bottom=border_side)
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[1].height = 22

        def _auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        def _style_data_rows(ws, start_row: int = 2):
            alt_fill = PatternFill("solid", fgColor="0f1e35")
            for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
                for cell in row:
                    cell.alignment = Alignment(vertical="center")
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

        # ── Crear libro ──
        wb = Workbook()

        # ─────────────────────────────────────────────
        # HOJA 1: EQUIPOS
        # ─────────────────────────────────────────────
        ws_eq = wb.active
        ws_eq.title = "Equipos"
        _header_row(ws_eq, ["#", "Equipo", "Capitán", "Teléfono", "Estado Inscripción", "Estado Pago", "Promoción"])
        eq_res = await session.execute(text("""
            SELECT nombre, capitan_nombre, capitan_telefono, estado_inscripcion, payment_status, promocion
            FROM torneos.equipos
            WHERE torneo_id = :tid
            ORDER BY nombre
        """), {"tid": torneo_id})
        for i, eq in enumerate(eq_res.fetchall(), 1):
            ws_eq.append([i, eq[0], eq[1] or "—", eq[2] or "—", eq[3], eq[4] or "—", eq[5] or 0])
        _style_data_rows(ws_eq)
        _auto_width(ws_eq)

        # ─────────────────────────────────────────────
        # HOJA 2: PLANTELES / JUGADORES
        # ─────────────────────────────────────────────
        ws_pl = wb.create_sheet("Planteles")
        _header_row(ws_pl, ["#", "Equipo", "Jugador", "DNI", "Camiseta", "Posición", "Estado", "Año Egreso"])
        pl_res = await session.execute(text("""
            SELECT te.nombre as equipo, tp.nombre, tp.dni, tp.numero_camiseta,
                   tp.posicion, tp.estado, tp.egreso_ano
            FROM torneos.tournament_players tp
            JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            WHERE te.torneo_id = :tid
            ORDER BY te.nombre, tp.nombre
        """), {"tid": torneo_id})
        for i, pl in enumerate(pl_res.fetchall(), 1):
            ws_pl.append([i, pl[0], pl[1], pl[2], pl[3] or "—", pl[4] or "—", pl[5], pl[6] or "—"])
        _style_data_rows(ws_pl)
        _auto_width(ws_pl)

        # ─────────────────────────────────────────────
        # HOJA 3: FIXTURE
        # ─────────────────────────────────────────────
        ws_fx = wb.create_sheet("Fixture")
        _header_row(ws_fx, ["Jornada", "Fase", "Fecha / Hora", "Local", "Goles L", "Goles V", "Visitante", "Estado"])
        fx_res = await session.execute(text("""
            SELECT tp.jornada, tp.fase,
                   TO_CHAR(tp.fecha_hora AT TIME ZONE 'America/Asuncion', 'DD/MM/YYYY HH24:MI'),
                   el.nombre, tp.goles_local, tp.goles_visitante, ev.nombre, tp.estado
            FROM torneos.partidos tp
            JOIN torneos.equipos el ON el.id = tp.equipo_local_id
            JOIN torneos.equipos ev ON ev.id = tp.equipo_visitante_id
            WHERE tp.torneo_id = :tid
            ORDER BY tp.jornada, tp.fecha_hora
        """), {"tid": torneo_id})
        for p in fx_res.fetchall():
            ws_fx.append([p[0], p[1] or "—", p[2] or "—", p[3],
                          p[4] if p[4] is not None else "—",
                          p[5] if p[5] is not None else "—",
                          p[6], p[7]])
        _style_data_rows(ws_fx)
        _auto_width(ws_fx)

        # ─────────────────────────────────────────────
        # HOJA 4: TABLA DE POSICIONES
        # ─────────────────────────────────────────────
        ws_pos = wb.create_sheet("Posiciones")
        _header_row(ws_pos, ["Pos", "Equipo", "PJ", "PG", "PE", "PP", "GF", "GC", "DG", "PTS"])
        pos_res = await session.execute(text("""
            SELECT po.posicion, te.nombre,
                   po.pj, po.pg, po.pe, po.pp, po.gf, po.gc, (po.gf - po.gc), po.pts
            FROM torneos.posiciones po
            JOIN torneos.equipos te ON te.id = po.equipo_id
            WHERE po.torneo_id = :tid
            ORDER BY po.posicion
        """), {"tid": torneo_id})
        for p in pos_res.fetchall():
            ws_pos.append(list(p))
        _style_data_rows(ws_pos)
        _auto_width(ws_pos)

        # ─────────────────────────────────────────────
        # HOJA 5: FAIR PLAY
        # ─────────────────────────────────────────────
        ws_fp = wb.create_sheet("Fair Play")
        _header_row(ws_fp, ["#", "Equipo", "Amarillas", "Rojas", "Doble Amarilla", "Pts Disciplina (desc)"])
        fp_res = await session.execute(text("""
            SELECT te.nombre,
                   COUNT(*) FILTER (WHERE tt.tipo IN ('amarilla')) as amarillas,
                   COUNT(*) FILTER (WHERE tt.tipo IN ('roja_directa')) as rojas,
                   COUNT(*) FILTER (WHERE tt.tipo = 'roja_segunda') as doble_amarilla,
                   COALESCE(SUM(tt.pts_fair_play), 0) as total_pts_fp
            FROM torneos.tarjetas tt
            JOIN torneos.partidos tp ON tt.partido_id = tp.id
            JOIN torneos.equipos te ON te.id = tt.equipo_id
            WHERE tp.torneo_id = :tid
            GROUP BY te.nombre
            ORDER BY total_pts_fp DESC
        """), {"tid": torneo_id})
        for i, fp in enumerate(fp_res.fetchall(), 1):
            ws_fp.append([i, fp[0], fp[1], fp[2], fp[3], fp[4]])
        _style_data_rows(ws_fp)
        _auto_width(ws_fp)

        # ── Serializar a bytes y devolver como descarga ──
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"Torneo_{torneo_nombre.replace(' ', '_')}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ============================================================
# ENDPOINTS — CUENTA CORRIENTE (MÓDULO B)
# ============================================================

class CargoCreate(BaseModel):
    torneo_id: str
    concepto: str
    monto: float
    partido_id: Optional[str] = None

@router.get("/equipos/{equipo_id}/cuenta_corriente", summary="Obtener cuenta corriente de un equipo")
async def get_cuenta_corriente(equipo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        # Obtener los movimientos
        result = await session.execute(
            text("""
                SELECT id, concepto, monto, estado, creado_en, partido_id, referencia_pago_id
                FROM cancha.cuenta_corriente_equipos
                WHERE equipo_id = :eid
                ORDER BY creado_en DESC
            """),
            {"eid": equipo_id}
        )
        movimientos = []
        deuda_total = 0.0
        
        for row in result.fetchall():
            m = _row_to_dict(["id", "concepto", "monto", "estado", "creado_en", "partido_id", "referencia_pago_id"], row)
            movimientos.append(m)
            if m["estado"] == "pendiente":
                deuda_total += float(m["monto"])
                
        # Obtener configuración del torneo para saber los límites
        # Necesitamos el torneo_id del equipo
        t_res = await session.execute(
            text("SELECT torneo_id FROM torneos.equipos WHERE id = :eid"),
            {"eid": equipo_id}
        )
        torneo_id = t_res.scalar()
        
        torneo_cfg = {"limite_deuda_habilitado": False, "limite_deuda_monto": 0.0}
        if torneo_id:
            cfg_res = await session.execute(
                text("SELECT limite_deuda_habilitado, limite_deuda_monto FROM torneos.torneos WHERE id = :tid"),
                {"tid": str(torneo_id)}
            )
            cfg_row = cfg_res.fetchone()
            if cfg_row:
                torneo_cfg["limite_deuda_habilitado"] = cfg_row[0]
                torneo_cfg["limite_deuda_monto"] = float(cfg_row[1] or 0)
                
        bloqueado = False
        if torneo_cfg["limite_deuda_habilitado"] and deuda_total > torneo_cfg["limite_deuda_monto"]:
            bloqueado = True

        return {
            "equipo_id": equipo_id,
            "torneo_id": torneo_id,
            "deuda_total": deuda_total,
            "bloqueado": bloqueado,
            "limite_habilitado": torneo_cfg["limite_deuda_habilitado"],
            "limite_monto": torneo_cfg["limite_deuda_monto"],
            "movimientos": movimientos
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/equipos/{equipo_id}/cuenta_corriente/cargos", summary="Agregar un cargo manual a la cuenta corriente")
async def add_cargo_manual(equipo_id: str, payload: CargoCreate, session: AsyncSession = Depends(get_session)):
    try:
        cargo_id = str(uuid.uuid4())
        await session.execute(
            text("""
                INSERT INTO cancha.cuenta_corriente_equipos
                (id, torneo_id, equipo_id, concepto, monto, estado, partido_id)
                VALUES (:id, :tid, :eid, :concepto, :monto, 'pendiente', :pid)
            """),
            {
                "id": cargo_id,
                "tid": payload.torneo_id,
                "eid": equipo_id,
                "concepto": payload.concepto,
                "monto": payload.monto,
                "pid": payload.partido_id
            }
        )
        await session.commit()
        return {"status": "ok", "cargo_id": cargo_id}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/equipos/{equipo_id}/cuenta_corriente/{cargo_id}/pagar", summary="Registrar el pago de un cargo")
async def pagar_cargo(equipo_id: str, cargo_id: str, session: AsyncSession = Depends(get_session)):
    try:
        # Verificar que exista y este pendiente
        check = await session.execute(
            text("SELECT estado FROM cancha.cuenta_corriente_equipos WHERE id = :cid AND equipo_id = :eid"),
            {"cid": cargo_id, "eid": equipo_id}
        )
        row = check.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Cargo no encontrado")
        if row[0] != 'pendiente':
            raise HTTPException(status_code=400, detail="El cargo ya se encuentra pagado")
            
        await session.execute(
            text("UPDATE cancha.cuenta_corriente_equipos SET estado = 'pagado' WHERE id = :cid"),
            {"cid": cargo_id}
        )
        await session.commit()
        return {"status": "ok", "mensaje": "Cargo marcado como pagado"}
    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# ENDPOINTS — ASISTENCIA (CHECK-IN FACIAL)
# ============================================================

@router.post("/checkin-torneo/{torneo_id}", summary="Check-in facial al torneo (venue)")
async def checkin_torneo_facial(
    torneo_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    """
    Recibe foto de un jugador y lo identifica contra todos los jugadores del torneo.
    Si lo reconoce, registra su asistencia al torneo (una sola vez por día).
    """
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")
        file_bytes = await file.read()

        # Extraer encoding de la foto enviada
        try:
            encoding_test = FacialRecognitionService.extract_encoding(file_bytes)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))

        # Traer todos los jugadores del torneo que tienen face_encoding
        result = await session.execute(text("""
            SELECT tp.id, tp.nombre, tp.face_encoding
            FROM cancha.tournament_players tp
            INNER JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            WHERE te.torneo_id = :torneo_id AND tp.face_encoding IS NOT NULL
        """), {"torneo_id": torneo_id})
        players = result.fetchall()

        recognized = None
        for row in players:
            try:
                db_enc = row.face_encoding if isinstance(row.face_encoding, list) else json.loads(row.face_encoding)
                if FacialRecognitionService.compare_encodings(db_enc, encoding_test):
                    recognized = {"id": str(row.id), "nombre": row.nombre}
                    break
            except Exception:
                continue

        if not recognized:
            return {"status": "no_match", "match": False, "message": "No se reconoció el rostro en este torneo"}

        # Registrar asistencia (ignorar duplicado por UNIQUE)
        try:
            await session.execute(text("""
                INSERT INTO cancha.asistencia_torneo (jugador_id, torneo_id, metodo)
                VALUES (:jid, :tid, 'facial')
                ON CONFLICT (jugador_id, torneo_id) DO NOTHING
            """), {"jid": recognized["id"], "tid": torneo_id})
            await session.commit()
        except Exception:
            await session.rollback()

        return {"status": "ok", "match": True, "jugador": recognized,
                "message": f"✅ ¡Bienvenido, {recognized['nombre']}! Asistencia registrada al torneo."}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/checkin-partido/{partido_id}", summary="Check-in facial por partido")
async def checkin_partido_facial(
    partido_id: str,
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_session)
):
    """
    Identifica al jugador y registra su participación en un partido específico.
    Primero busca en los equipos que juegan ese partido.
    """
    try:
        if not file.content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="El archivo debe ser una imagen")
        file_bytes = await file.read()

        try:
            encoding_test = FacialRecognitionService.extract_encoding(file_bytes)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))

        # Obtener info del partido (equipos y torneo)
        partido_res = await session.execute(text("""
            SELECT equipo_local_id, equipo_visitante_id, torneo_id
            FROM torneos.partidos WHERE id = :pid
        """), {"pid": partido_id})
        partido = partido_res.fetchone()
        if not partido:
            raise HTTPException(status_code=404, detail="Partido no encontrado")

        equipo_ids = [str(partido.equipo_local_id), str(partido.equipo_visitante_id)]
        torneo_id = str(partido.torneo_id)

        # Traer jugadores de ambos equipos con encoding
        result = await session.execute(text("""
            SELECT tp.id, tp.nombre, tp.face_encoding, te.nombre AS equipo_nombre
            FROM cancha.tournament_players tp
            INNER JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            WHERE te.id = ANY(:eids) AND tp.face_encoding IS NOT NULL
        """), {"eids": equipo_ids})
        players = result.fetchall()

        recognized = None
        for row in players:
            try:
                db_enc = row.face_encoding if isinstance(row.face_encoding, list) else json.loads(row.face_encoding)
                if FacialRecognitionService.compare_encodings(db_enc, encoding_test):
                    recognized = {"id": str(row.id), "nombre": row.nombre, "equipo": row.equipo_nombre}
                    break
            except Exception:
                continue

        if not recognized:
            return {"status": "no_match", "match": False, "message": "No se reconoció al jugador en este partido"}

        # Registrar asistencia al partido
        try:
            await session.execute(text("""
                INSERT INTO cancha.asistencia_partido (jugador_id, partido_id, torneo_id, metodo)
                VALUES (:jid, :pid, :tid, 'facial')
                ON CONFLICT (jugador_id, partido_id) DO NOTHING
            """), {"jid": recognized["id"], "pid": partido_id, "tid": torneo_id})
            await session.commit()
        except Exception:
            await session.rollback()

        return {"status": "ok", "match": True, "jugador": recognized,
                "message": f"✅ {recognized['nombre']} ({recognized['equipo']}) registrado en el partido."}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/asistencia-torneo/{torneo_id}", summary="Listar asistencias al torneo")
async def get_asistencias_torneo(torneo_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT at.id, at.jugador_id, at.fecha_hora, at.metodo,
               tp.nombre AS jugador_nombre, te.nombre AS equipo_nombre
        FROM cancha.asistencia_torneo at
        INNER JOIN cancha.tournament_players tp ON tp.id = at.jugador_id
        INNER JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
        WHERE at.torneo_id = :tid
        ORDER BY at.fecha_hora DESC
    """), {"tid": torneo_id})
    return [dict(r._mapping) for r in result.fetchall()]


@router.get("/asistencia-partido/{partido_id}", summary="Listar asistencias de un partido")
async def get_asistencias_partido(partido_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT ap.id, ap.jugador_id, ap.fecha_hora, ap.metodo,
               tp.nombre AS jugador_nombre, te.nombre AS equipo_nombre
        FROM cancha.asistencia_partido ap
        INNER JOIN cancha.tournament_players tp ON tp.id = ap.jugador_id
        INNER JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
        WHERE ap.partido_id = :pid
        ORDER BY ap.fecha_hora DESC
    """), {"pid": partido_id})
    return [dict(r._mapping) for r in result.fetchall()]


# ============================================================
# ENDPOINTS — JERARQUIA REGIONAL Y PLAYOFFS
# ============================================================

@router.post("/eventos/{evento_id}/regiones", summary="Crear Región")
async def create_region(evento_id: str, payload: RegionCreate, session: AsyncSession = Depends(get_session)):
    try:
        region_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.regiones (id, evento_id, nombre, determinar_campeon_regional)
            VALUES (:id, :eid, :nombre, :dcr)
        """), {"id": region_id, "eid": evento_id, "nombre": payload.nombre, "dcr": payload.determinar_campeon_regional})
        await session.commit()
        return {"id": region_id, "mensaje": "Región creada exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/eventos/{evento_id}/regiones", summary="Listar Regiones del Evento")
async def get_regiones(evento_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT id, nombre, determinar_campeon_regional, creado_en 
        FROM torneos.regiones 
        WHERE evento_id = :eid
    """), {"eid": evento_id})
    return [dict(r._mapping) for r in result.fetchall()]


@router.post("/regiones/{region_id}/ciudades", summary="Crear Ciudad")
async def create_ciudad(region_id: str, payload: CiudadCreate, session: AsyncSession = Depends(get_session)):
    try:
        ciudad_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.ciudades (id, region_id, nombre)
            VALUES (:id, :rid, :nombre)
        """), {"id": ciudad_id, "rid": region_id, "nombre": payload.nombre})
        await session.commit()
        return {"id": ciudad_id, "mensaje": "Ciudad creada exitosamente"}
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/regiones/{region_id}/ciudades", summary="Listar Ciudades de una Región")
async def get_ciudades(region_id: str, session: AsyncSession = Depends(get_session)):
    result = await session.execute(text("""
        SELECT id, nombre, creado_en 
        FROM torneos.ciudades 
        WHERE region_id = :rid
    """), {"rid": region_id})
    return [dict(r._mapping) for r in result.fetchall()]


@router.post("/regiones/{region_id}/generar-playoff-regional", summary="Generar Campeonato de Campeones Regionales")
async def generar_playoff_regional(region_id: str, payload: PlayoffRegionalCreate, session: AsyncSession = Depends(get_session)):
    try:
        # 1. Validar que la región permita campeón regional
        r_res = await session.execute(text("""
            SELECT evento_id, nombre, determinar_campeon_regional 
            FROM torneos.regiones WHERE id = :rid
        """), {"rid": region_id})
        r_row = r_res.fetchone()
        if not r_row:
            raise HTTPException(status_code=404, detail="Región no encontrada")
        if not r_row.determinar_campeon_regional:
            raise HTTPException(status_code=400, detail="Esta región no permite generar un Playoff Regional")
        
        evento_id = str(r_row.evento_id)
        region_nombre = r_row.nombre

        # 2. Buscar todas las ciudades de esta región
        c_res = await session.execute(text("SELECT id FROM torneos.ciudades WHERE region_id = :rid"), {"rid": region_id})
        ciudades_ids = [str(r[0]) for r in c_res.fetchall()]
        if not ciudades_ids:
            raise HTTPException(status_code=400, detail="No hay ciudades en esta región")

        # 3. Buscar los torneos asociados a esas ciudades
        t_res = await session.execute(text("""
            SELECT id FROM torneos.torneos WHERE ciudad_id = ANY(:cids) AND estado = 'finalizado'
        """), {"cids": ciudades_ids})
        torneos_ids = [str(r[0]) for r in t_res.fetchall()]
        
        if not torneos_ids:
            raise HTTPException(status_code=400, detail="No hay campeonatos locales finalizados en esta región para clonar")

        # 4. Obtener los mejores N equipos de cada campeonato
        equipos_clasificados = []
        for tid in torneos_ids:
            eq_res = await session.execute(text("""
                SELECT equipo_id FROM torneos.posiciones
                WHERE torneo_id = :tid
                ORDER BY pts DESC, (gf - gc) DESC
                LIMIT :limite
            """), {"tid": tid, "limite": payload.cupos_por_ciudad})
            equipos_clasificados.extend([str(r[0]) for r in eq_res.fetchall()])

        if not equipos_clasificados:
            raise HTTPException(status_code=400, detail="No se encontraron equipos clasificados.")

        # 5. Crear el nuevo "Campeonato de Campeones Regionales"
        nuevo_torneo_id = str(uuid.uuid4())
        await session.execute(text("""
            INSERT INTO torneos.torneos 
            (id, evento_id, nombre, deporte, formato, fecha_inicio, estado, configuracion)
            VALUES 
            (:id, :eid, :nombre, 'Futbol', 'eliminacion_simple', CURRENT_DATE, 'abierto', '{}'::jsonb)
        """), {"id": nuevo_torneo_id, "eid": evento_id, "nombre": f"Playoff Regional - {region_nombre}"})

        # 6. Clonar Equipos y Jugadores
        for eid in equipos_clasificados:
            # Obtener datos del equipo
            eq_info = await session.execute(text("SELECT nombre, capitan_nombre, capitan_telefono, capitan_email, logo_url FROM torneos.equipos WHERE id = :eid"), {"eid": eid})
            eq = eq_info.fetchone()
            if not eq: continue

            nuevo_equipo_id = str(uuid.uuid4())
            await session.execute(text("""
                INSERT INTO torneos.equipos (id, torneo_id, nombre, capitan_nombre, capitan_telefono, capitan_email, logo_url, estado_inscripcion)
                VALUES (:nid, :tid, :nombre, :cn, :ct, :ce, :logo, 'confirmado')
            """), {"nid": nuevo_equipo_id, "tid": nuevo_torneo_id, "nombre": eq.nombre, "cn": eq.capitan_nombre, "ct": eq.capitan_telefono, "ce": eq.capitan_email, "logo": eq.logo_url})

            # Clonar Jugadores de Buena Fe
            pl_res = await session.execute(text("""
                SELECT nombre, dni, fecha_nacimiento, numero_camiseta, posicion, foto_url, face_encoding, estado
                FROM cancha.tournament_players
                WHERE tournament_team_id = :eid AND estado IN ('habilitado', 'en_revision')
            """), {"eid": eid})
            jugadores = pl_res.fetchall()

            for pl in jugadores:
                await session.execute(text("""
                    INSERT INTO cancha.tournament_players 
                    (torneo_equipo_id, nombre, dni, fecha_nacimiento, numero_camiseta, posicion, foto_url, face_encoding, estado)
                    VALUES (:tid, :nombre, :dni, :fn, :nc, :pos, :foto, :face, :estado)
                """), {
                    "tid": nuevo_equipo_id, "nombre": pl.nombre, "dni": pl.dni, 
                    "fn": pl.fecha_nacimiento, "nc": pl.numero_camiseta, "pos": pl.posicion, 
                    "foto": pl.foto_url, "face": pl.face_encoding, "estado": pl.estado
                })

        await session.commit()
        return {
            "status": "ok", 
            "mensaje": f"Playoff Regional creado exitosamente con {len(equipos_clasificados)} equipos clasificados.",
            "nuevo_torneo_id": nuevo_torneo_id
        }

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        raise HTTPException(status_code=500, detail=str(e))

