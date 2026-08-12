import os
import uuid
from io import BytesIO
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
import io
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from database import get_session

import qrcode
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm

router = APIRouter(
    prefix="/cancha/torneos",
    tags=["Reportes"]
)

class ColumnasEquiposExport(BaseModel):
    nombre: bool = True
    escudo: bool = True
    entrenador: bool = True
    enlace: bool = False
    puntos: bool = False
    juegos: bool = False
    ganados: bool = False
    empates: bool = False
    perdido: bool = False
    golesFavor: bool = False
    golesContra: bool = False
    diferenciaGoles: bool = False
    promedioGoles: bool = False
    aprovechamiento: bool = False
    puntosExtras: bool = False
    tarjetaRoja: bool = False
    tarjetaAmarilla: bool = False
    tarjetaAzul: bool = False
    todasTarjetas: bool = False
    juegoLimpio: bool = False
    indexTechnique: bool = False

class ColumnasJugadoresExport(BaseModel):
    equipo: bool = True
    nombre: bool = True
    dni: bool = True
    camiseta: bool = True
    posicion: bool = True
    estado: bool = True
    partidosJugados: bool = True
    goles: bool = True
    amarillas: bool = True
    rojas: bool = True
    foto: bool = True

class CarnetExportConfig(BaseModel):
    titulo: str = ""
    subtitulo: str = ""
    color: str = "#0b5cd5"
    incluirEscudo: bool = True
    espacio1: str = "numero_camiseta"
    espacio2: str = "nombre_abreviado"
    espacio3: str = "posicion"
    tamano: str = "86x59"
    modo: str = "multiple"
    equipo_ids: list[str] = []

@router.post("/{torneo_id}/reportes/equipos/excel", summary="Exportar Equipos a Excel")
async def exportar_equipos_excel(torneo_id: str, columnas: ColumnasEquiposExport, session: AsyncSession = Depends(get_session)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 1. Fetch Tournament Name
        t_res = await session.execute(text("SELECT nombre FROM torneos.torneos WHERE id = :tid"), {"tid": torneo_id})
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        torneo_nombre = t_row[0]

        # 2. Construct the big query
        sql = """
            SELECT te.id as equipo_id, te.nombre as nombre_equipo, te.capitan_nombre as entrenador,
                   po.pts as puntos, po.pj as juegos, po.pg as ganados, po.pe as empates, po.pp as perdido,
                   po.gf as goles_favor, po.gc as goles_contra,
                   (po.gf - po.gc) as diferencia_goles,
                   -- Cards
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('amarilla')), 0) as amarilla,
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('roja_directa', 'roja_segunda')), 0) as roja,
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('azul')), 0) as azul,
                   COALESCE((SELECT SUM(tt.pts_fair_play) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id), 0) as pts_disciplina
            FROM torneos.equipos te
            LEFT JOIN torneos.posiciones po ON po.equipo_id = te.id AND po.torneo_id = te.torneo_id
            WHERE te.torneo_id = :tid AND te.estado_inscripcion != 'eliminado'
            ORDER BY te.nombre
        """
        res = await session.execute(text(sql), {"tid": torneo_id})
        equipos = res.fetchall()

        # Build Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos"

        # Headers
        headers = []
        if columnas.nombre: headers.append("Nombre")
        if columnas.escudo: headers.append("Escudo")
        if columnas.entrenador: headers.append("Entrenador")
        if columnas.enlace: headers.append("Enlace de Edición")
        if columnas.puntos: headers.append("Puntos")
        if columnas.juegos: headers.append("Juegos")
        if columnas.ganados: headers.append("Ganados")
        if columnas.empates: headers.append("Empates")
        if columnas.perdido: headers.append("Perdido")
        if columnas.golesFavor: headers.append("Goles a Favor")
        if columnas.golesContra: headers.append("Goles Contra")
        if columnas.diferenciaGoles: headers.append("Diferencia de Goles")
        if columnas.promedioGoles: headers.append("Promedio de Goles")
        if columnas.aprovechamiento: headers.append("Aprovechamiento (%)")
        if columnas.puntosExtras: headers.append("Puntos Extras")
        if columnas.tarjetaRoja: headers.append("Tarjeta Roja")
        if columnas.tarjetaAmarilla: headers.append("Tarjeta Amarilla")
        if columnas.tarjetaAzul: headers.append("Tarjeta Azul")
        if columnas.todasTarjetas: headers.append("Todas las tarjetas")
        if columnas.juegoLimpio: headers.append("Juego Limpio")
        if columnas.indexTechnique: headers.append("Index Technique")

        # Fill headers
        fill_hex = PatternFill("solid", fgColor="1e3a5f")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        border_side = Side(style="thin", color="2d4a6e")
        border = Border(bottom=border_side)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = fill_hex
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 22

        # Fill rows
        for row_idx, eq in enumerate(equipos, 2):
            equipo_id, nombre_equipo, entrenador = eq[0], eq[1], eq[2]
            puntos, juegos, ganados, empates, perdido = eq[3] or 0, eq[4] or 0, eq[5] or 0, eq[6] or 0, eq[7] or 0
            goles_favor, goles_contra, diferencia_goles = eq[8] or 0, eq[9] or 0, eq[10] or 0
            amarilla, roja, azul, pts_disciplina = eq[11], eq[12], eq[13], eq[14]
            todas_tarjetas = amarilla + roja + azul

            promedio_goles = round(goles_favor / juegos, 2) if juegos > 0 else 0
            aprovechamiento = round((puntos / (juegos * 3)) * 100, 2) if juegos > 0 else 0
            
            puntos_extras = 0
            index_technique = "N/A"
            juego_limpio = pts_disciplina or 0
            
            enlace = f"https://micancha.com/admin-torneo/{torneo_id}/equipos/{equipo_id}"

            row_data = []
            if columnas.nombre: row_data.append(nombre_equipo)
            if columnas.escudo: row_data.append("N/A") 
            if columnas.entrenador: row_data.append(entrenador or "N/A")
            if columnas.enlace: row_data.append(enlace)
            if columnas.puntos: row_data.append(puntos)
            if columnas.juegos: row_data.append(juegos)
            if columnas.ganados: row_data.append(ganados)
            if columnas.empates: row_data.append(empates)
            if columnas.perdido: row_data.append(perdido)
            if columnas.golesFavor: row_data.append(goles_favor)
            if columnas.golesContra: row_data.append(goles_contra)
            if columnas.diferenciaGoles: row_data.append(diferencia_goles)
            if columnas.promedioGoles: row_data.append(promedio_goles)
            if columnas.aprovechamiento: row_data.append(aprovechamiento)
            if columnas.puntosExtras: row_data.append(puntos_extras)
            if columnas.tarjetaRoja: row_data.append(roja)
            if columnas.tarjetaAmarilla: row_data.append(amarilla)
            if columnas.tarjetaAzul: row_data.append(azul)
            if columnas.todasTarjetas: row_data.append(todas_tarjetas)
            if columnas.juegoLimpio: row_data.append(juego_limpio)
            if columnas.indexTechnique: row_data.append(index_technique)

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center", horizontal="center" if isinstance(val, (int, float)) else "left")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="f1f5f9")

        # Auto width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"Reporte_Equipos_{torneo_nombre.replace(' ', '_')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"', "Access-Control-Expose-Headers": "Content-Disposition"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")

@router.post("/{torneo_id}/reportes/equipos/pdf", summary="Exportar Equipos a PDF")
async def exportar_equipos_pdf(torneo_id: str, columnas: ColumnasEquiposExport, session: AsyncSession = Depends(get_session)):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        
        # Nuevo apartado: Consultar el nombre del torneo para el título del PDF
        t_res = await session.execute(text("SELECT nombre FROM torneos.torneos WHERE id = :tid"), {"tid": torneo_id})
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        torneo_nombre = t_row[0]

        # Nuevo apartado: Consulta cruzada para recuperar la estadística y sumatorias de tarjetas
        sql = """
            SELECT te.id as equipo_id, te.nombre as nombre_equipo, te.capitan_nombre as entrenador,
                   po.pts as puntos, po.pj as juegos, po.pg as ganados, po.pe as empates, po.pp as perdido,
                   po.gf as goles_favor, po.gc as goles_contra,
                   (po.gf - po.gc) as diferencia_goles,
                   -- Cards
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('amarilla')), 0) as amarilla,
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('roja_directa', 'roja_segunda')), 0) as roja,
                   COALESCE((SELECT COUNT(*) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id AND tt.tipo IN ('azul')), 0) as azul,
                   COALESCE((SELECT SUM(tt.pts_fair_play) FROM torneos.tarjetas tt JOIN torneos.partidos tp ON tt.partido_id = tp.id WHERE tt.equipo_id = te.id AND tp.torneo_id = te.torneo_id), 0) as pts_disciplina
            FROM torneos.equipos te
            LEFT JOIN torneos.posiciones po ON po.equipo_id = te.id AND po.torneo_id = te.torneo_id
            WHERE te.torneo_id = :tid AND te.estado_inscripcion != 'eliminado'
            ORDER BY te.nombre
        """
        res = await session.execute(text(sql), {"tid": torneo_id})
        equipos = res.fetchall()

        # Nuevo apartado: Configurar ReportLab en orientación horizontal (landscape) para que quepan muchas columnas
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, fontSize=16, spaceAfter=20)
        
        elements.append(Paragraph(f"<b>Reporte de Equipos: {torneo_nombre}</b>", title_style))

        # Nuevo apartado: Condicionar dinámicamente las cabeceras según lo marcado en la UI
        headers = []
        if columnas.nombre: headers.append("Nombre")
        if columnas.escudo: headers.append("Escudo")
        if columnas.entrenador: headers.append("Entrenador")
        if columnas.enlace: headers.append("Enlace")
        if columnas.puntos: headers.append("PTS")
        if columnas.juegos: headers.append("PJ")
        if columnas.ganados: headers.append("PG")
        if columnas.empates: headers.append("PE")
        if columnas.perdido: headers.append("PP")
        if columnas.golesFavor: headers.append("GF")
        if columnas.golesContra: headers.append("GC")
        if columnas.diferenciaGoles: headers.append("DG")
        if columnas.promedioGoles: headers.append("P.GF")
        if columnas.aprovechamiento: headers.append("Aprov(%)")
        if columnas.puntosExtras: headers.append("PtsExt")
        if columnas.tarjetaRoja: headers.append("Rojas")
        if columnas.tarjetaAmarilla: headers.append("Amaril")
        if columnas.tarjetaAzul: headers.append("Azules")
        if columnas.todasTarjetas: headers.append("T.Tarj")
        if columnas.juegoLimpio: headers.append("J.Limpio")
        if columnas.indexTechnique: headers.append("IdxTech")

        # Nuevo apartado: Formatear filas
        data = [headers]
        for eq in equipos:
            equipo_id, nombre_equipo, entrenador = eq[0], eq[1], eq[2]
            puntos, juegos, ganados, empates, perdido = eq[3] or 0, eq[4] or 0, eq[5] or 0, eq[6] or 0, eq[7] or 0
            goles_favor, goles_contra, diferencia_goles = eq[8] or 0, eq[9] or 0, eq[10] or 0
            amarilla, roja, azul, pts_disciplina = eq[11], eq[12], eq[13], eq[14]
            todas_tarjetas = amarilla + roja + azul

            promedio_goles = round(goles_favor / juegos, 2) if juegos > 0 else 0
            aprovechamiento = round((puntos / (juegos * 3)) * 100, 2) if juegos > 0 else 0
            puntos_extras = 0
            index_technique = "N/A"
            juego_limpio = pts_disciplina or 0
            
            enlace = "Ver" if columnas.enlace else ""

            row_data = []
            if columnas.nombre: row_data.append(str(nombre_equipo)[:20]) # Acortar para PDF
            if columnas.escudo: row_data.append("N/A") 
            if columnas.entrenador: row_data.append(str(entrenador or "N/A")[:15])
            if columnas.enlace: row_data.append(enlace)
            if columnas.puntos: row_data.append(str(puntos))
            if columnas.juegos: row_data.append(str(juegos))
            if columnas.ganados: row_data.append(str(ganados))
            if columnas.empates: row_data.append(str(empates))
            if columnas.perdido: row_data.append(str(perdido))
            if columnas.golesFavor: row_data.append(str(goles_favor))
            if columnas.golesContra: row_data.append(str(goles_contra))
            if columnas.diferenciaGoles: row_data.append(str(diferencia_goles))
            if columnas.promedioGoles: row_data.append(str(promedio_goles))
            if columnas.aprovechamiento: row_data.append(str(aprovechamiento))
            if columnas.puntosExtras: row_data.append(str(puntos_extras))
            if columnas.tarjetaRoja: row_data.append(str(roja))
            if columnas.tarjetaAmarilla: row_data.append(str(amarilla))
            if columnas.tarjetaAzul: row_data.append(str(azul))
            if columnas.todasTarjetas: row_data.append(str(todas_tarjetas))
            if columnas.juegoLimpio: row_data.append(str(juego_limpio))
            if columnas.indexTechnique: row_data.append(str(index_technique))
            
            data.append(row_data)

        # Nuevo apartado: Aplicar estilos visuales a la tabla PDF
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"Reporte_Equipos_{torneo_nombre.replace(' ', '_')}.pdf"
        
        return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}", "Access-Control-Expose-Headers": "Content-Disposition"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")

@router.post("/{torneo_id}/reportes/jugadores/excel", summary="Exportar Jugadores a Excel")
async def exportar_jugadores_excel(torneo_id: str, columnas: ColumnasJugadoresExport, session: AsyncSession = Depends(get_session)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        t_res = await session.execute(text("SELECT nombre FROM torneos.torneos WHERE id = :tid"), {"tid": torneo_id})
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        torneo_nombre = t_row[0]

        sql = """
            SELECT 
                te.nombre as equipo,
                tp.nombre,
                tp.dni,
                tp.numero_camiseta,
                tp.posicion,
                tp.estado,
                tp.partidos_jugados,
                (SELECT COUNT(*) FROM torneos.goles tg JOIN torneos.partidos tp_part ON tg.partido_id = tp_part.id WHERE tg.player_id = tp.id AND tp_part.torneo_id = te.torneo_id AND tg.anulado = false) as goles,
                tp.amarillas_acum,
                tp.rojas_acum,
                tp.foto_url
            FROM torneos.tournament_players tp
            JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            WHERE te.torneo_id = :tid
            ORDER BY te.nombre, tp.nombre
        """
        res = await session.execute(text(sql), {"tid": torneo_id})
        jugadores = res.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Jugadores"

        headers = []
        if columnas.equipo: headers.append("Equipo")
        if columnas.nombre: headers.append("Nombre")
        if columnas.dni: headers.append("DNI")
        if columnas.camiseta: headers.append("Camiseta")
        if columnas.posicion: headers.append("Posición")
        if columnas.estado: headers.append("Estado")
        if columnas.partidosJugados: headers.append("PJ")
        if columnas.goles: headers.append("Goles")
        if columnas.amarillas: headers.append("Amarillas")
        if columnas.rojas: headers.append("Rojas")
        if columnas.foto: headers.append("Enlace Foto")

        fill_hex = PatternFill("solid", fgColor="1e3a5f")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        border_side = Side(style="thin", color="2d4a6e")
        border = Border(bottom=border_side)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = fill_hex
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 22

        for row_idx, jug in enumerate(jugadores, 2):
            equipo, nombre, dni, camiseta, posicion, estado, pj, goles, amarillas, rojas, foto = jug
            
            row_data = []
            if columnas.equipo: row_data.append(equipo)
            if columnas.nombre: row_data.append(nombre)
            if columnas.dni: row_data.append(dni)
            if columnas.camiseta: row_data.append(camiseta or "N/A")
            if columnas.posicion: row_data.append(posicion or "N/A")
            if columnas.estado: row_data.append(estado)
            if columnas.partidosJugados: row_data.append(pj or 0)
            if columnas.goles: row_data.append(goles or 0)
            if columnas.amarillas: row_data.append(amarillas or 0)
            if columnas.rojas: row_data.append(rojas or 0)
            if columnas.foto: row_data.append(foto or "N/A")

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="center", horizontal="center" if isinstance(val, (int, float)) else "left")
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="f1f5f9")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        filename = f"Reporte_Jugadores_{torneo_nombre.replace(' ', '_')}.xlsx"
        
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"', "Access-Control-Expose-Headers": "Content-Disposition"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Excel de jugadores: {str(e)}")

@router.post("/{torneo_id}/reportes/jugadores/pdf", summary="Exportar Jugadores a PDF")
async def exportar_jugadores_pdf(torneo_id: str, columnas: ColumnasJugadoresExport, session: AsyncSession = Depends(get_session)):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        
        t_res = await session.execute(text("SELECT nombre FROM torneos.torneos WHERE id = :tid"), {"tid": torneo_id})
        t_row = t_res.fetchone()
        if not t_row:
            raise HTTPException(status_code=404, detail="Torneo no encontrado")
        torneo_nombre = t_row[0]

        sql = """
            SELECT 
                te.nombre as equipo,
                tp.nombre,
                tp.dni,
                tp.numero_camiseta,
                tp.posicion,
                tp.estado,
                tp.partidos_jugados,
                (SELECT COUNT(*) FROM torneos.goles tg JOIN torneos.partidos tp_part ON tg.partido_id = tp_part.id WHERE tg.player_id = tp.id AND tp_part.torneo_id = te.torneo_id AND tg.anulado = false) as goles,
                tp.amarillas_acum,
                tp.rojas_acum,
                tp.foto_url
            FROM torneos.tournament_players tp
            JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            WHERE te.torneo_id = :tid
            ORDER BY te.nombre, tp.nombre
        """
        res = await session.execute(text(sql), {"tid": torneo_id})
        jugadores = res.fetchall()

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, fontSize=16, spaceAfter=20)
        
        elements.append(Paragraph(f"<b>Reporte de Jugadores: {torneo_nombre}</b>", title_style))

        headers = []
        if columnas.equipo: headers.append("Equipo")
        if columnas.nombre: headers.append("Nombre")
        if columnas.dni: headers.append("DNI")
        if columnas.camiseta: headers.append("Cmis")
        if columnas.posicion: headers.append("Posición")
        if columnas.estado: headers.append("Estado")
        if columnas.partidosJugados: headers.append("PJ")
        if columnas.goles: headers.append("Goles")
        if columnas.amarillas: headers.append("Amaril")
        if columnas.rojas: headers.append("Rojas")
        if columnas.foto: headers.append("Foto")

        data = [headers]
        for jug in jugadores:
            equipo, nombre, dni, camiseta, posicion, estado, pj, goles, amarillas, rojas, foto = jug
            
            row_data = []
            if columnas.equipo: row_data.append(str(equipo)[:15])
            if columnas.nombre: row_data.append(str(nombre)[:20])
            if columnas.dni: row_data.append(str(dni))
            if columnas.camiseta: row_data.append(str(camiseta or "-"))
            if columnas.posicion: row_data.append(str(posicion or "-")[:10])
            if columnas.estado: row_data.append(str(estado)[:10])
            if columnas.partidosJugados: row_data.append(str(pj or 0))
            if columnas.goles: row_data.append(str(goles or 0))
            if columnas.amarillas: row_data.append(str(amarillas or 0))
            if columnas.rojas: row_data.append(str(rojas or 0))
            if columnas.foto: row_data.append("Sí" if foto else "No")
            
            data.append(row_data)

        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(t)
        doc.build(elements)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"Reporte_Jugadores_{torneo_nombre.replace(' ', '_')}.pdf"
        
        return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}", "Access-Control-Expose-Headers": "Content-Disposition"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF de jugadores: {str(e)}")

@router.get("/{torneo_id}/partidos/{partido_id}/planilla-pdf", summary="Generar Planilla de Partido en PDF")
async def generar_planilla_pdf(torneo_id: str, partido_id: str, session: AsyncSession = Depends(get_session)):
    try:
        # 1. Obtener detalles del partido
        partido_sql = """
            SELECT p.id, p.fecha, p.hora_inicio, p.fase,
                   el.id as local_id, el.nombre as local_nombre, 
                   ev.id as visitante_id, ev.nombre as visitante_nombre,
                   t.nombre as torneo_nombre, t.deporte, t.formato
            FROM torneos.partidos p
            JOIN torneos.equipos el ON p.equipo_local_id = el.id
            JOIN torneos.equipos ev ON p.equipo_visitante_id = ev.id
            JOIN torneos.torneos t ON p.torneo_id = t.id
            WHERE p.id = :pid AND p.torneo_id = :tid
        """
        res_partido = await session.execute(text(partido_sql), {"pid": partido_id, "tid": torneo_id})
        partido = res_partido.fetchone()
        
        if not partido:
            raise HTTPException(status_code=404, detail="Partido no encontrado")

        # 2. Obtener plantillas
        plantilla_sql = """
            SELECT torneo_equipo_id, nombre, numero_camiseta, dni
            FROM cancha.tournament_players
            WHERE (torneo_equipo_id = :el_id OR torneo_equipo_id = :ev_id)
              AND estado = 'habilitado'
            ORDER BY torneo_equipo_id, numero_camiseta ASC NULLS LAST, nombre ASC
        """
        res_plantilla = await session.execute(text(plantilla_sql), {"el_id": partido.local_id, "ev_id": partido.visitante_id})
        jugadores = res_plantilla.fetchall()

        jugadores_local = [j for j in jugadores if str(j.torneo_equipo_id) == str(partido.local_id)]
        jugadores_visitante = [j for j in jugadores if str(j.torneo_equipo_id) == str(partido.visitante_id)]

        # 3. Preparar el buffer PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, fontSize=16, spaceAfter=10)
        subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], alignment=1, fontSize=12, spaceAfter=20)
        
        # 4. Generar QR Code
        # En una app real, la URL del QR apuntaría a la pantalla del veedor.
        qr_url = f"https://micancha.com/admin/torneos/{torneo_id}?tab=acta&partido={partido_id}"
        qr = qrcode.QRCode(version=1, box_size=4, border=1)
        qr.add_data(qr_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Guardar QR temporalmente en /tmp (o en el directorio actual)
        tmp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "static", "uploads", "tmp")
        os.makedirs(tmp_dir, exist_ok=True)
        qr_path = os.path.join(tmp_dir, f"qr_{uuid.uuid4().hex[:8]}.png")
        img.save(qr_path)

        # Encabezado con QR
        fecha_str = partido.fecha.isoformat() if isinstance(partido.fecha, date) else str(partido.fecha)
        elements.append(Paragraph(f"<b>PLANILLA OFICIAL DE PARTIDO</b>", title_style))
        elements.append(Paragraph(f"Torneo: {partido.torneo_nombre} ({partido.deporte})", subtitle_style))
        elements.append(Paragraph(f"<b>{partido.local_nombre}</b> vs <b>{partido.visitante_nombre}</b>", title_style))
        elements.append(Paragraph(f"Fase: {partido.fase} | Fecha: {fecha_str} | Hora: {partido.hora_inicio}", subtitle_style))
        
        elements.append(Image(qr_path, width=1.5*inch, height=1.5*inch))
        elements.append(Spacer(1, 0.2*inch))

        # 5. Tablas de Jugadores
        def crear_tabla_equipo(titulo, lista_jugadores):
            data = [[titulo, '', '', '', '', ''],
                    ['#', 'Nombre', 'DNI', 'Firma/Check', 'Goles', 'Tarjetas']]
            for i, j in enumerate(lista_jugadores):
                cam = str(j.numero_camiseta) if j.numero_camiseta else '-'
                data.append([cam, j.nombre, j.dni, '', '', ''])
            
            # Completar filas vacías para llenar la hoja (mínimo 12 filas)
            for _ in range(max(0, 12 - len(lista_jugadores))):
                data.append(['', '', '', '', '', ''])

            t = Table(data, colWidths=[30, 180, 70, 80, 60, 60])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('SPAN', (0, 0), (-1, 0)),
                
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e2e8f0')),
                ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                ('ALIGN', (0, 2), (0, -1), 'CENTER'), # Camiseta center
                ('ALIGN', (2, 2), (2, -1), 'CENTER'), # DNI center
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            return t

        elements.append(crear_tabla_equipo(f"LOCAL: {partido.local_nombre}", jugadores_local))
        elements.append(Spacer(1, 0.3*inch))
        elements.append(crear_tabla_equipo(f"VISITANTE: {partido.visitante_nombre}", jugadores_visitante))
        
        elements.append(Spacer(1, 0.5*inch))
        
        # Firmas
        firmas_data = [['Firma Capitán Local', 'Firma Capitán Visitante', 'Firma Árbitro/Veedor'],
                       ['', '', '']]
        t_firmas = Table(firmas_data, colWidths=[150, 150, 150], rowHeights=[20, 60])
        t_firmas.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ]))
        elements.append(t_firmas)

        doc.build(elements)
        
        # Limpiar QR
        if os.path.exists(qr_path):
            os.remove(qr_path)

        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"Planilla_{partido.torneo_nombre}_{partido.local_nombre}_vs_{partido.visitante_nombre}.pdf"
        # Limpiar filename de caracteres raros
        filename = filename.replace(' ', '_').replace('/', '-')
        
        return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")

@router.post("/{torneo_id}/reportes/carnets/pdf", summary="Exportar Carnets a PDF")
async def exportar_carnets_pdf(torneo_id: str, config: CarnetExportConfig, session: AsyncSession = Depends(get_session)):
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        
        where_clause = "WHERE te.torneo_id::text = :tid"
        if config.equipo_ids and len(config.equipo_ids) > 0:
            clean_eids = [f"'{eid}'" for eid in config.equipo_ids if eid]
            if clean_eids:
                equipo_ids_str = ",".join(clean_eids)
                where_clause += f" AND te.id::text IN ({equipo_ids_str})"
            
        sql = f"""
            SELECT te.id as equipo_id, te.nombre as nombre_equipo, te.logo_url as escudo_equipo,
                   tp.nombre as nombre_jugador,
                   tp.dni as dni,
                   tp.numero_camiseta as camiseta,
                   tp.posicion as posicion,
                   tp.telefono as telefono,
                   tp.foto_url as foto_jugador
            FROM torneos.tournament_players tp
            JOIN torneos.equipos te ON te.id = tp.torneo_equipo_id
            {where_clause}
            ORDER BY te.nombre, tp.nombre
        """
        res = await session.execute(text(sql), {"tid": torneo_id})
        jugadores = res.fetchall()

        if config.tamano == "86x59":
            c_width = 59 * mm
            c_height = 86 * mm
        else: # 85x54
            c_width = 54 * mm
            c_height = 85.6 * mm

        buffer = io.BytesIO()
        
        hex_color = config.color if config.color and config.color.startswith("#") else f"#{config.color or '0b5cd5'}"
        try:
            banner_color = colors.HexColor(hex_color)
        except Exception:
            banner_color = colors.HexColor("#0b5cd5")

        def draw_carnet(c, jug, x, y):
            c.setFillColor(banner_color)
            c.rect(x, y + c_height - 25*mm, c_width, 25*mm, fill=1, stroke=0)
            
            c.setStrokeColor(colors.black)
            c.setLineWidth(0.5)
            c.rect(x, y, c_width, c_height, fill=0, stroke=1)
            
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(x + c_width/2, y + c_height - 10*mm, (config.titulo or "")[:25])
            c.setFont("Helvetica", 8)
            c.drawCentredString(x + c_width/2, y + c_height - 15*mm, (config.subtitulo or "")[:30])
            
            c.setFillColor(colors.lightgrey)
            foto_w, foto_h = 25*mm, 30*mm
            foto_x = x + (c_width - foto_w)/2
            foto_y = y + c_height - 25*mm - 5*mm - foto_h
            c.rect(foto_x, foto_y, foto_w, foto_h, fill=1, stroke=1)
            
            if config.incluirEscudo:
                c.setFillColor(colors.white)
                c.rect(x + 2*mm, y + c_height - 25*mm - 12*mm, 10*mm, 10*mm, fill=1, stroke=1)
            
            def get_text(esp):
                if esp == "numero_camiseta": return f"Dorsal: {jug.camiseta or ''}"
                elif esp == "nombre_abreviado": return str(jug.nombre_jugador or '')[:20]
                elif esp == "documento": return f"DNI: {jug.dni or ''}"
                elif esp == "posicion": return f"Pos: {jug.posicion or ''}"
                elif esp == "telefono": return f"Tel: {jug.telefono or ''}"
                return ""
            
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(x + c_width/2, foto_y - 6*mm, get_text(config.espacio1))
            c.setFont("Helvetica", 8)
            c.drawCentredString(x + c_width/2, foto_y - 12*mm, get_text(config.espacio2))
            c.drawCentredString(x + c_width/2, foto_y - 18*mm, get_text(config.espacio3))

        if config.modo == "ajuste_tamano":
            c = canvas.Canvas(buffer, pagesize=(c_width, c_height))
            for jug in jugadores:
                draw_carnet(c, jug, 0, 0)
                c.showPage()
        else:
            c = canvas.Canvas(buffer, pagesize=A4)
            a4_w, a4_h = A4
            margin_x = 10 * mm
            margin_y = 10 * mm
            
            x = margin_x
            y = a4_h - margin_y - c_height
            
            for jug in jugadores:
                draw_carnet(c, jug, x, y)
                x += c_width + 5*mm
                if x + c_width > a4_w - margin_x:
                    x = margin_x
                    y -= (c_height + 5*mm)
                    if y < margin_y:
                        c.showPage()
                        x = margin_x
                        y = a4_h - margin_y - c_height
                        
            c.showPage()
            
        c.save()
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        filename = f"Carnets_{torneo_id}.pdf"
        return Response(content=pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}", "Access-Control-Expose-Headers": "Content-Disposition"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando Carnets: {str(e)}")
