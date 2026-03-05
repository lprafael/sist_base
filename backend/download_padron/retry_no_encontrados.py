"""
retry_no_encontrados.py
Lee el archivo PADRON_SAN_LORENZO_COTEJADO.xlsx, extrae los CIs con 
observación "No encontrado" en la columna P, e intenta descargarlos 
nuevamente de la API de la ANR.
Si los encuentra, los inserta en la BD y actualiza el Excel.
"""
import os
import sys
import asyncio
import logging
import time
from datetime import datetime
from collections import deque

try:
    import httpx
except ImportError:
    print("Instala httpx: pip install httpx")
    sys.exit(1)

try:
    import openpyxl
    import asyncpg
    from dotenv import load_dotenv
except ImportError as e:
    print(f"Dependencia faltante: {e} — instala con pip install openpyxl asyncpg python-dotenv")
    sys.exit(1)

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("retry_no_encontrados.log")
    ]
)
logger = logging.getLogger(__name__)

# Config
dotenv_path = os.path.join(os.path.dirname(__file__), '..', '.env')
load_dotenv(dotenv_path)

DATABASE_URL = os.getenv("DATABASE_URL", "")
DB_DSN = DATABASE_URL \
    .replace("postgresql+asyncpg://", "postgresql://") \
    .replace("postgresql+psycopg2://", "postgresql://")

if not os.path.exists('/.dockerenv') and not os.path.exists('/run/.containerenv'):
    if "@db:5432/" in DB_DSN:
        DB_DSN = DB_DSN.replace("@db:5432/", "@localhost:5434/")

BASE_URL = "https://www.anr.org.py/assets/p2026"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
    "Referer": "https://www.anr.org.py/pre-padron-2026/"
}
CONCURRENCY = 20
TIMEOUT = 15.0

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "PADRON_SAN_LORENZO_COTEJADO.xlsx")
# Columna P = índice 16 (0-based) o columna 16 en openpyxl (1-based)
COL_CI = 8          # Columna H tiene el CI (numero_ced)
COL_OBS = 16        # Columna P (observacion_sigel)


def build_url(cedula: str) -> str:
    ci = cedula.strip().lstrip("0")
    if not ci:
        return None
    # La URL usa los dígitos del CI divididos en carpetas
    parts = list(ci)
    # URL format: BASE/d1/d2/d3/d4/FULLCI.json
    if len(parts) < 4:
        return None
    path = "/".join(parts[:4])
    return f"{BASE_URL}/{path}/{ci}.json"


def read_no_encontrados(filepath: str):
    """Lee el Excel y retorna una lista de (nro_fila, cedula)."""
    logger.info(f"Leyendo archivo: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    no_encontrados = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ci = row[COL_CI - 1]      # Columna A
        obs = row[COL_OBS - 1]    # Columna P
        if ci and obs and str(obs).strip().lower() == "no encontrado":
            no_encontrados.append((i, str(ci).strip()))
    
    wb.close()
    logger.info(f"Encontrados {len(no_encontrados)} CIs con 'No encontrado'")
    return no_encontrados


async def fetch_one(client: httpx.AsyncClient, cedula: str):
    url = build_url(cedula)
    if not url:
        return cedula, None
    try:
        r = await client.get(url, timeout=TIMEOUT)
        if r.status_code == 200:
            return cedula, r.json()
        return cedula, None
    except Exception:
        return cedula, None


async def fetch_batch(cedulas: list):
    sem = asyncio.Semaphore(CONCURRENCY)
    results = {}

    async def fetch_with_sem(client, ci):
        async with sem:
            return await fetch_one(client, ci)

    async with httpx.AsyncClient(headers=HEADERS, follow_redirects=True) as client:
        tasks = [fetch_with_sem(client, ci) for ci in cedulas]
        for coro in asyncio.as_completed(tasks):
            ci, data = await coro
            results[ci] = data
    return results


async def save_found(conn, found_records: list):
    """Inserta los registros encontrados en la BD."""
    if not found_records:
        return 0
    
    # Verificar qué columnas tiene la tabla
    cols_sql = """
        SELECT column_name FROM information_schema.columns
        WHERE table_schema = 'electoral' AND table_name = 'anr_padron_2026'
    """
    cols = await conn.fetch(cols_sql)
    col_names = [r['column_name'] for r in cols]
    logger.info(f"Columnas de la tabla: {col_names}")

    inserted = 0
    for ci, data in found_records:
        try:
            # Mapear campos del JSON a la tabla (ajustar según estructura real de la API)
            record = {
                "cedula": str(ci),
                "nombres": data.get("nombres") or data.get("nombre") or "",
                "apellidos": data.get("apellidos") or data.get("apellido") or "",
                "nacimiento": data.get("fechaNacimiento") or data.get("fecha_nac") or None,
                "departamento": data.get("departamento") or None,
                "distrito": data.get("distrito") or None,
                "seccional": data.get("seccional") or None,
                "local": data.get("local") or None,
                "mesa": data.get("mesa") or None,
                "orden": data.get("orden") or None,
            }
            
            await conn.execute("""
                INSERT INTO electoral.anr_padron_2026
                    (cedula, nombres, apellidos, nacimiento, departamento, distrito, seccional, local, mesa, orden)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                ON CONFLICT (cedula) DO NOTHING
            """,
                record["cedula"], record["nombres"], record["apellidos"],
                record["nacimiento"], record["departamento"], record["distrito"],
                record["seccional"], record["local"], record["mesa"], record["orden"]
            )
            inserted += 1
        except Exception as e:
            logger.warning(f"Error insertando CI {ci}: {e}")
    
    return inserted


def update_excel(filepath: str, row_updates: dict):
    """Abre el Excel en modo escritura y actualiza la columna P de las filas indicadas."""
    logger.info("Actualizando el Excel con los nuevos datos...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for row_num, obs_value in row_updates.items():
        ws.cell(row=row_num, column=COL_OBS, value=obs_value)
    wb.save(filepath)
    wb.close()
    logger.info("Excel actualizado.")


async def main():
    # Paso 1: Leer CIs "No encontrado"
    no_enc = read_no_encontrados(EXCEL_FILE)
    if not no_enc:
        logger.info("No hay CIs con 'No encontrado'. ¡Nada que procesar!")
        return

    cedulas = [ci for _, ci in no_enc]
    row_map = {ci: row for row, ci in no_enc}

    logger.info(f"Procesando {len(cedulas)} CIs en lotes de {CONCURRENCY}...")

    # Paso 2: Descargar desde la API
    start = time.time()
    results = await fetch_batch(cedulas)
    elapsed = time.time() - start
    logger.info(f"Descarga completada en {elapsed:.1f}s")

    found = [(ci, data) for ci, data in results.items() if data]
    not_found = [ci for ci, data in results.items() if not data]

    logger.info(f"Encontrados ahora: {len(found)} | Siguen sin encontrar: {len(not_found)}")

    # Paso 3: Guardar en BD
    if found:
        conn = await asyncpg.connect(DB_DSN)
        try:
            inserted = await save_found(conn, found)
            logger.info(f"Insertados en BD: {inserted}")
        finally:
            await conn.close()

    # Paso 4: Actualizar el Excel
    row_updates = {}
    for ci, data in results.items():
        row = row_map[ci]
        if data:
            row_updates[row] = "Encontrado (reintento)"
        else:
            row_updates[row] = "No encontrado (reintento)"

    update_excel(EXCEL_FILE, row_updates)

    logger.info("=== RESUMEN ===")
    logger.info(f"  Total 'No encontrado' procesados : {len(cedulas)}")
    logger.info(f"  Encontrados ahora                : {len(found)}")
    logger.info(f"  Siguen sin encontrarse           : {len(not_found)}")


if __name__ == "__main__":
    asyncio.run(main())
